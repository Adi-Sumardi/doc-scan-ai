"""
Enhanced Excel Template Generator
Professional Excel templates with advanced styling, formulas, and formatting
"""

import logging
from datetime import datetime, timezone
from typing import Dict, Any, List
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

try:
    from .exporters.faktur_pajak_exporter import FakturPajakExporter
except ImportError:  # Fallback when imported as top-level module
    from exporters.faktur_pajak_exporter import FakturPajakExporter  # type: ignore

logger = logging.getLogger(__name__)


class EnhancedExcelTemplate:
    """
    Create professional Excel templates with advanced styling
    """
    
    # Color scheme
    HEADER_COLOR = "1F4E78"  # Dark Blue
    SUBHEADER_COLOR = "4472C4"  # Medium Blue
    ALT_ROW_COLOR = "F2F2F2"  # Light Gray
    TOTAL_COLOR = "FFC000"  # Orange
    SUCCESS_COLOR = "70AD47"  # Green
    WARNING_COLOR = "FFC000"  # Orange
    ERROR_COLOR = "C00000"  # Red
    
    def __init__(self):
        self.wb = None
        self.ws = None
    
    def create_workbook(self, title: str = "Tax Document Report") -> Workbook:
        """Create new workbook with styling"""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = title[:31]  # Excel limit 31 chars
        return self.wb
    
    def add_header(self, title: str, subtitle: str = None, logo_path: str = None):
        """
        Add professional header with logo
        
        Args:
            title: Main title
            subtitle: Optional subtitle
            logo_path: Optional path to logo image
        """
        # Merge cells for title
        self.ws.merge_cells('A1:J1')
        
        # Title
        title_cell = self.ws['A1']
        title_cell.value = title
        title_cell.font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type='solid')
        self.ws.row_dimensions[1].height = 30
        
        # Subtitle
        if subtitle:
            self.ws.merge_cells('A2:J2')
            subtitle_cell = self.ws['A2']
            subtitle_cell.value = subtitle
            subtitle_cell.font = Font(name='Arial', size=12, color='666666')
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
            self.ws.row_dimensions[2].height = 20
        
        # Logo
        if logo_path and Path(logo_path).exists():
            try:
                img = OpenpyxlImage(logo_path)
                img.width = 80
                img.height = 80
                self.ws.add_image(img, 'A1')
            except Exception as e:
                logger.warning(f"⚠️ Failed to add logo: {e}")
    
    def add_info_section(self, start_row: int, info: Dict[str, str]):
        """
        Add information section (metadata)
        
        Args:
            start_row: Starting row number
            info: Dictionary of label: value pairs
        """
        current_row = start_row
        
        for label, value in info.items():
            # Label
            label_cell = self.ws.cell(row=current_row, column=1)
            label_cell.value = label
            label_cell.font = Font(name='Arial', size=10, bold=True)
            label_cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Value
            value_cell = self.ws.cell(row=current_row, column=2)
            value_cell.value = value
            value_cell.font = Font(name='Arial', size=10)
            value_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            current_row += 1
        
        return current_row
    
    def add_table_header(self, row: int, headers: List[str], widths: List[int] = None):
        """
        Add styled table header
        
        Args:
            row: Row number for header
            headers: List of header titles
            widths: Optional list of column widths
        """
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color=self.SUBHEADER_COLOR, end_color=self.SUBHEADER_COLOR, fill_type='solid')
            
            # Border
            cell.border = Border(
                left=Side(style='thin', color='FFFFFF'),
                right=Side(style='thin', color='FFFFFF'),
                top=Side(style='thin', color='FFFFFF'),
                bottom=Side(style='thin', color='FFFFFF')
            )
            
            # Set column width
            if widths and col_idx <= len(widths):
                self.ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]
            else:
                self.ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        self.ws.row_dimensions[row].height = 25
    
    def add_table_row(self, row: int, data: List[Any], is_total: bool = False, is_alternate: bool = False):
        """
        Add data row with styling
        
        Args:
            row: Row number
            data: List of cell values
            is_total: Whether this is a total row
            is_alternate: Whether to use alternate row color
        """
        for col_idx, value in enumerate(data, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = value
            
            # Styling
            if is_total:
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.fill = PatternFill(start_color=self.TOTAL_COLOR, end_color=self.TOTAL_COLOR, fill_type='solid')
            else:
                cell.font = Font(name='Arial', size=10)
                if is_alternate:
                    cell.fill = PatternFill(start_color=self.ALT_ROW_COLOR, end_color=self.ALT_ROW_COLOR, fill_type='solid')
            
            # Alignment
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                # Format numbers
                if isinstance(value, int) and value > 1000:
                    cell.number_format = '#,##0'
                elif isinstance(value, float):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Border
            cell.border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0')
            )
    
    def add_summary_section(self, start_row: int, summary: Dict[str, Any]):
        """
        Add summary section at bottom
        
        Args:
            start_row: Starting row for summary
            summary: Dictionary of summary statistics
        """
        current_row = start_row + 2
        
        # Title
        self.ws.merge_cells(f'A{current_row}:B{current_row}')
        title_cell = self.ws[f'A{current_row}']
        title_cell.value = "📊 RINGKASAN"
        title_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type='solid')
        current_row += 1
        
        # Summary items
        for label, value in summary.items():
            label_cell = self.ws.cell(row=current_row, column=1)
            label_cell.value = label
            label_cell.font = Font(name='Arial', size=10, bold=True)
            label_cell.alignment = Alignment(horizontal='right', vertical='center')
            
            value_cell = self.ws.cell(row=current_row, column=2)
            value_cell.value = value
            value_cell.font = Font(name='Arial', size=10)
            value_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            if isinstance(value, (int, float)) and value > 1000:
                value_cell.number_format = '#,##0'
            
            current_row += 1
        
        return current_row
    
    def add_footer(self, row: int):
        """Add footer with generation timestamp"""
        footer_row = row + 2
        self.ws.merge_cells(f'A{footer_row}:J{footer_row}')
        
        footer_cell = self.ws[f'A{footer_row}']
        footer_cell.value = f"Generated on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} | Doc-Scan AI System"
        footer_cell.font = Font(name='Arial', size=9, italic=True, color='999999')
        footer_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def freeze_panes(self, row: int, col: int = 1):
        """Freeze header rows and columns"""
        freeze_cell = self.ws.cell(row=row, column=col)
        self.ws.freeze_panes = freeze_cell
    
    def auto_filter(self, start_row: int, end_row: int, start_col: int = 1, end_col: int = 10):
        """Add auto filter to table"""
        start_cell = self.ws.cell(row=start_row, column=start_col).coordinate
        end_cell = self.ws.cell(row=end_row, column=end_col).coordinate
        self.ws.auto_filter.ref = f"{start_cell}:{end_cell}"
    
    def save(self, output_path: str):
        """Save workbook to file"""
        try:
            self.wb.save(output_path)
            logger.info(f"✅ Excel file saved: {output_path}")
            return True
        except Exception as e:
            logger.error(f"❌ Failed to save Excel: {e}")
            return False


# Helper function to create batch export
def create_batch_excel_export(batch_results: List[Dict[str, Any]], output_path: str, document_type: str = "faktur_pajak") -> bool:
    """
    Create professional batch Excel export
    
    Args:
        batch_results: List of processing results
        output_path: Path to save Excel file
        document_type: Type of documents
    
    Returns:
        True if successful
    """
    try:
        template = EnhancedExcelTemplate()
        template.create_workbook(f"{document_type.upper()} Batch Report")
        
        # Header
        template.add_header(
            title=f"📄 {document_type.upper().replace('_', ' ')} - BATCH REPORT",
            subtitle=f"Batch Processing Results | {len(batch_results)} Documents"
        )
        
        # Info section
        info = {
            "Tanggal Export": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            "Jumlah Dokumen": len(batch_results),
            "Tipe Dokumen": document_type.replace('_', ' ').title()
        }
        current_row = template.add_info_section(4, info)
        current_row += 2
        
        # Table header
        if document_type == 'faktur_pajak':
            headers = [
                "Nama",
                "Tgl",
                "NPWP",
                "Nomor Faktur",
                "Alamat",
                "DPP",
                "PPN",
                "Total",
                "Invoice",
                "Nama Barang Kena Pajak / Jasa Kena Pajak"
            ]
            widths = [25, 14, 22, 22, 45, 18, 18, 18, 18, 50]
        elif document_type == 'pph21':
            headers = ["No", "Filename", "Nomor", "Masa Pajak", "Nama", "NPWP/NIK", "Bruto", "PPh", "Confidence"]
            widths = [5, 25, 20, 15, 25, 20, 15, 15, 12]
        elif document_type == 'rekening_koran':
            headers = ["No", "Filename", "Tanggal", "Uang Masuk", "Uang Keluar", "Saldo", "Confidence"]
            widths = [5, 30, 15, 15, 15, 15, 12]
        else:
            headers = ["No", "Filename", "Document Type", "Confidence", "Status"]
            widths = [5, 30, 20, 12, 15]
        
        header_row = current_row
        template.add_table_header(header_row, headers, widths)
        current_row += 1
        
        # Data rows
        total_confidence = 0
        faktur_sanitizer = FakturPajakExporter() if document_type == 'faktur_pajak' else None
        for idx, result in enumerate(batch_results, start=1):
            extracted_data = result.get('extracted_data', {})
            
            if document_type == 'faktur_pajak':
                # PRIORITY 1: Use Smart Mapper data if available (already clean from GPT-4o)
                smart_mapped = extracted_data.get('smart_mapped', {})
                if smart_mapped and faktur_sanitizer:
                    logger.info(f"✅ Row {idx}: Using Smart Mapper GPT data (CLEAN, NO POST-PROCESSING)")
                    cleaned = faktur_sanitizer._convert_smart_mapped_to_structured(smart_mapped)
                else:
                    # FALLBACK: Use legacy structured_data with post-processing
                    logger.info(f"⚠️ Row {idx}: Smart Mapper not available, using legacy parser")
                    structured_data = extracted_data.get('structured_data', {})
                    cleaned = faktur_sanitizer._prepare_structured_fields(
                        structured_data,
                        extracted_data.get('raw_text', '')
                    ) if faktur_sanitizer else structured_data
                
                row_data = [
                    cleaned.get('nama') or 'N/A',
                    cleaned.get('tanggal') or 'N/A',
                    cleaned.get('npwp') or 'N/A',
                    cleaned.get('nomor_faktur') or 'N/A',
                    cleaned.get('alamat') or 'N/A',
                    cleaned.get('dpp') or 'N/A',
                    cleaned.get('ppn') or 'N/A',
                    cleaned.get('total') or 'N/A',
                    cleaned.get('invoice') or 'N/A',
                    cleaned.get('nama_barang_jasa') or 'N/A'
                ]
            elif document_type == 'pph21':
                row_data = [
                    idx,
                    result.get('filename', 'N/A'),
                    extracted_data.get('nomor', 'N/A'),
                    extracted_data.get('masa_pajak', 'N/A'),
                    extracted_data.get('identitas_penerima_penghasilan', {}).get('nama', 'N/A'),
                    extracted_data.get('identitas_penerima_penghasilan', {}).get('npwp_nik', 'N/A'),
                    extracted_data.get('penghasilan_bruto', 0),
                    extracted_data.get('pph', 0),
                    f"{result.get('confidence', 0):.1%}"
                ]
            elif document_type == 'rekening_koran':
                row_data = [
                    idx,
                    result.get('filename', 'N/A'),
                    extracted_data.get('tanggal', 'N/A'),
                    extracted_data.get('nilai_uang_masuk', 0),
                    extracted_data.get('nilai_uang_keluar', 0),
                    extracted_data.get('saldo', 0),
                    f"{result.get('confidence', 0):.1%}"
                ]
            else:
                row_data = [
                    idx,
                    result.get('filename', 'N/A'),
                    result.get('document_type', 'N/A'),
                    f"{result.get('confidence', 0):.1%}",
                    "✓ Processed"
                ]
            
            template.add_table_row(current_row, row_data, is_alternate=(idx % 2 == 0))
            total_confidence += result.get('confidence', 0)
            current_row += 1
        
        # Summary - REMOVED per user request
        # summary = {
        #     "Total Dokumen": len(batch_results),
        #     "Rata-rata Confidence": f"{(total_confidence / len(batch_results)):.1%}" if batch_results else "0%",
        #     "Status": "✓ Berhasil Diproses"
        # }
        # current_row = template.add_summary_section(current_row, summary)
        
        # Footer
        template.add_footer(current_row)
        
        # Freeze panes and auto filter
        template.freeze_panes(header_row + 1, 1)
        template.auto_filter(header_row, current_row - 1, 1, len(headers))
        
        # Save
        return template.save(output_path)
        
    except Exception as e:
        logger.error(f"❌ Failed to create batch Excel export: {e}")
        return False
