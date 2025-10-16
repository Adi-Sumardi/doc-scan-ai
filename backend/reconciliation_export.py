"""
Excel Export untuk Reconciliation Results
Generate Excel report untuk hasil rekonsiliasi
"""

from typing import List
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from pathlib import Path

from database import (
    ReconciliationProject,
    TaxInvoice,
    BankTransaction,
    ReconciliationMatch
)


class ReconciliationExporter:
    """
    Excel export untuk reconciliation results
    """

    # Colors
    COLOR_HEADER = "4472C4"
    COLOR_MATCHED = "C6EFCE"
    COLOR_UNMATCHED = "FFC7CE"
    COLOR_HIGH_CONFIDENCE = "D9E1F2"
    COLOR_MEDIUM_CONFIDENCE = "FFF2CC"
    COLOR_LOW_CONFIDENCE = "FCE4D6"

    def __init__(self, db: Session):
        self.db = db

    def export_project(
        self,
        project_id: str,
        output_path: str
    ) -> str:
        """
        Export complete reconciliation project to Excel

        Returns:
            Path ke file Excel yang di-generate
        """
        # Get project
        project = self.db.query(ReconciliationProject).filter(
            ReconciliationProject.id == project_id
        ).first()

        if not project:
            raise ValueError(f"Project not found: {project_id}")

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Add sheets
        self._create_summary_sheet(wb, project)
        self._create_matches_sheet(wb, project)
        self._create_unmatched_invoices_sheet(wb, project)
        self._create_unmatched_transactions_sheet(wb, project)
        self._create_all_invoices_sheet(wb, project)
        self._create_all_transactions_sheet(wb, project)

        # Save workbook
        output_file = Path(output_path)
        wb.save(output_file)

        return str(output_file)

    def _create_summary_sheet(self, wb, project: ReconciliationProject):
        """
        Create summary sheet dengan project overview
        """
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "LAPORAN REKONSILIASI PAJAK"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Project info
        row = 3
        ws[f'A{row}'] = "Nama Proyek:"
        ws[f'B{row}'] = project.name
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Periode:"
        ws[f'B{row}'] = f"{project.period_start.strftime('%d/%m/%Y')} - {project.period_end.strftime('%d/%m/%Y')}"
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Status:"
        ws[f'B{row}'] = project.status.upper()
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Tanggal Generate:"
        ws[f'B{row}'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        ws[f'A{row}'].font = Font(bold=True)

        # Statistics
        row += 2
        ws[f'A{row}'] = "STATISTIK"
        ws[f'A{row}'].font = Font(size=14, bold=True)

        row += 1
        headers = ['Kategori', 'Jumlah', 'Total Amount', 'Persentase']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Data
        stats = [
            ('Total Faktur Pajak', project.total_invoices, project.total_invoice_amount, ''),
            ('Total Transaksi Bank', project.total_transactions, project.total_transaction_amount, ''),
            ('Matched', project.matched_count, '', f"{(project.matched_count / project.total_invoices * 100) if project.total_invoices > 0 else 0:.1f}%"),
            ('Unmatched Invoices', project.unmatched_invoices, '', f"{(project.unmatched_invoices / project.total_invoices * 100) if project.total_invoices > 0 else 0:.1f}%"),
            ('Unmatched Transactions', project.unmatched_transactions, '', f"{(project.unmatched_transactions / project.total_transactions * 100) if project.total_transactions > 0 else 0:.1f}%"),
        ]

        for stat in stats:
            row += 1
            ws[f'A{row}'] = stat[0]
            ws[f'B{row}'] = stat[1]
            ws[f'C{row}'] = f"Rp {stat[2]:,.2f}" if stat[2] else '-'
            ws[f'D{row}'] = stat[3]

        # Variance
        row += 2
        ws[f'A{row}'] = "VARIANCE"
        ws[f'A{row}'].font = Font(size=14, bold=True)

        row += 1
        ws[f'A{row}'] = "Selisih Amount:"
        ws[f'B{row}'] = f"Rp {project.variance_amount:,.2f}"
        ws[f'A{row}'].font = Font(bold=True)

        if project.variance_amount > 0:
            ws[f'B{row}'].fill = PatternFill(start_color=self.COLOR_UNMATCHED, fill_type="solid")

        # Auto-fit columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _create_matches_sheet(self, wb, project: ReconciliationProject):
        """
        Create matches sheet dengan matched pairs
        """
        ws = wb.create_sheet("Matched Pairs")

        # Headers
        headers = [
            'No Faktur',
            'Tanggal Faktur',
            'Vendor',
            'Amount Faktur',
            'Tanggal Transaksi',
            'Deskripsi',
            'Amount Transaksi',
            'Variance',
            'Confidence',
            'Type',
            'Status'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Get matches
        matches = self.db.query(ReconciliationMatch).filter(
            ReconciliationMatch.project_id == project.id,
            ReconciliationMatch.status == 'active'
        ).order_by(ReconciliationMatch.match_confidence.desc()).all()

        # Data
        row = 2
        for match in matches:
            # Get invoice and transaction
            invoice = self.db.query(TaxInvoice).filter(
                TaxInvoice.id == match.invoice_id
            ).first()

            transaction = self.db.query(BankTransaction).filter(
                BankTransaction.id == match.transaction_id
            ).first()

            if not invoice or not transaction:
                continue

            # Write data
            ws[f'A{row}'] = invoice.invoice_number
            ws[f'B{row}'] = invoice.invoice_date.strftime('%d/%m/%Y')
            ws[f'C{row}'] = invoice.vendor_name or '-'
            ws[f'D{row}'] = invoice.total_amount
            ws[f'D{row}'].number_format = '#,##0.00'

            ws[f'E{row}'] = transaction.transaction_date.strftime('%d/%m/%Y')
            ws[f'F{row}'] = transaction.description or '-'
            trans_amount = transaction.credit if transaction.credit > 0 else transaction.debit
            ws[f'G{row}'] = trans_amount
            ws[f'G{row}'].number_format = '#,##0.00'

            ws[f'H{row}'] = match.amount_variance
            ws[f'H{row}'].number_format = '#,##0.00'

            ws[f'I{row}'] = f"{match.match_confidence:.1%}"
            ws[f'J{row}'] = match.match_type.upper()
            ws[f'K{row}'] = "CONFIRMED" if match.confirmed else "PENDING"

            # Color coding by confidence
            if match.match_confidence >= 0.90:
                fill_color = self.COLOR_HIGH_CONFIDENCE
            elif match.match_confidence >= 0.70:
                fill_color = self.COLOR_MEDIUM_CONFIDENCE
            else:
                fill_color = self.COLOR_LOW_CONFIDENCE

            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill_color,
                    fill_type="solid"
                )

            row += 1

        # Auto-fit columns
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_unmatched_invoices_sheet(self, wb, project: ReconciliationProject):
        """
        Create unmatched invoices sheet
        """
        ws = wb.create_sheet("Unmatched Invoices")

        # Headers
        headers = [
            'No Faktur',
            'Tanggal',
            'Vendor',
            'NPWP',
            'DPP',
            'PPN',
            'Total',
            'Notes'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Get unmatched invoices
        invoices = self.db.query(TaxInvoice).filter(
            TaxInvoice.project_id == project.id,
            TaxInvoice.match_status == 'unmatched'
        ).order_by(TaxInvoice.invoice_date.desc()).all()

        # Data
        row = 2
        for invoice in invoices:
            ws[f'A{row}'] = invoice.invoice_number
            ws[f'B{row}'] = invoice.invoice_date.strftime('%d/%m/%Y')
            ws[f'C{row}'] = invoice.vendor_name or '-'
            ws[f'D{row}'] = invoice.vendor_npwp or '-'
            ws[f'E{row}'] = invoice.dpp
            ws[f'E{row}'].number_format = '#,##0.00'
            ws[f'F{row}'] = invoice.ppn
            ws[f'F{row}'].number_format = '#,##0.00'
            ws[f'G{row}'] = invoice.total_amount
            ws[f'G{row}'].number_format = '#,##0.00'
            ws[f'H{row}'] = invoice.notes or '-'

            # Highlight unmatched
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=self.COLOR_UNMATCHED,
                    fill_type="solid"
                )

            row += 1

        # Auto-fit columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_unmatched_transactions_sheet(self, wb, project: ReconciliationProject):
        """
        Create unmatched transactions sheet
        """
        ws = wb.create_sheet("Unmatched Transactions")

        # Headers
        headers = [
            'Tanggal',
            'Bank',
            'Deskripsi',
            'Reference',
            'Debit',
            'Credit',
            'Saldo',
            'Notes'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Get unmatched transactions
        transactions = self.db.query(BankTransaction).filter(
            BankTransaction.project_id == project.id,
            BankTransaction.match_status == 'unmatched'
        ).order_by(BankTransaction.transaction_date.desc()).all()

        # Data
        row = 2
        for trans in transactions:
            ws[f'A{row}'] = trans.transaction_date.strftime('%d/%m/%Y')
            ws[f'B{row}'] = trans.bank_name or '-'
            ws[f'C{row}'] = trans.description or '-'
            ws[f'D{row}'] = trans.reference_number or '-'
            ws[f'E{row}'] = trans.debit if trans.debit > 0 else '-'
            if trans.debit > 0:
                ws[f'E{row}'].number_format = '#,##0.00'
            ws[f'F{row}'] = trans.credit if trans.credit > 0 else '-'
            if trans.credit > 0:
                ws[f'F{row}'].number_format = '#,##0.00'
            ws[f'G{row}'] = trans.balance if trans.balance else '-'
            if trans.balance:
                ws[f'G{row}'].number_format = '#,##0.00'
            ws[f'H{row}'] = trans.notes or '-'

            # Highlight unmatched
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=self.COLOR_UNMATCHED,
                    fill_type="solid"
                )

            row += 1

        # Auto-fit columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_all_invoices_sheet(self, wb, project: ReconciliationProject):
        """
        Create all invoices sheet (complete data)
        """
        ws = wb.create_sheet("All Invoices")

        # Headers
        headers = [
            'No Faktur',
            'Tanggal',
            'Tipe',
            'Vendor',
            'NPWP',
            'DPP',
            'PPN',
            'Total',
            'Match Status',
            'Confidence',
            'Notes'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Get all invoices
        invoices = self.db.query(TaxInvoice).filter(
            TaxInvoice.project_id == project.id
        ).order_by(TaxInvoice.invoice_date.desc()).all()

        # Data
        row = 2
        for invoice in invoices:
            ws[f'A{row}'] = invoice.invoice_number
            ws[f'B{row}'] = invoice.invoice_date.strftime('%d/%m/%Y')
            ws[f'C{row}'] = invoice.invoice_type or '-'
            ws[f'D{row}'] = invoice.vendor_name or '-'
            ws[f'E{row}'] = invoice.vendor_npwp or '-'
            ws[f'F{row}'] = invoice.dpp
            ws[f'F{row}'].number_format = '#,##0.00'
            ws[f'G{row}'] = invoice.ppn
            ws[f'G{row}'].number_format = '#,##0.00'
            ws[f'H{row}'] = invoice.total_amount
            ws[f'H{row}'].number_format = '#,##0.00'
            ws[f'I{row}'] = invoice.match_status.upper()
            ws[f'J{row}'] = f"{invoice.match_confidence:.1%}" if invoice.match_confidence > 0 else '-'
            ws[f'K{row}'] = invoice.notes or '-'

            # Color coding by match status
            if invoice.match_status in ['auto_matched', 'manual_matched']:
                fill_color = self.COLOR_MATCHED
            else:
                fill_color = self.COLOR_UNMATCHED

            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill_color,
                    fill_type="solid"
                )

            row += 1

        # Auto-fit columns
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_all_transactions_sheet(self, wb, project: ReconciliationProject):
        """
        Create all transactions sheet (complete data)
        """
        ws = wb.create_sheet("All Transactions")

        # Headers
        headers = [
            'Tanggal',
            'Bank',
            'Account',
            'Deskripsi',
            'Tipe',
            'Reference',
            'Debit',
            'Credit',
            'Saldo',
            'Match Status',
            'Confidence',
            'Notes'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Get all transactions
        transactions = self.db.query(BankTransaction).filter(
            BankTransaction.project_id == project.id
        ).order_by(BankTransaction.transaction_date.desc()).all()

        # Data
        row = 2
        for trans in transactions:
            ws[f'A{row}'] = trans.transaction_date.strftime('%d/%m/%Y')
            ws[f'B{row}'] = trans.bank_name or '-'
            ws[f'C{row}'] = trans.account_number or '-'
            ws[f'D{row}'] = trans.description or '-'
            ws[f'E{row}'] = trans.transaction_type or '-'
            ws[f'F{row}'] = trans.reference_number or '-'
            ws[f'G{row}'] = trans.debit if trans.debit > 0 else '-'
            if trans.debit > 0:
                ws[f'G{row}'].number_format = '#,##0.00'
            ws[f'H{row}'] = trans.credit if trans.credit > 0 else '-'
            if trans.credit > 0:
                ws[f'H{row}'].number_format = '#,##0.00'
            ws[f'I{row}'] = trans.balance if trans.balance else '-'
            if trans.balance:
                ws[f'I{row}'].number_format = '#,##0.00'
            ws[f'J{row}'] = trans.match_status.upper()
            ws[f'K{row}'] = f"{trans.match_confidence:.1%}" if trans.match_confidence > 0 else '-'
            ws[f'L{row}'] = trans.notes or '-'

            # Color coding by match status
            if trans.match_status in ['auto_matched', 'manual_matched']:
                fill_color = self.COLOR_MATCHED
            else:
                fill_color = self.COLOR_UNMATCHED

            for col in range(1, 13):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill_color,
                    fill_type="solid"
                )

            row += 1

        # Auto-fit columns
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 20
