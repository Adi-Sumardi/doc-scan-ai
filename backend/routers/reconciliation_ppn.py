"""
PPN (Tax) Reconciliation API Router
Comprehensive reconciliation for Indonesian tax documents:
- Point A: Faktur Pajak Keluaran (Output Tax)
- Point B: Faktur Pajak Masukan (Input Tax)
- Point C: Bukti Potong Lawan Transaksi (Withholding Tax)
- Point E: Rekening Koran (Bank Statement)
"""

from fastapi import APIRouter, HTTPException, Depends, Body, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import List, Optional, Dict, Any
from pydantic import BaseModel, Field
from datetime import datetime
import uuid
import logging
import io
import os
import tempfile
import shutil
from sqlalchemy.orm import Session
from database import get_db, PPNProject, ReconciliationSession, ReconciliationMessage, User
from auth import get_current_user
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from services.ppn_reconciliation_service import (
    run_ppn_reconciliation, detect_file_type, extract_company_npwp
)
from services.ppn_ai_reconciliation_service import run_ppn_ai_reconciliation
from excel_reader_service import excel_reader_service
import pandas as pd
import json
from openai import OpenAI

logger = logging.getLogger(__name__)

# Directory for temporary file storage
TEMP_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "temp_uploads")
os.makedirs(TEMP_DIR, exist_ok=True)

router = APIRouter(
    prefix="/api/reconciliation-ppn",
    tags=["reconciliation-ppn"]
)


# ==================== Pydantic Models ====================

class ProjectCreate(BaseModel):
    """Project creation request"""
    name: str = Field(..., min_length=1, max_length=255)
    periode_start: str = Field(..., description="ISO date format: YYYY-MM-DD")
    periode_end: str = Field(..., description="ISO date format: YYYY-MM-DD")
    company_npwp: str = Field(..., min_length=15, max_length=20, description="Company NPWP for auto-split logic")

    class Config:
        json_schema_extra = {
            "example": {
                "name": "PPN Reconciliation Q4 2024",
                "periode_start": "2024-10-01",
                "periode_end": "2024-12-31",
                "company_npwp": "01.234.567.8-901.000"
            }
        }


class ProjectResponse(BaseModel):
    """Project response"""
    id: str
    name: str
    periode_start: str
    periode_end: str
    company_npwp: str
    status: str  # draft, in_progress, completed
    created_at: str
    updated_at: str
    point_a_count: Optional[int] = 0
    point_b_count: Optional[int] = 0
    point_c_count: Optional[int] = 0
    point_e_count: Optional[int] = 0


class DataSourceSelection(BaseModel):
    """Data source selection for reconciliation"""
    point_a_b_source: Optional[Dict[str, Any]] = Field(
        None,
        description="Source for Point A & B (Faktur Pajak). Type: 'scanned' or 'upload'"
    )
    point_c_source: Optional[Dict[str, Any]] = Field(
        None,
        description="Source for Point C (Bukti Potong). Type: 'scanned' or 'upload'"
    )
    point_e_source: Optional[Dict[str, Any]] = Field(
        None,
        description="Source for Point E (Rekening Koran). Type: 'scanned' or 'upload'"
    )

    class Config:
        json_schema_extra = {
            "example": {
                "point_a_b_source": {
                    "type": "scanned",
                    "file_id": "uuid-here"
                },
                "point_c_source": {
                    "type": "upload",
                    "file_path": "/tmp/bukti_potong.xlsx"
                },
                "point_e_source": {
                    "type": "scanned",
                    "file_id": "uuid-here"
                }
            }
        }


class ReconciliationRequest(BaseModel):
    """Request to run reconciliation"""
    project_id: str
    data_sources: DataSourceSelection


class ReconciliationResult(BaseModel):
    """Reconciliation result summary"""
    project_id: str
    status: str  # processing, completed, failed
    point_a_count: int
    point_b_count: int
    point_c_count: int
    point_e_count: int
    matches: Dict[str, Any]
    mismatches: Dict[str, Any]
    summary: Dict[str, Any]
    created_at: str


# ==================== API Endpoints ====================

@router.post("/projects", response_model=ProjectResponse)
async def create_project(
    project: ProjectCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Create a new PPN reconciliation project

    Args:
        project: Project creation data with name, period, and company NPWP

    Returns:
        Created project with generated ID and timestamps
    """
    try:
        # Validate dates
        try:
            start_date = datetime.fromisoformat(project.periode_start)
            end_date = datetime.fromisoformat(project.periode_end)

            if start_date > end_date:
                raise HTTPException(
                    status_code=400,
                    detail="periode_end must be after periode_start"
                )
        except ValueError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid date format: {str(e)}"
            )

        # Validate NPWP format (15 digits with optional separators)
        npwp_clean = project.company_npwp.replace('.', '').replace('-', '')
        if not npwp_clean.isdigit() or len(npwp_clean) != 15:
            raise HTTPException(
                status_code=400,
                detail="NPWP must be 15 digits"
            )

        # Create project in database
        db_project = PPNProject(
            id=str(uuid.uuid4()),
            user_id=current_user.id,
            name=project.name,
            periode_start=start_date,
            periode_end=end_date,
            company_npwp=project.company_npwp,
            status="draft",
            point_a_count=0,
            point_b_count=0,
            point_c_count=0,
            point_e_count=0
        )

        db.add(db_project)
        db.commit()
        db.refresh(db_project)

        logger.info(f"Created PPN reconciliation project: {db_project.id} - {project.name} for user {current_user.username}")

        return ProjectResponse(
            id=db_project.id,
            name=db_project.name,
            periode_start=db_project.periode_start.isoformat(),
            periode_end=db_project.periode_end.isoformat(),
            company_npwp=db_project.company_npwp,
            status=db_project.status,
            created_at=db_project.created_at.isoformat(),
            updated_at=db_project.updated_at.isoformat(),
            point_a_count=db_project.point_a_count,
            point_b_count=db_project.point_b_count,
            point_c_count=db_project.point_c_count,
            point_e_count=db_project.point_e_count
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to create project: {str(e)}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create project: {str(e)}")


@router.get("/projects", response_model=List[ProjectResponse])
async def list_projects(
    status: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    List all PPN reconciliation projects for current user

    Args:
        status: Filter by status (draft, in_progress, completed)
        limit: Maximum number of results
        offset: Pagination offset

    Returns:
        List of projects
    """
    try:
        # Query projects for current user
        query = db.query(PPNProject).filter(PPNProject.user_id == current_user.id)

        # Filter by status if provided
        if status:
            query = query.filter(PPNProject.status == status)

        # Sort by created_at descending (newest first)
        query = query.order_by(PPNProject.created_at.desc())

        # Apply pagination
        projects = query.offset(offset).limit(limit).all()

        return [
            ProjectResponse(
                id=p.id,
                name=p.name,
                periode_start=p.periode_start.isoformat(),
                periode_end=p.periode_end.isoformat(),
                company_npwp=p.company_npwp,
                status=p.status,
                created_at=p.created_at.isoformat(),
                updated_at=p.updated_at.isoformat(),
                point_a_count=p.point_a_count,
                point_b_count=p.point_b_count,
                point_c_count=p.point_c_count,
                point_e_count=p.point_e_count
            )
            for p in projects
        ]

    except Exception as e:
        logger.error(f"Failed to list projects: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to list projects: {str(e)}")


@router.get("/projects/{project_id}", response_model=ProjectResponse)
async def get_project(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get a specific PPN reconciliation project by ID

    Args:
        project_id: Project UUID

    Returns:
        Project details
    """
    try:
        project = db.query(PPNProject).filter(
            PPNProject.id == project_id,
            PPNProject.user_id == current_user.id
        ).first()

        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        return ProjectResponse(
            id=project.id,
            name=project.name,
            periode_start=project.periode_start.isoformat(),
            periode_end=project.periode_end.isoformat(),
            company_npwp=project.company_npwp,
            status=project.status,
            created_at=project.created_at.isoformat(),
            updated_at=project.updated_at.isoformat(),
            point_a_count=project.point_a_count,
            point_b_count=project.point_b_count,
            point_c_count=project.point_c_count,
            point_e_count=project.point_e_count
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get project: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to get project: {str(e)}")


@router.put("/projects/{project_id}", response_model=ProjectResponse)
async def update_project(
    project_id: str,
    project_update: ProjectCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Update a PPN reconciliation project

    Args:
        project_id: Project UUID
        project_update: Updated project data

    Returns:
        Updated project
    """
    try:
        project = db.query(PPNProject).filter(
            PPNProject.id == project_id,
            PPNProject.user_id == current_user.id
        ).first()

        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        # Update project fields
        project.name = project_update.name
        project.company_npwp = project_update.company_npwp

        # Parse date strings to datetime objects for database
        project.periode_start = datetime.fromisoformat(project_update.periode_start)
        project.periode_end = datetime.fromisoformat(project_update.periode_end)

        project.updated_at = datetime.utcnow()

        db.commit()
        db.refresh(project)

        logger.info(f"Updated PPN reconciliation project: {project_id} for user {current_user.username}")

        # Return properly formatted response
        return ProjectResponse(
            id=project.id,
            name=project.name,
            periode_start=project.periode_start.isoformat(),
            periode_end=project.periode_end.isoformat(),
            company_npwp=project.company_npwp,
            status=project.status,
            created_at=project.created_at.isoformat(),
            updated_at=project.updated_at.isoformat(),
            point_a_count=project.point_a_count,
            point_b_count=project.point_b_count,
            point_c_count=project.point_c_count,
            point_e_count=project.point_e_count
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update project {project_id}: {str(e)}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update project: {str(e)}")


@router.delete("/projects/{project_id}")
async def delete_project(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Delete a PPN reconciliation project

    Args:
        project_id: Project UUID

    Returns:
        Success message
    """
    try:
        project = db.query(PPNProject).filter(
            PPNProject.id == project_id,
            PPNProject.user_id == current_user.id
        ).first()

        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        db.delete(project)
        db.commit()

        logger.info(f"Deleted PPN reconciliation project: {project_id} for user {current_user.username}")

        return {"message": "Project deleted successfully", "project_id": project_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete project: {str(e)}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete project: {str(e)}")


@router.post("/reconcile-with-files")
async def run_reconciliation_with_files(
    project_id: str = Form(...),
    faktur_pajak_file: Optional[UploadFile] = File(None),
    faktur_pajak_file_id: Optional[str] = Form(None),
    bukti_potong_file: Optional[UploadFile] = File(None),
    bukti_potong_file_id: Optional[str] = Form(None),
    rekening_koran_file: Optional[UploadFile] = File(None),
    rekening_koran_file_id: Optional[str] = Form(None),
    use_ai: bool = Form(True),  # Enable AI by default
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Run PPN reconciliation with direct file uploads or file IDs

    **NEW: AI-Enhanced Reconciliation with GPT-4o**
    - Primary: Uses GPT-4o for intelligent fuzzy matching
    - Fallback: Automatically switches to rule-based if GPT-4o fails or quota exceeded
    - Toggle: Set use_ai=False to use only rule-based matching

    This endpoint accepts:
    - Direct file uploads (multipart/form-data)
    - File IDs from previously scanned files
    - Or a combination of both
    - use_ai flag to enable/disable AI matching (default: True)
    """
    try:
        # Validate project exists
        project = db.query(PPNProject).filter(
            PPNProject.id == project_id,
            PPNProject.user_id == current_user.id
        ).first()

        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        # Update project status
        project.status = "in_progress"
        project.updated_at = datetime.utcnow()
        db.commit()

        # Helper to get file path (returns None if no file provided)
        def get_or_save_file(uploaded_file: Optional[UploadFile], file_id: Optional[str], source_name: str, required: bool = True) -> Optional[str]:
            if uploaded_file:
                # Save uploaded file temporarily
                file_extension = os.path.splitext(uploaded_file.filename)[1]
                temp_path = os.path.join(TEMP_DIR, f"{uuid.uuid4()}{file_extension}")

                with open(temp_path, "wb") as f:
                    content = uploaded_file.file.read()
                    f.write(content)

                logger.info(f"Saved uploaded file for {source_name}: {temp_path}")
                return temp_path
            elif file_id:
                # Get file from Excel exports
                excel_file = excel_reader_service.get_excel_file_by_id(file_id)
                if not excel_file:
                    raise HTTPException(status_code=404, detail=f"File not found for {source_name}")
                return excel_file['file_path']
            else:
                if required:
                    raise HTTPException(status_code=400, detail=f"Either file or file_id required for {source_name}")
                return None

        # Get file paths (only Faktur Pajak is required)
        faktur_pajak_path = get_or_save_file(faktur_pajak_file, faktur_pajak_file_id, "Faktur Pajak", required=True)
        bukti_potong_path = get_or_save_file(bukti_potong_file, bukti_potong_file_id, "Bukti Potong", required=False)
        rekening_koran_path = get_or_save_file(rekening_koran_file, rekening_koran_file_id, "Rekening Koran", required=False)

        logger.info(f"Running reconciliation with files: FP={faktur_pajak_path}, BP={bukti_potong_path}, RK={rekening_koran_path}, AI={use_ai}")

        # Run reconciliation (AI-enhanced or rule-based)
        try:
            if use_ai:
                logger.info("🤖 Using AI-Enhanced Reconciliation with GPT-4o (with fallback)")
                reconciliation_result = run_ppn_ai_reconciliation(
                    faktur_pajak_path=faktur_pajak_path,
                    bukti_potong_path=bukti_potong_path,
                    rekening_koran_path=rekening_koran_path,
                    company_npwp=project.company_npwp,
                    use_ai=True
                )
            else:
                logger.info("⚙️ Using Rule-Based Reconciliation")
                reconciliation_result = run_ppn_reconciliation(
                    faktur_pajak_path=faktur_pajak_path,
                    bukti_potong_path=bukti_potong_path,
                    rekening_koran_path=rekening_koran_path,
                    company_npwp=project.company_npwp
                )

            result = {
                "project_id": project_id,
                "status": "completed",
                "point_a_count": reconciliation_result['point_a_count'],
                "point_b_count": reconciliation_result['point_b_count'],
                "point_c_count": reconciliation_result['point_c_count'],
                "point_e_count": reconciliation_result['point_e_count'],
                "matches": reconciliation_result['matches'],
                "mismatches": reconciliation_result['mismatches'],
                "summary": reconciliation_result['summary'],
                "created_at": datetime.utcnow().isoformat()
            }

            # Update project with results
            project.status = "completed"
            project.point_a_count = reconciliation_result['point_a_count']
            project.point_b_count = reconciliation_result['point_b_count']
            project.point_c_count = reconciliation_result['point_c_count']
            project.point_e_count = reconciliation_result['point_e_count']
            project.updated_at = datetime.utcnow()
            db.commit()

            return result

        except Exception as e:
            logger.error(f"Reconciliation failed: {str(e)}", exc_info=True)
            project.status = "draft"
            db.commit()
            raise HTTPException(status_code=500, detail=f"Reconciliation failed: {str(e)}")
        finally:
            # Clean up temporary files
            for path in [faktur_pajak_path, bukti_potong_path, rekening_koran_path]:
                if path and path.startswith(TEMP_DIR) and os.path.exists(path):
                    try:
                        os.remove(path)
                        logger.info(f"Cleaned up temp file: {path}")
                    except Exception as e:
                        logger.warning(f"Failed to clean up temp file {path}: {e}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Reconciliation endpoint failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/reconcile", response_model=ReconciliationResult)
async def run_reconciliation(
    request: ReconciliationRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Run PPN reconciliation for a project

    This endpoint:
    1. Loads data from selected sources (scanned files or uploads)
    2. Auto-splits Faktur Pajak into Point A and B based on company NPWP
    3. Runs reconciliation algorithms:
       - Point A vs Point C (Faktur Keluaran vs Bukti Potong)
       - Point B vs Point E (Faktur Masukan vs Rekening Koran)
    4. Returns matched and unmatched results

    Args:
        request: Reconciliation request with project_id and data_sources

    Returns:
        Reconciliation results with matches and mismatches
    """
    try:
        # Validate project exists and belongs to user
        project = db.query(PPNProject).filter(
            PPNProject.id == request.project_id,
            PPNProject.user_id == current_user.id
        ).first()

        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        # Validate required data sources
        if not request.data_sources.point_a_b_source:
            raise HTTPException(status_code=400, detail="Point A & B source (Faktur Pajak) is required")
        if not request.data_sources.point_c_source:
            raise HTTPException(status_code=400, detail="Point C source (Bukti Potong) is required")
        if not request.data_sources.point_e_source:
            raise HTTPException(status_code=400, detail="Point E source (Rekening Koran) is required")

        # Update project status
        project.status = "in_progress"
        project.updated_at = datetime.utcnow()
        db.commit()

        # Helper function to resolve file path
        def get_file_path(source: Dict[str, Any], source_name: str) -> str:
            if source.get('type') == 'scanned' and source.get('file_id'):
                # Get file from Excel exports
                file_id = source['file_id']
                excel_file = excel_reader_service.get_excel_file_by_id(file_id)
                if not excel_file:
                    raise HTTPException(status_code=404, detail=f"File not found for {source_name}")
                return excel_file['file_path']
            elif source.get('type') == 'upload':
                # This shouldn't happen with current implementation
                # Uploaded files need to be sent via multipart/form-data
                raise HTTPException(
                    status_code=400,
                    detail=f"Direct file upload not supported yet for {source_name}. Please use scanned files."
                )
            else:
                raise HTTPException(status_code=400, detail=f"Invalid source type for {source_name}")

        # Get file paths for all three sources
        faktur_pajak_path = get_file_path(request.data_sources.point_a_b_source, "Point A & B")
        bukti_potong_path = get_file_path(request.data_sources.point_c_source, "Point C")
        rekening_koran_path = get_file_path(request.data_sources.point_e_source, "Point E")

        logger.info(f"Running reconciliation with files: FP={faktur_pajak_path}, BP={bukti_potong_path}, RK={rekening_koran_path}")

        # Run actual reconciliation
        try:
            reconciliation_result = run_ppn_reconciliation(
                faktur_pajak_path=faktur_pajak_path,
                bukti_potong_path=bukti_potong_path,
                rekening_koran_path=rekening_koran_path,
                company_npwp=project.company_npwp
            )

            result = ReconciliationResult(
                project_id=request.project_id,
                status="completed",
                point_a_count=reconciliation_result['point_a_count'],
                point_b_count=reconciliation_result['point_b_count'],
                point_c_count=reconciliation_result['point_c_count'],
                point_e_count=reconciliation_result['point_e_count'],
                matches=reconciliation_result['matches'],
                mismatches=reconciliation_result['mismatches'],
                summary=reconciliation_result['summary'],
                created_at=datetime.utcnow().isoformat()
            )

        except Exception as e:
            logger.error(f"Reconciliation failed: {str(e)}", exc_info=True)
            project.status = "draft"
            db.commit()
            raise HTTPException(status_code=500, detail=f"Reconciliation failed: {str(e)}")

        # Update project status to completed
        project.status = "completed"
        project.point_a_count = result.point_a_count
        project.point_b_count = result.point_b_count
        project.point_c_count = result.point_c_count
        project.point_e_count = result.point_e_count
        db.commit()

        logger.info(f"Completed PPN reconciliation for project: {request.project_id}")

        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to run reconciliation: {str(e)}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to run reconciliation: {str(e)}")


@router.get("/reconciliation/{project_id}/results")
async def get_reconciliation_results(project_id: str):
    """
    Get reconciliation results for a project

    Args:
        project_id: Project UUID

    Returns:
        Detailed reconciliation results with all matches and mismatches
    """
    try:
        # TODO: Fetch results from database
        # For now, return empty results

        return {
            "project_id": project_id,
            "status": "not_started",
            "message": "No reconciliation has been run for this project yet"
        }

    except Exception as e:
        logger.error(f"Failed to get reconciliation results: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to get results: {str(e)}")


@router.post("/reconciliation/{project_id}/export")
async def export_reconciliation_results(
    project_id: str,
    reconciliation_data: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export reconciliation results to Excel

    Args:
        project_id: Project UUID
        reconciliation_data: Full reconciliation result data from frontend

    Returns:
        Excel file with formatted reconciliation results
    """
    try:
        # Validate project exists and belongs to user
        project = db.query(PPNProject).filter(
            PPNProject.id == project_id,
            PPNProject.user_id == current_user.id
        ).first()

        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        # Create Excel workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["PPN Reconciliation Summary"])
        ws_summary.append([])
        ws_summary.append(["Project Name:", project.name])
        ws_summary.append(["Period:", f"{project.periode_start.strftime('%d/%m/%Y')} - {project.periode_end.strftime('%d/%m/%Y')}"])
        ws_summary.append(["Company NPWP:", project.company_npwp])
        ws_summary.append(["Generated:", datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S UTC')])
        ws_summary.append([])

        # Summary statistics
        ws_summary.append(["Reconciliation Statistics"])
        ws_summary.append(["Metric", "Value"])
        summary = reconciliation_data.get('summary', {})
        ws_summary.append(["Total Matched", summary.get('total_matched', 0)])
        ws_summary.append(["Total Unmatched", summary.get('total_unmatched', 0)])
        ws_summary.append(["Match Rate", f"{summary.get('match_rate', 0):.2f}%"])
        ws_summary.append([])

        ws_summary.append(["Point Counts"])
        ws_summary.append(["Point", "Count"])
        ws_summary.append(["Point A (Faktur Keluaran)", reconciliation_data.get('point_a_count', 0)])
        ws_summary.append(["Point B (Faktur Masukan)", reconciliation_data.get('point_b_count', 0)])
        ws_summary.append(["Point C (Bukti Potong)", reconciliation_data.get('point_c_count', 0)])
        ws_summary.append(["Point E (Rekening Koran)", reconciliation_data.get('point_e_count', 0)])

        # Style summary sheet
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40

        # Sheet 2: Point A vs C Matches (Enhanced with more details)
        matches_a_vs_c = reconciliation_data.get('matches', {}).get('point_a_vs_c', [])
        if matches_a_vs_c:
            ws_a_vs_c = wb.create_sheet("Point A vs C - Matched")
            headers = ["Nomor Faktur", "Tanggal", "Nama Pembeli", "NPWP Pembeli",
                      "Nama Barang/Jasa", "Quantity", "DPP (IDR)", "PPN (IDR)", "Total (IDR)",
                      "Match Score", "Keterangan"]
            ws_a_vs_c.append(headers)

            # Style headers
            for col_num, header in enumerate(headers, 1):
                cell = ws_a_vs_c.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Add data rows
            row_num = 2
            for match in matches_a_vs_c:
                details = match.get('details', {})
                ws_a_vs_c.append([
                    details.get('nomor_faktur', ''),
                    details.get('tanggal', ''),
                    details.get('vendor_name', ''),
                    details.get('npwp', ''),
                    details.get('nama_barang', '-'),
                    details.get('quantity', '-'),
                    details.get('dpp', 0),
                    details.get('ppn', 0),
                    details.get('amount', 0),
                    f"{match.get('match_confidence', 1.0) * 100:.1f}%",
                    details.get('keterangan', 'NPWP Match')
                ])

                # Apply currency format to DPP, PPN, Total columns (columns 7, 8, 9)
                for col_idx in [7, 8, 9]:
                    cell = ws_a_vs_c.cell(row=row_num, column=col_idx)
                    cell.number_format = 'Rp #,##0'
                row_num += 1

            # Auto-adjust column widths
            for col in ws_a_vs_c.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_a_vs_c.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Sheet 3: Point B vs E Matches (Enhanced with more details)
        matches_b_vs_e = reconciliation_data.get('matches', {}).get('point_b_vs_e', [])
        ws_b_vs_e = wb.create_sheet("Point B vs E - Matched")
        headers = ["Nomor Faktur", "Tanggal Faktur", "Nama Penjual", "NPWP Penjual",
                  "Nama Barang/Jasa", "Quantity", "DPP (IDR)", "PPN (IDR)", "Total (IDR)",
                  "Tanggal Bank", "Bank Amount (IDR)", "Selisih (IDR)", "Match Score", "Keterangan"]
        ws_b_vs_e.append(headers)

        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws_b_vs_e.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Add data rows or "No data" message
        if matches_b_vs_e:
            row_num = 2
            for match in matches_b_vs_e:
                details = match.get('details', {})
                amount = details.get('amount', 0)
                bank_amount = details.get('bank_amount', amount)
                selisih = bank_amount - amount

                ws_b_vs_e.append([
                    details.get('nomor_faktur', ''),
                    details.get('tanggal', ''),
                    details.get('vendor_name', ''),
                    details.get('npwp', ''),
                    details.get('nama_barang', '-'),
                    details.get('quantity', '-'),
                    details.get('dpp', 0),
                    details.get('ppn', 0),
                    amount,
                    details.get('bank_date', ''),
                    bank_amount,
                    selisih,
                    f"{match.get('match_confidence', 1.0) * 100:.1f}%",
                    details.get('keterangan', '')
                ])

                # Apply currency format to DPP, PPN, Total, Bank Amount, Selisih columns (columns 7, 8, 9, 11, 12)
                for col_idx in [7, 8, 9, 11, 12]:
                    cell = ws_b_vs_e.cell(row=row_num, column=col_idx)
                    cell.number_format = 'Rp #,##0'
                row_num += 1
        else:
            # Add "No data" message
            ws_b_vs_e.append(["No matched data available", "", "", "", "", "", "", "", "", "", "", "", "", ""])

        # Auto-adjust column widths
        for column in ws_b_vs_e.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_b_vs_e.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Sheet 4: Unmatched Items (Always create)
        mismatches = reconciliation_data.get('mismatches', {})
        ws_unmatched = wb.create_sheet("Unmatched Items")

        # Point A - Unmatched (Faktur Keluaran) - Enhanced with more details
        ws_unmatched.append(["Point A - Unmatched (Faktur Keluaran)"])
        ws_unmatched.merge_cells('A1:I1')
        cell = ws_unmatched['A1']
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')

        headers_a = ["Nomor Faktur", "Tanggal Faktur", "Nama Pembeli", "NPWP Pembeli",
                    "Nama Barang/Jasa", "Quantity", "DPP", "PPN", "Total"]
        ws_unmatched.append(headers_a)
        for col_num, header in enumerate(headers_a, 1):
            cell = ws_unmatched.cell(row=2, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        point_a_items = mismatches.get('point_a_unmatched', [])
        point_a_start_row = 3
        if point_a_items:
            for item in point_a_items:
                ws_unmatched.append([
                    item.get('Nomor Faktur', ''),
                    item.get('Tanggal Faktur', ''),
                    item.get('Nama Pembeli', ''),
                    item.get('NPWP Pembeli', ''),
                    item.get('Nama Barang', '-'),
                    item.get('Quantity', '-'),
                    item.get('DPP', 0),
                    item.get('PPN', 0),
                    item.get('Total', 0)
                ])

                # Apply currency format to DPP, PPN, Total columns (columns 7, 8, 9)
                for col_idx in [7, 8, 9]:
                    cell = ws_unmatched.cell(row=point_a_start_row, column=col_idx)
                    cell.number_format = 'Rp #,##0'
                point_a_start_row += 1
        else:
            ws_unmatched.append(["No unmatched data", "", "", "", "", "", "", "", ""])

        ws_unmatched.append([])

        # Point C - Unmatched (Bukti Potong)
        current_row = ws_unmatched.max_row + 1
        ws_unmatched.append(["Point C - Unmatched (Bukti Potong)"])
        ws_unmatched.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws_unmatched[f'A{current_row}']
        cell.fill = PatternFill(start_color="9966FF", end_color="9966FF", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')

        headers_c = ["Nomor Bukti Potong", "Tanggal", "Nama Pemotong", "NPWP Pemotong", "Jenis Penghasilan", "Jumlah Bruto", "PPh Dipotong"]
        ws_unmatched.append(headers_c)
        header_row = ws_unmatched.max_row
        for col_num, header in enumerate(headers_c, 1):
            cell = ws_unmatched.cell(row=header_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        point_c_items = mismatches.get('point_c_unmatched', [])
        point_c_start_row = header_row + 1
        if point_c_items:
            for item in point_c_items:
                ws_unmatched.append([
                    item.get('Nomor Bukti Potong', ''),
                    item.get('Tanggal', ''),
                    item.get('Nama Pemotong', ''),
                    item.get('NPWP Pemotong', ''),
                    item.get('Jenis Penghasilan', ''),
                    item.get('Jumlah Bruto', 0),
                    item.get('PPh Dipotong', 0)
                ])

                # Apply currency format to Jumlah Bruto and PPh Dipotong columns (columns 6, 7)
                for col_idx in [6, 7]:
                    cell = ws_unmatched.cell(row=point_c_start_row, column=col_idx)
                    cell.number_format = 'Rp #,##0'
                point_c_start_row += 1
        else:
            ws_unmatched.append(["No unmatched data", "", "", "", "", "", ""])

        ws_unmatched.append([])

        # Point B - Unmatched (Faktur Masukan) - Enhanced with more details
        current_row = ws_unmatched.max_row + 1
        ws_unmatched.append(["Point B - Unmatched (Faktur Masukan)"])
        ws_unmatched.merge_cells(f'A{current_row}:I{current_row}')
        cell = ws_unmatched[f'A{current_row}']
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')

        headers_b = ["Nomor Faktur", "Tanggal Faktur", "Nama Penjual", "NPWP Penjual",
                    "Nama Barang/Jasa", "Quantity", "DPP", "PPN", "Total"]
        ws_unmatched.append(headers_b)
        header_row = ws_unmatched.max_row
        for col_num, header in enumerate(headers_b, 1):
            cell = ws_unmatched.cell(row=header_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        point_b_items = mismatches.get('point_b_unmatched', [])
        point_b_start_row = header_row + 1
        if point_b_items:
            for item in point_b_items:
                ws_unmatched.append([
                    item.get('Nomor Faktur', ''),
                    item.get('Tanggal Faktur', ''),
                    item.get('Nama Penjual', ''),
                    item.get('NPWP Penjual', ''),
                    item.get('Nama Barang', '-'),
                    item.get('Quantity', '-'),
                    item.get('DPP', 0),
                    item.get('PPN', 0),
                    item.get('Total', 0)
                ])

                # Apply currency format to DPP, PPN, Total columns (columns 7, 8, 9)
                for col_idx in [7, 8, 9]:
                    cell = ws_unmatched.cell(row=point_b_start_row, column=col_idx)
                    cell.number_format = 'Rp #,##0'
                point_b_start_row += 1
        else:
            ws_unmatched.append(["No unmatched data", "", "", "", "", "", "", "", ""])

        ws_unmatched.append([])

        # Point E - Unmatched (Rekening Koran)
        current_row = ws_unmatched.max_row + 1
        ws_unmatched.append(["Point E - Unmatched (Rekening Koran)"])
        ws_unmatched.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws_unmatched[f'A{current_row}']
        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')

        headers_e = ["Tanggal", "Keterangan", "Cabang", "Debet", "Kredit", "Saldo"]
        ws_unmatched.append(headers_e)
        header_row = ws_unmatched.max_row
        for col_num, header in enumerate(headers_e, 1):
            cell = ws_unmatched.cell(row=header_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        point_e_items = mismatches.get('point_e_unmatched', [])
        point_e_start_row = header_row + 1
        if point_e_items:
            for item in point_e_items:
                ws_unmatched.append([
                    item.get('Tanggal', ''),
                    item.get('Keterangan', ''),
                    item.get('Cabang', ''),
                    item.get('Debet', 0),
                    item.get('Kredit', 0),
                    item.get('Saldo', 0)
                ])

                # Apply currency format to Debet, Kredit, Saldo columns (columns 4, 5, 6)
                for col_idx in [4, 5, 6]:
                    cell = ws_unmatched.cell(row=point_e_start_row, column=col_idx)
                    cell.number_format = 'Rp #,##0'
                point_e_start_row += 1
        else:
            ws_unmatched.append(["No unmatched data", "", "", "", "", ""])

        # Auto-adjust column widths for unmatched sheet
        for col_idx in range(1, 10):  # A to I columns (now 9 columns instead of 7)
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in ws_unmatched.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    # Skip merged cells
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        continue
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws_unmatched.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename (sanitize user input)
        import re as _re
        safe_name = _re.sub(r'[^a-zA-Z0-9._-]', '_', project.name[:30])
        filename = f"PPN_Reconciliation_{safe_name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

        logger.info(f"Exported reconciliation results for project: {project_id}")

        # Return file
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export reconciliation results: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export results: {str(e)}")


# ==================== Chat-Based Reconciliation Endpoints ====================

@router.post("/sessions")
async def create_session(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new reconciliation chat session"""
    session = ReconciliationSession(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        title="New Reconciliation"
    )
    db.add(session)
    db.commit()
    db.refresh(session)

    return {
        "id": session.id,
        "title": session.title,
        "created_at": session.created_at.isoformat() if session.created_at else None,
        "updated_at": session.updated_at.isoformat() if session.updated_at else None,
    }


@router.get("/sessions")
async def list_sessions(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List all reconciliation chat sessions for current user"""
    sessions = db.query(ReconciliationSession).filter(
        ReconciliationSession.user_id == current_user.id
    ).order_by(ReconciliationSession.created_at.desc()).all()

    return [
        {
            "id": s.id,
            "title": s.title,
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "updated_at": s.updated_at.isoformat() if s.updated_at else None,
        }
        for s in sessions
    ]


@router.get("/sessions/{session_id}")
async def get_session(
    session_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get session with all messages"""
    session = db.query(ReconciliationSession).filter(
        ReconciliationSession.id == session_id,
        ReconciliationSession.user_id == current_user.id
    ).first()

    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    messages = db.query(ReconciliationMessage).filter(
        ReconciliationMessage.session_id == session_id
    ).order_by(
        ReconciliationMessage.created_at.asc(),
        ReconciliationMessage.role.desc()  # 'user' before 'assistant' when timestamps match
    ).all()

    return {
        "id": session.id,
        "title": session.title,
        "created_at": session.created_at.isoformat() if session.created_at else None,
        "updated_at": session.updated_at.isoformat() if session.updated_at else None,
        "messages": [
            {
                "id": m.id,
                "role": m.role,
                "content": m.content,
                "attachments": m.attachments,
                "results": m.results,
                "created_at": m.created_at.isoformat() if m.created_at else None,
            }
            for m in messages
        ]
    }


@router.delete("/sessions/{session_id}")
async def delete_session(
    session_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a reconciliation session and its messages"""
    session = db.query(ReconciliationSession).filter(
        ReconciliationSession.id == session_id,
        ReconciliationSession.user_id == current_user.id
    ).first()

    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    # Delete messages first
    db.query(ReconciliationMessage).filter(
        ReconciliationMessage.session_id == session_id
    ).delete()

    db.delete(session)
    db.commit()

    return {"status": "deleted", "id": session_id}


# ==================== AI Chat Helper ====================

async def _chat_with_ai(prompt: str, session_id: str, db: Session) -> str:
    """Conversational AI chat about tax reconciliation using OpenAI GPT-4.1"""
    try:
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            return "AI chat tidak tersedia. OPENAI_API_KEY belum dikonfigurasi."

        client = OpenAI(api_key=api_key)

        # Load conversation history from this session
        history_messages = db.query(ReconciliationMessage).filter(
            ReconciliationMessage.session_id == session_id
        ).order_by(
            ReconciliationMessage.created_at.asc(),
            ReconciliationMessage.role.desc()
        ).all()

        # Build messages for OpenAI
        messages = [
            {
                "role": "system",
                "content": (
                    "Kamu adalah asisten ahli rekonsiliasi pajak Indonesia. "
                    "Kamu membantu pengguna dengan pertanyaan seputar:\n"
                    "- Faktur Pajak (PPN Keluaran & Masukan)\n"
                    "- Bukti Potong (PPh 21, 23, dll)\n"
                    "- Rekening Koran (Bank Statement)\n"
                    "- Rekonsiliasi pajak: mencocokkan Faktur Pajak vs Bukti Potong, Faktur vs Rekening Koran\n"
                    "- Peraturan perpajakan Indonesia terkait PPN, PPh, e-Faktur, e-Bupot\n"
                    "- Format NPWP, nomor faktur, kode transaksi\n\n"
                    "Jawab dalam Bahasa Indonesia yang profesional dan mudah dipahami. "
                    "Jika pengguna bertanya di luar topik pajak/akuntansi, tetap jawab dengan sopan "
                    "tapi arahkan kembali ke topik rekonsiliasi pajak.\n\n"
                    "Jika pengguna ingin melakukan rekonsiliasi, minta mereka upload file Excel "
                    "(Faktur Pajak, Bukti Potong, atau Rekening Koran) beserta instruksi."
                )
            }
        ]

        # Add conversation history (last 20 messages to keep context manageable)
        for msg in history_messages[-20:]:
            if msg.role == "user":
                messages.append({"role": "user", "content": msg.content or ""})
            elif msg.role == "assistant":
                # Include result summary if available, not full data
                content = msg.content or ""
                if msg.results and msg.results.get("reconciliation", {}).get("status") == "completed":
                    summary = msg.results.get("reconciliation", {}).get("summary", {})
                    content += f"\n[Hasil rekonsiliasi sebelumnya: match_rate={summary.get('match_rate', 0)}%, "
                    content += f"matched={summary.get('total_auto_matched', 0)}, "
                    content += f"unmatched={summary.get('total_unmatched', 0)}]"
                messages.append({"role": "assistant", "content": content})

        # Add current user prompt
        messages.append({"role": "user", "content": prompt})

        response = client.chat.completions.create(
            model="gpt-4.1",
            messages=messages,
            max_tokens=2000,
            temperature=0.7,
        )

        return response.choices[0].message.content or "Maaf, saya tidak bisa menjawab saat ini."

    except Exception as e:
        logger.error(f"AI chat error: {e}", exc_info=True)
        return f"Maaf, terjadi kesalahan saat memproses pertanyaan Anda. Silakan coba lagi."


@router.post("/sessions/{session_id}/chat")
async def chat_reconciliation(
    session_id: str,
    prompt: str = Form(""),
    files: List[UploadFile] = File(default=[]),
    use_ai: bool = Form(True),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Send a chat message with files + prompt, returns assistant response with reconciliation results.
    Auto-detects file types and extracts NPWP from faktur pajak.
    """
    session = db.query(ReconciliationSession).filter(
        ReconciliationSession.id == session_id,
        ReconciliationSession.user_id == current_user.id
    ).first()

    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    if not files and not prompt.strip():
        raise HTTPException(status_code=400, detail="Please attach files or type a message")

    temp_paths = []
    try:
        # Save uploaded files and detect types
        files_detected = []
        faktur_pajak_path = None
        bukti_potong_path = None
        rekening_koran_path = None

        for uploaded_file in files:
            ext = os.path.splitext(uploaded_file.filename)[1]
            temp_path = os.path.join(TEMP_DIR, f"{uuid.uuid4()}{ext}")

            with open(temp_path, "wb") as f:
                content = uploaded_file.file.read()
                f.write(content)

            temp_paths.append(temp_path)

            # Auto-detect file type from column headers
            try:
                df = pd.read_excel(temp_path)
                detected_type = detect_file_type(df)
                row_count = len(df)

                files_detected.append({
                    "name": uploaded_file.filename,
                    "detected_type": detected_type,
                    "row_count": row_count,
                    "size": len(content),
                })

                if detected_type == 'faktur_pajak':
                    faktur_pajak_path = temp_path
                elif detected_type == 'bukti_potong':
                    bukti_potong_path = temp_path
                elif detected_type == 'rekening_koran':
                    rekening_koran_path = temp_path
                else:
                    files_detected[-1]["error"] = "Tipe file tidak dikenali dari kolom Excel"

            except Exception as e:
                files_detected.append({
                    "name": uploaded_file.filename,
                    "detected_type": "error",
                    "row_count": 0,
                    "error": f"Gagal membaca file: {str(e)}",
                })

        # Save user message (set explicit timestamp so ordering is guaranteed)
        user_created_at = datetime.utcnow()
        user_msg = ReconciliationMessage(
            id=str(uuid.uuid4()),
            session_id=session_id,
            role="user",
            content=prompt,
            attachments=files_detected,
            created_at=user_created_at,
        )
        db.add(user_msg)

        # Update session title from first prompt
        existing_msgs = db.query(ReconciliationMessage).filter(
            ReconciliationMessage.session_id == session_id
        ).count()
        if existing_msgs == 0 and prompt.strip():
            session.title = prompt.strip()[:50]
        elif existing_msgs == 0 and files_detected:
            session.title = f"Rekonsiliasi {files_detected[0]['name'][:40]}"

        # Run reconciliation if we have at least faktur pajak
        assistant_content = ""
        results_data = {"files_detected": files_detected}

        if faktur_pajak_path:
            # Auto-extract NPWP
            try:
                fp_df = pd.read_excel(faktur_pajak_path)
                company_npwp = extract_company_npwp(fp_df)
            except Exception as e:
                logger.warning(f"Failed to extract NPWP: {e}")
                company_npwp = ""

            results_data["company_npwp"] = company_npwp

            if not company_npwp:
                assistant_content = "Tidak dapat mendeteksi NPWP perusahaan dari file Faktur Pajak. Pastikan kolom 'NPWP Penjual' terisi."
            else:
                logger.info(f"Chat reconciliation: NPWP={company_npwp}, AI={use_ai}")

                try:
                    if use_ai:
                        reconciliation_result = run_ppn_ai_reconciliation(
                            faktur_pajak_path=faktur_pajak_path,
                            bukti_potong_path=bukti_potong_path,
                            rekening_koran_path=rekening_koran_path,
                            company_npwp=company_npwp,
                            use_ai=True
                        )
                    else:
                        reconciliation_result = run_ppn_reconciliation(
                            faktur_pajak_path=faktur_pajak_path,
                            bukti_potong_path=bukti_potong_path,
                            rekening_koran_path=rekening_koran_path,
                            company_npwp=company_npwp
                        )

                    results_data["reconciliation"] = {
                        "status": "completed",
                        "point_a_count": reconciliation_result['point_a_count'],
                        "point_b_count": reconciliation_result['point_b_count'],
                        "point_c_count": reconciliation_result['point_c_count'],
                        "point_e_count": reconciliation_result['point_e_count'],
                        "matches": reconciliation_result['matches'],
                        "mismatches": reconciliation_result['mismatches'],
                        "summary": reconciliation_result['summary'],
                    }

                    summary = reconciliation_result['summary']
                    match_rate = summary.get('match_rate', 0)
                    total_matched = summary.get('total_auto_matched', 0) + summary.get('total_suggested', 0)
                    total_unmatched = summary.get('total_unmatched', 0)

                    assistant_content = (
                        f"Rekonsiliasi selesai. "
                        f"Match rate: {match_rate:.1f}% | "
                        f"Matched: {total_matched} | "
                        f"Unmatched: {total_unmatched}"
                    )

                except Exception as e:
                    logger.error(f"Chat reconciliation failed: {e}", exc_info=True)
                    assistant_content = f"Rekonsiliasi gagal: {str(e)}"
                    results_data["reconciliation"] = {"status": "failed", "error": str(e)}
        else:
            if files_detected:
                unknown_files = [f['name'] for f in files_detected if f['detected_type'] in ('unknown', 'error')]
                if unknown_files:
                    assistant_content = (
                        f"File yang diupload tidak terdeteksi sebagai Faktur Pajak. "
                        f"Pastikan file Excel memiliki kolom: Nomor Faktur, NPWP Penjual, NPWP Pembeli, dll."
                    )
                else:
                    assistant_content = "Tidak ada file Faktur Pajak. Upload minimal file Faktur Pajak untuk memulai rekonsiliasi."
            else:
                # No files — use OpenAI for conversational chat about reconciliation
                if prompt.strip():
                    assistant_content = await _chat_with_ai(prompt, session_id, db)
                else:
                    assistant_content = "Silakan upload file Excel untuk memulai rekonsiliasi, atau tanyakan sesuatu tentang rekonsiliasi pajak."

        # Save assistant message (timestamp after user message for correct ordering)
        assistant_msg = ReconciliationMessage(
            id=str(uuid.uuid4()),
            session_id=session_id,
            role="assistant",
            content=assistant_content,
            results=results_data,
            created_at=datetime.utcnow(),
        )
        db.add(assistant_msg)

        session.updated_at = datetime.utcnow()
        db.commit()

        return {
            "user_message": {
                "id": user_msg.id,
                "role": "user",
                "content": user_msg.content,
                "attachments": user_msg.attachments,
                "created_at": user_msg.created_at.isoformat() if user_msg.created_at else None,
            },
            "assistant_message": {
                "id": assistant_msg.id,
                "role": "assistant",
                "content": assistant_msg.content,
                "results": assistant_msg.results,
                "created_at": assistant_msg.created_at.isoformat() if assistant_msg.created_at else None,
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Chat reconciliation error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Clean up temp files
        for path in temp_paths:
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                except Exception as e:
                    logger.warning(f"Failed to clean up temp file {path}: {e}")


@router.post("/sessions/{session_id}/export")
async def export_chat_results(
    session_id: str,
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export reconciliation results from a chat message to Excel"""
    session = db.query(ReconciliationSession).filter(
        ReconciliationSession.id == session_id,
        ReconciliationSession.user_id == current_user.id
    ).first()

    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    message_id = body.get("message_id")
    if not message_id:
        raise HTTPException(status_code=400, detail="message_id is required")

    message = db.query(ReconciliationMessage).filter(
        ReconciliationMessage.id == message_id,
        ReconciliationMessage.session_id == session_id,
        ReconciliationMessage.role == "assistant"
    ).first()

    if not message or not message.results:
        raise HTTPException(status_code=404, detail="Results not found")

    reconciliation = message.results.get("reconciliation", {})
    if reconciliation.get("status") != "completed":
        raise HTTPException(status_code=400, detail="No completed reconciliation results to export")

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Summary sheet
        ws = wb.create_sheet("Summary")
        ws.append(["Reconciliation Summary"])
        ws.append([])
        ws.append(["Session:", session.title])
        ws.append(["NPWP:", message.results.get("company_npwp", "-")])
        ws.append(["Generated:", datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S UTC')])
        ws.append([])

        summary = reconciliation.get("summary", {})
        ws.append(["Metric", "Value"])
        for key, val in summary.items():
            ws.append([key.replace("_", " ").title(), str(val)])

        # Matched A vs C
        matches_ac = reconciliation.get("matches", {}).get("point_a_vs_c", [])
        if matches_ac:
            ws_ac = wb.create_sheet("Matched A vs C")
            headers = ["No", "Nomor Faktur", "Vendor", "Amount", "Match Type", "Confidence"]
            ws_ac.append(headers)
            for i, cell in enumerate(ws_ac[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            for idx, m in enumerate(matches_ac, 1):
                d = m.get("details", {})
                ws_ac.append([
                    idx, d.get("nomor_faktur", ""), d.get("vendor_name", ""),
                    d.get("amount", 0), m.get("match_type", ""), m.get("match_confidence", 0)
                ])

        # Matched B vs E
        matches_be = reconciliation.get("matches", {}).get("point_b_vs_e", [])
        if matches_be:
            ws_be = wb.create_sheet("Matched B vs E")
            headers = ["No", "Nomor Faktur", "Vendor", "Amount", "Bank Amount", "Match Type", "Confidence"]
            ws_be.append(headers)
            for i, cell in enumerate(ws_be[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            for idx, m in enumerate(matches_be, 1):
                d = m.get("details", {})
                ws_be.append([
                    idx, d.get("nomor_faktur", ""), d.get("vendor_name", ""),
                    d.get("amount", 0), d.get("bank_amount", 0),
                    m.get("match_type", ""), m.get("match_confidence", 0)
                ])

        # Unmatched sheet
        mismatches = reconciliation.get("mismatches", {})
        has_unmatched = any(len(v) > 0 for v in mismatches.values() if isinstance(v, list))
        if has_unmatched:
            ws_um = wb.create_sheet("Unmatched")
            for point_key, items in mismatches.items():
                if not items:
                    continue
                ws_um.append([point_key.replace("_", " ").title()])
                if items and isinstance(items[0], dict):
                    keys = list(items[0].keys())
                    ws_um.append(keys)
                    for item in items:
                        ws_um.append([item.get(k, "") for k in keys])
                ws_um.append([])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        import re as _re
        safe_title = _re.sub(r'[^a-zA-Z0-9._-]', '_', session.title[:30])
        filename = f"Reconciliation_{safe_title}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
