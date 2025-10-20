"""
Reconciliation Exporter
Exports reconciliation project data to Excel with multiple sheets
"""

from typing import Dict, Any, List
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

# Check for openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("⚠️ openpyxl not installed - Excel export disabled")


class ReconciliationExporter:
    """Exporter for Tax Reconciliation projects"""

    def __init__(self):
        self.document_type = "reconciliation"

    def export_project_to_excel(
        self,
        project: Any,
        invoices: List[Any],
        transactions: List[Any],
        matches: List[Any],
        output_path: str
    ) -> bool:
        """
        Export complete reconciliation project to Excel with 4 sheets:
        - Summary: Project overview and statistics
        - Invoices: All tax invoices
        - Transactions: All bank transactions
        - Matches: All matched pairs with confidence scores
        """

        if not HAS_OPENPYXL:
            logger.error("❌ openpyxl not installed - cannot export to Excel")
            return False

        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Sheet 1: Summary
            self._create_summary_sheet(wb, project)

            # Sheet 2: Invoices
            self._create_invoices_sheet(wb, invoices)

            # Sheet 3: Transactions
            self._create_transactions_sheet(wb, transactions)

            # Sheet 4: Matches
            self._create_matches_sheet(wb, matches)

            # Save workbook
            wb.save(output_path)
            logger.info(f"✅ Reconciliation report exported to: {output_path}")
            return True

        except Exception as e:
            logger.error(f"❌ Failed to export reconciliation to Excel: {e}")
            return False

    def _create_summary_sheet(self, wb: Workbook, project: Any):
        """Create summary sheet with project stats"""
        ws = wb.create_sheet("Summary", 0)

        # Header
        ws['A1'] = "Tax Reconciliation Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Project info
        row = 3
        info_data = [
            ("Project Name", project.name),
            ("Description", project.description or "-"),
            ("Period", f"{self._format_date(project.period_start)} to {self._format_date(project.period_end)}"),
            ("Status", project.status.upper()),
            ("", ""),
            ("Statistics", ""),
            ("Total Invoices", f"{project.total_invoices:,}"),
            ("Total Transactions", f"{project.total_transactions:,}"),
            ("Matched Count", f"{project.matched_count:,}"),
            ("Unmatched Invoices", f"{project.unmatched_invoices:,}"),
            ("Unmatched Transactions", f"{project.unmatched_transactions:,}"),
            ("", ""),
            ("Amounts", ""),
            ("Total Invoice Amount", self._format_rupiah(project.total_invoice_amount)),
            ("Total Transaction Amount", self._format_rupiah(project.total_transaction_amount)),
            ("Variance", self._format_rupiah(project.variance_amount)),
            ("", ""),
            ("Generated", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        ]

        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if label in ["Statistics", "Amounts", "Project Name"]:
                ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def _create_invoices_sheet(self, wb: Workbook, invoices: List[Any]):
        """Create invoices sheet"""
        ws = wb.create_sheet("Tax Invoices")

        # Headers
        headers = [
            "Invoice Number", "Date", "Type", "Vendor Name", "Vendor NPWP",
            "DPP", "PPN", "Total Amount", "Match Status", "Confidence",
            "AI Model", "Extraction Confidence"
        ]

        self._apply_header_style(ws, headers)

        # Data
        for idx, inv in enumerate(invoices, start=2):
            ws[f'A{idx}'] = inv.invoice_number or "-"
            ws[f'B{idx}'] = self._format_date(inv.invoice_date)
            ws[f'C{idx}'] = (inv.invoice_type or "").title()
            ws[f'D{idx}'] = inv.vendor_name or "-"
            ws[f'E{idx}'] = inv.vendor_npwp or "-"
            ws[f'F{idx}'] = inv.dpp or 0
            ws[f'G{idx}'] = inv.ppn or 0
            ws[f'H{idx}'] = inv.total_amount or 0
            ws[f'I{idx}'] = (inv.match_status or "unmatched").upper()
            ws[f'J{idx}'] = f"{(inv.match_confidence or 0) * 100:.1f}%" if inv.match_confidence else "-"
            ws[f'K{idx}'] = inv.ai_model_used or "gpt-4o"
            ws[f'L{idx}'] = f"{(inv.extraction_confidence or 0) * 100:.1f}%" if inv.extraction_confidence else "-"

            # Apply number format to currency columns
            for col in ['F', 'G', 'H']:
                ws[f'{col}{idx}'].number_format = '#,##0'

            # Color code match status
            self._apply_match_status_color(ws[f'I{idx}'], inv.match_status)

        # Auto-size columns
        for col_idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _create_transactions_sheet(self, wb: Workbook, transactions: List[Any]):
        """Create transactions sheet"""
        ws = wb.create_sheet("Bank Transactions")

        # Headers
        headers = [
            "Date", "Bank", "Account", "Description", "Reference",
            "Debit", "Credit", "Balance", "Match Status", "Confidence",
            "Extracted Vendor", "Extracted Invoice", "AI Model"
        ]

        self._apply_header_style(ws, headers)

        # Data
        for idx, txn in enumerate(transactions, start=2):
            ws[f'A{idx}'] = self._format_date(txn.transaction_date)
            ws[f'B{idx}'] = txn.bank_name or "-"
            ws[f'C{idx}'] = txn.account_number or "-"
            ws[f'D{idx}'] = txn.description or "-"
            ws[f'E{idx}'] = txn.reference_number or "-"
            ws[f'F{idx}'] = txn.debit or 0
            ws[f'G{idx}'] = txn.credit or 0
            ws[f'H{idx}'] = txn.balance or 0
            ws[f'I{idx}'] = (txn.match_status or "unmatched").upper()
            ws[f'J{idx}'] = f"{(txn.match_confidence or 0) * 100:.1f}%" if txn.match_confidence else "-"
            ws[f'K{idx}'] = txn.extracted_vendor_name or "-"
            ws[f'L{idx}'] = txn.extracted_invoice_number or "-"
            ws[f'M{idx}'] = txn.ai_model_used or "claude-sonnet-4"

            # Apply number format to currency columns
            for col in ['F', 'G', 'H']:
                ws[f'{col}{idx}'].number_format = '#,##0'

            # Color code match status
            self._apply_match_status_color(ws[f'I{idx}'], txn.match_status)

        # Auto-size columns
        for col_idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _create_matches_sheet(self, wb: Workbook, matches: List[Any]):
        """Create matches sheet"""
        ws = wb.create_sheet("Matches")

        # Headers
        headers = [
            "Match Type", "Confidence", "Invoice Number", "Invoice Date", "Invoice Amount",
            "Transaction Date", "Transaction Amount", "Amount Variance", "Date Variance (days)",
            "Score - Amount", "Score - Date", "Score - Vendor", "Score - Reference", "Status"
        ]

        self._apply_header_style(ws, headers)

        # Data
        for idx, match in enumerate(matches, start=2):
            ws[f'A{idx}'] = (match.match_type or "auto").upper()
            ws[f'B{idx}'] = f"{(match.match_confidence or 0) * 100:.1f}%"

            # Get related invoice and transaction
            invoice = match.invoice if hasattr(match, 'invoice') else None
            transaction = match.transaction if hasattr(match, 'transaction') else None

            ws[f'C{idx}'] = invoice.invoice_number if invoice else "-"
            ws[f'D{idx}'] = self._format_date(invoice.invoice_date) if invoice else "-"
            ws[f'E{idx}'] = invoice.total_amount if invoice else 0
            ws[f'F{idx}'] = self._format_date(transaction.transaction_date) if transaction else "-"
            ws[f'G{idx}'] = (transaction.credit or transaction.debit) if transaction else 0
            ws[f'H{idx}'] = match.amount_variance or 0
            ws[f'I{idx}'] = match.date_variance_days or 0
            ws[f'J{idx}'] = f"{(match.score_amount or 0) * 100:.1f}%"
            ws[f'K{idx}'] = f"{(match.score_date or 0) * 100:.1f}%"
            ws[f'L{idx}'] = f"{(match.score_vendor or 0) * 100:.1f}%"
            ws[f'M{idx}'] = f"{(match.score_reference or 0) * 100:.1f}%"
            ws[f'N{idx}'] = (match.status or "active").upper()

            # Apply number format to currency columns
            for col in ['E', 'G', 'H']:
                ws[f'{col}{idx}'].number_format = '#,##0'

            # Color code confidence
            self._apply_confidence_color(ws[f'B{idx}'], match.match_confidence or 0)

        # Auto-size columns
        for col_idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _apply_header_style(self, ws, headers):
        """Apply styling to header row"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _apply_match_status_color(self, cell, status):
        """Apply color based on match status"""
        if status == "auto_matched":
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        elif status == "manual_matched":
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        elif status == "unmatched":
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

    def _apply_confidence_color(self, cell, confidence):
        """Apply color based on confidence level"""
        if confidence >= 0.90:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        elif confidence >= 0.70:
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        else:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

    def _format_rupiah(self, value) -> str:
        """Format number to Rupiah"""
        if not value or value == 0:
            return "Rp 0"
        try:
            return f"Rp {int(value):,}".replace(',', '.')
        except:
            return str(value)

    def _format_date(self, date_obj) -> str:
        """Format date to DD/MM/YYYY"""
        if not date_obj:
            return "-"
        if isinstance(date_obj, datetime):
            return date_obj.strftime("%d/%m/%Y")
        return str(date_obj)
