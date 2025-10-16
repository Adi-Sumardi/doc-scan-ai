"""
Faktur Pajak Exporter
Handles Excel and PDF export specifically for Indonesian Tax Invoice (Faktur Pajak) documents
"""

from typing import Dict, Any, Tuple
from datetime import datetime
import logging
import re

from .base_exporter import BaseExporter

logger = logging.getLogger(__name__)

# Check for required libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("⚠️ openpyxl not installed - Excel export disabled")

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    logger.warning("⚠️ reportlab not installed - PDF export disabled")


class FakturPajakExporter(BaseExporter):
    """Exporter for Faktur Pajak (Indonesian Tax Invoice) documents"""
    
    def __init__(self):
        super().__init__("faktur_pajak")
        self.columns = [
            "Nama Seller",
            "Alamat Seller",
            "NPWP Seller",
            "Nama Buyer",
            "Alamat Buyer",
            "NPWP Buyer",
            "Email Buyer",
            "Tgl",
            "Nomor Faktur",
            "DPP",
            "PPN",
            "Total",
            "Invoice",
            "Nama Barang Kena Pajak / Jasa Kena Pajak",
            "Quantity",
            "Nilai Barang",
            "Total Nilai Barang"
        ]

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _format_rupiah(self, value) -> str:
        """Format number to Rupiah without decimals"""
        if not value or value in ['-', 'N/A', '', '0', 0, 'None', 'null']:
            return '-'
        try:
            # If already a number (int or float), convert to int first
            if isinstance(value, (int, float)):
                value_int = int(value)
            else:
                # If string, clean and parse
                value_str = str(value).replace('Rp', '').replace('IDR', '').replace('.', '').replace(',', '').strip()
                if not value_str or value_str == '-':
                    return '-'
                value_int = int(float(value_str))

            if value_int == 0:
                return '-'
            formatted = f"Rp {value_int:,}"
            formatted = formatted.replace(',', '.')
            return formatted
        except (ValueError, AttributeError):
            logger.warning(f"⚠️ Failed to format rupiah value: {value}")
            return str(value) if value else '-'

    def _format_date(self, date_str) -> str:
        """Format date to DD/MM/YYYY"""
        if not date_str or date_str in ['N/A', '', 'None', 'null']:
            return 'N/A'

        import re
        from datetime import datetime

        date_str = str(date_str).strip()

        # Try various date formats
        formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%Y',
            '%d %B %Y', '%d %b %Y', '%d-%m-%y', '%d/%m/%y',
            '%d %B %Y', '%d %b %Y'  # Indonesian month names
        ]

        for fmt in formats:
            try:
                date_obj = datetime.strptime(date_str, fmt)
                return date_obj.strftime('%d/%m/%Y')
            except:
                continue

        # Try to parse "DD Bulan YYYY" format (current format in exporter)
        # Example: "12 Oktober 2025"
        month_map = {
            'januari': 1, 'februari': 2, 'maret': 3, 'april': 4,
            'mei': 5, 'juni': 6, 'juli': 7, 'agustus': 8,
            'september': 9, 'oktober': 10, 'november': 11, 'desember': 12,
            'january': 1, 'february': 2, 'march': 3, 'may': 5,
            'june': 6, 'july': 7, 'august': 8, 'october': 10, 'december': 12
        }

        match = re.search(r'(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})', date_str, re.IGNORECASE)
        if match:
            day = int(match.group(1))
            month_name = match.group(2).lower()
            year = int(match.group(3))
            month = month_map.get(month_name)
            if month and 1 <= day <= 31:
                return f"{day:02d}/{month:02d}/{year}"

        # If no format worked, try to extract DD, MM, YYYY
        match = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', date_str)
        if match:
            return f"{match.group(1).zfill(2)}/{match.group(2).zfill(2)}/{match.group(3)}"

        return date_str  # Return as-is if can't format

    def _parse_amount(self, raw_value: Any) -> Any:
        """Convert Rupiah strings into integers for Excel numeric formatting."""
        if raw_value in (None, "", "N/A"):
            return "N/A"
        if isinstance(raw_value, (int, float)):
            return raw_value

        text = str(raw_value).strip()
        if not text:
            return "N/A"

        import re
        digits = re.sub(r"[^0-9]", "", text)
        if digits:
            try:
                return int(digits)
            except ValueError:
                return text
        return text

    def _create_items_description_list(self, items: list) -> str:
        """
        Create simple description list for items (no qty, no price)

        Format for 1 item:
        Laptop Dell Inspiron 15

        Format for multiple items:
        1. Laptop Dell
        2. Mouse Wireless
        3. Keyboard Mechanical
        """
        if not items or len(items) == 0:
            return "N/A"

        # If only 1 item, show simple format (just description)
        if len(items) == 1:
            item = items[0]
            desc = item.get('description', 'N/A')
            return desc

        # Multiple items: Create numbered list with descriptions only
        items_text = []
        for idx, item in enumerate(items, start=1):
            desc = item.get('description', 'N/A')
            items_text.append(f"{idx}. {desc}")

        return "\n".join(items_text)

    def _calculate_total_quantity(self, items: list) -> str:
        """
        Calculate total quantity from all items
        Returns formatted quantity string
        """
        if not items or len(items) == 0:
            return "-"

        total_qty = 0
        for item in items:
            qty_str = item.get('quantity', '-')
            if qty_str and qty_str != '-':
                try:
                    import re
                    qty_clean = re.sub(r'[^\d.]', '', str(qty_str))
                    if qty_clean:
                        qty = float(qty_clean)
                        total_qty += qty
                except:
                    pass

        if total_qty == 0:
            return "-"

        # Return integer if whole number, otherwise with decimals
        if total_qty == int(total_qty):
            return str(int(total_qty))
        else:
            return f"{total_qty:.2f}"

    def _calculate_nilai_barang_satuan(self, items: list) -> str:
        """
        Get unit prices (nilai barang satuan) from ALL items, one per line
        Returns formatted string with each item's unit price

        Format for 1 item:
        Rp 5.000.000

        Format for multiple items:
        1. Rp 5.000.000
        2. Rp 3.000.000
        3. Rp 7.500.000
        """
        if not items or len(items) == 0:
            return "-"

        # If only 1 item, show simple format (just price)
        if len(items) == 1:
            item = items[0]
            price_str = item.get('unit_price', '-')
            if not price_str or price_str == '-':
                return "-"
            price = self._parse_price(price_str)
            if price == 0:
                return "-"
            return self._format_rupiah(price)

        # Multiple items: Create numbered list with unit prices
        price_lines = []
        for idx, item in enumerate(items, start=1):
            price_str = item.get('unit_price', '-')
            if not price_str or price_str == '-':
                price_lines.append(f"{idx}. -")
            else:
                price = self._parse_price(price_str)
                if price == 0:
                    price_lines.append(f"{idx}. -")
                else:
                    price_lines.append(f"{idx}. {self._format_rupiah(price)}")

        return "\n".join(price_lines)

    def _parse_price(self, price_str: str) -> int:
        """
        Parse price string to integer
        Handles Indonesian format (dots as thousand separator, comma as decimal)
        """
        if not price_str or price_str == '-':
            return 0

        try:
            import re
            price_text = str(price_str).strip()

            # Remove currency symbols
            price_text = price_text.replace('Rp', '').replace('IDR', '').strip()

            # Detect format based on separators
            has_comma = ',' in price_text
            has_dot = '.' in price_text

            if has_comma and has_dot:
                # Both comma and dot present
                # Indonesian format: 5.000.000,00 (dots = thousand, comma = decimal)
                # Remove dots, replace comma with dot
                price_text = price_text.replace('.', '').replace(',', '.')
                price = float(price_text)
                # Buang decimal (take integer part only)
                return int(price)
            elif has_comma and not has_dot:
                # Only comma: could be Indonesian 5000000,00 or mistake
                # Assume comma is decimal separator
                price_text = price_text.replace(',', '.')
                price = float(price_text)
                return int(price)
            elif has_dot and not has_comma:
                # Only dot: could be Indonesian 5.000.000 or US 5000000.00
                # Count dots to determine
                dot_count = price_text.count('.')
                if dot_count >= 2:
                    # Multiple dots = Indonesian thousand separator (5.000.000)
                    price_text = price_text.replace('.', '')
                    return int(price_text)
                else:
                    # Single dot - check position
                    parts = price_text.split('.')
                    if len(parts) == 2 and len(parts[1]) == 2:
                        # Last part is 2 digits = decimal (5000000.00)
                        price = float(price_text)
                        return int(price)
                    else:
                        # Probably thousand separator or just clean number
                        price_text = price_text.replace('.', '')
                        return int(price_text)
            else:
                # No separators - just digits
                return int(price_text)
        except Exception as e:
            logger.warning(f"⚠️ Failed to parse unit price '{price_str}': {e}")
            return 0

    def _calculate_total_nilai_barang(self, items: list) -> str:
        """
        Calculate total nilai barang (quantity × unit_price for all items)
        Returns formatted rupiah string
        """
        if not items or len(items) == 0:
            return "-"

        grand_total = 0
        for idx, item in enumerate(items, 1):
            qty_str = item.get('quantity', '-')
            price_str = item.get('unit_price', '-')

            # Parse quantity
            qty = 0
            if qty_str and qty_str != '-':
                try:
                    import re
                    qty_clean = re.sub(r'[^\d.]', '', str(qty_str))
                    if qty_clean:
                        qty = float(qty_clean)
                except:
                    qty = 0

            # Parse price using helper function
            price = self._parse_price(price_str)

            # Calculate subtotal (convert to int to avoid float issues)
            subtotal = int(qty * price) if qty > 0 and price > 0 else 0
            grand_total += subtotal

        if grand_total == 0:
            return "-"

        return self._format_rupiah(grand_total)

    def _format_barang_jasa(self, value: Any) -> str:
        """Normalize barang/jasa data into a readable comma-separated list."""
        if value is None:
            return "N/A"

        # Prebuild parts list while preserving order without duplicates
        parts = []

        def _consume(raw):
            if raw is None:
                return
            text = str(raw).replace('\r', '\n')
            for piece in text.replace('\n', ',').split(','):
                cleaned = piece.strip()
                if not cleaned:
                    continue
                lowered = cleaned.lower()
                if lowered in ('alamat', 'npwp'):
                    continue
                if lowered.startswith(('alamat', 'npwp', 'nomor', 'no invoice', 'invoice', 'kode')):
                    continue
                if re.match(r'^nama\s+barang', lowered):
                    continue
                if re.match(r'^nama\s*(penjual|pembeli|pt|cv)', lowered):
                    continue
                if cleaned not in parts:
                    parts.append(cleaned)

        if isinstance(value, list):
            for item in value:
                _consume(item)
        else:
            _consume(value)

        return ', '.join(parts) if parts else "N/A"

    def _get_invoice_reference(self, structured: Dict[str, Any]) -> str:
        """Retrieve invoice reference from structured payload with flexible key lookup."""
        if not structured:
            return "N/A"

        candidates = [
            structured.get('invoice'),
            structured.get('nomor_invoice'),
            structured.get('invoice_number'),
            structured.get('referensi_invoice'),
        ]

        for value in candidates:
            if value and str(value).strip():
                return str(value).strip()

        return "N/A"

    # ------------------------------------------------------------------
    # Advanced field sanitizers
    # ------------------------------------------------------------------
    MONTH_NAMES_ID = {
        1: "Januari",
        2: "Februari",
        3: "Maret",
        4: "April",
        5: "Mei",
        6: "Juni",
        7: "Juli",
        8: "Agustus",
        9: "September",
        10: "Oktober",
        11: "November",
        12: "Desember",
    }

    MONTH_LOOKUP = {
        "jan": 1,
        "januari": 1,
        "january": 1,
        "feb": 2,
        "februari": 2,
        "february": 2,
        "mar": 3,
        "maret": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "mei": 5,
        "jun": 6,
        "juni": 6,
        "jun": 6,
        "jul": 7,
        "juli": 7,
        "july": 7,
        "agu": 8,
        "agt": 8,
        "agus": 8,
        "agustus": 8,
        "aug": 8,
        "september": 9,
        "sept": 9,
        "sep": 9,
        "oct": 10,
        "okt": 10,
        "oktober": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "des": 12,
        "desember": 12,
        "december": 12,
    }

    def _standardize_amount(self, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if not text or text.upper() == "N/A":
            return ""
        digits = re.sub(r"[^0-9]", "", text)
        if not digits:
            return text
        try:
            return f"Rp {int(digits):,}"
        except ValueError:
            return text

    def _extract_amount_from_text(self, raw_text: str, labels: Tuple[str, ...]) -> str:
        if not raw_text:
            return ""
        pattern = rf"(?:{'|'.join(labels)})\s*:?\s*(?:Rp\.?\s*)?([0-9.,]+)"
        match = re.search(pattern, raw_text, re.IGNORECASE)
        if not match:
            return ""
        return self._standardize_amount(match.group(1))

    def _normalize_amount_field(self, value: Any, raw_text: str, labels: Tuple[str, ...]) -> str:
        normalized = self._standardize_amount(value)
        if normalized:
            return normalized
        fallback = self._extract_amount_from_text(raw_text, labels)
        return fallback if fallback else ""

    def _format_npwp(self, value: str) -> str:
        digits = re.sub(r"[^0-9]", "", value)
        if len(digits) != 15:
            return value.strip()
        return f"{digits[0:2]}.{digits[2:5]}.{digits[5:8]}.{digits[8]}-{digits[9:12]}.{digits[12:15]}"

    def _extract_npwp(self, value: Any, raw_text: str) -> str:
        candidates = []
        if value:
            candidates.append(str(value))
        if raw_text:
            candidates.extend(re.findall(r"\d{2}\.?\d{3}\.?\d{3}\.?\d-?\d{3}\.?\d{3}", raw_text))
        for candidate in candidates:
            formatted = self._format_npwp(candidate)
            if formatted and len(re.sub(r"[^0-9]", "", formatted)) == 15:
                return formatted
        return ""

    def _extract_nomor_faktur(self, value: Any, raw_text: str) -> str:
        candidates = []
        if value:
            candidates.append(str(value))
        if raw_text:
            candidates.extend(re.findall(r"\d{3}\.\d{3}-\d{2}\.\d{8}", raw_text))
        for candidate in candidates:
            cleaned = candidate.strip().replace(" ", "")
            if re.fullmatch(r"\d{3}\.\d{3}-\d{2}\.\d{8}", cleaned):
                return cleaned
        return ""

    def _clean_company_name(self, value: Any, raw_text: str) -> str:
        text = str(value or "").strip()
        if not text and raw_text:
            match = re.search(
                r"Nama\s*(?:Penjual|Perusahaan)?\s*:?\s*([^\n]+)",
                raw_text,
                re.IGNORECASE,
            )
            if match:
                text = match.group(1).strip()
        if not text:
            return ""
        tokens = [tok.strip(" ,") for tok in re.split(r"[\r\n]+", text) if tok.strip()]
        filtered = []
        for token in tokens:
            if re.match(r"^(alamat|npwp|nomor|kode|invoice)", token, re.IGNORECASE):
                continue
            filtered.append(token)
        cleaned = ' '.join(filtered if filtered else tokens)
        cleaned = re.sub(r"\s+", " ", cleaned)
        return cleaned.strip(" ,")

    def _clean_address(self, value: Any, raw_text: str) -> str:
        text = str(value or "").strip()
        lines = [line.strip(" ,") for line in re.split(r"[\r\n]+", text) if line.strip()]
        filtered = []
        for line in lines:
            if re.match(r"^(npwp|nomor|no\.?|kode|invoice|nama)", line, re.IGNORECASE):
                break
            filtered.append(line)
        if not filtered and raw_text:
            pattern = re.compile(r"Alamat\s*:?\s*(.+?)(?:\n\s*(?:NPWP|Nomor|No\.?|Kode|Invoice)\b|$)", re.IGNORECASE | re.DOTALL)
            match = pattern.search(raw_text)
            if match:
                raw_addr = match.group(1).strip()
                raw_addr = re.sub(r"\s+", " ", raw_addr)
                filtered = [raw_addr]
        address = ', '.join(filtered)
        return address.strip(" ,")

    def _extract_barang_jasa_from_text(self, raw_text: str) -> str:
        if not raw_text:
            return "N/A"

        lines = [line.strip() for line in raw_text.splitlines()]
        collecting = False
        collected = []

        for line in lines:
            lower = line.lower()
            if not collecting:
                if 'nama barang kena pajak' in lower or 'nama barang' in lower:
                    collecting = True
                continue

            if not line:
                if collected:
                    break
                continue

            if re.match(r"^(harga|dpp|dasar|total|pajak|npwp|alamat|kode|nomor|no\.?|jumlah|invoice)", lower):
                if collected:
                    break
                else:
                    continue

            if re.fullmatch(r"\d+", line):
                continue
            if lower.startswith('rp'):
                continue

            cleaned = line.strip(' ,;')
            if not cleaned:
                continue
            if cleaned not in collected:
                collected.append(cleaned)

        return ', '.join(collected) if collected else "N/A"

    def _parse_date_candidate(self, candidate: str) -> str:
        if not candidate:
            return ""
        text = candidate.strip().strip(",.;")
        if not text:
            return ""

        # Month name pattern (Indonesian/English)
        match = re.search(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{2,4})", text, re.IGNORECASE)
        if match:
            day = int(match.group(1))
            month_name = match.group(2).lower()
            year = int(match.group(3))
            month = self.MONTH_LOOKUP.get(month_name)
            if month and 1 <= day <= 31:
                if year < 100:
                    year += 2000 if year < 50 else 1900
                if 1900 <= year <= 2100:
                    return f"{day:02d} {self.MONTH_NAMES_ID[month]} {year}"

        # Numeric patterns (dd/mm/yyyy, dd-mm-yyyy, dd.mm.yyyy)
        numeric_patterns = [
            (r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", "dmy"),
            (r"^(\d{1,2})\.(\d{1,2})\.(\d{2,4})$", "dmy"),
            (r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", "ymd"),
        ]
        for pattern, order in numeric_patterns:
            num_match = re.match(pattern, text)
            if not num_match:
                continue
            groups = [int(g) for g in num_match.groups()]
            if order == "ymd":
                year, month, day = groups
            else:
                day, month, year = groups
            if year < 100:
                year += 2000 if year < 50 else 1900
            if not (1 <= day <= 31 and 1 <= month <= 12 and 1900 <= year <= 2100):
                continue
            return f"{day:02d} {self.MONTH_NAMES_ID[month]} {year}"

        return ""

    def _normalize_date(self, value: Any, raw_text: str) -> str:
        candidates = []
        if value:
            candidates.append(str(value))
        if raw_text:
            label_pattern = re.compile(r"(?:Tanggal\s*(?:Faktur)?|Tgl\.?|Tanggal Invoice)\s*:?\s*([^\n]+)", re.IGNORECASE)
            for match in label_pattern.finditer(raw_text):
                candidates.append(match.group(1))
            generic_numeric = re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", raw_text)
            candidates.extend(generic_numeric)
            month_name_candidates = re.findall(r"\b\d{1,2}\s+[A-Za-z]+\s+\d{2,4}\b", raw_text)
            candidates.extend(month_name_candidates)

        seen = set()
        for candidate in candidates:
            normalized = self._parse_date_candidate(candidate)
            if normalized and normalized not in seen:
                return normalized
            seen.add(candidate)
        return ""

    def _convert_smart_mapped_to_structured(self, smart_mapped: Dict[str, Any]) -> Dict[str, Any]:
        """Convert Smart Mapper GPT output to legacy structured format for Excel export"""
        seller = smart_mapped.get('seller', {})
        buyer = smart_mapped.get('buyer', {})
        invoice = smart_mapped.get('invoice', {})
        financials = smart_mapped.get('financials', {})
        items = smart_mapped.get('items', [])
        
        # DEBUG: Log FULL smart_mapped data to see exact keys
        logger.info(f"🔍 DEBUG - FULL smart_mapped keys: {list(smart_mapped.keys())}")
        logger.info(f"🔍 DEBUG - financials type: {type(financials)}")
        logger.info(f"🔍 DEBUG - financials keys: {list(financials.keys()) if isinstance(financials, dict) else 'NOT A DICT'}")
        logger.info(f"🔍 DEBUG - financials full data: {financials}")
        
        # Try BOTH lowercase and uppercase keys (case-insensitive)
        dpp_value = financials.get('dpp') or financials.get('DPP') or ''
        ppn_value = financials.get('ppn') or financials.get('PPN') or ''
        total_value = financials.get('total') or financials.get('Total') or financials.get('TOTAL') or ''
        
        logger.info(f"🔍 DEBUG - Extracted DPP: '{dpp_value}'")
        logger.info(f"🔍 DEBUG - Extracted PPN: '{ppn_value}'")
        logger.info(f"🔍 DEBUG - Extracted Total: '{total_value}'")
        
        # Extract item descriptions
        item_descriptions = []
        for item in items:
            if isinstance(item, dict):
                desc = item.get('description', '')
                if desc:
                    item_descriptions.append(desc)
        
        structured = {
            # Seller data
            'nama_seller': seller.get('name', ''),
            'alamat_seller': seller.get('address', ''),
            'npwp_seller': seller.get('npwp', ''),

            # Buyer data
            'nama_buyer': buyer.get('name', ''),
            'alamat_buyer': buyer.get('address', ''),
            'npwp_buyer': buyer.get('npwp', ''),
            'email_buyer': buyer.get('email', ''),

            # Invoice data
            'nomor_faktur': invoice.get('number', ''),
            'tanggal': invoice.get('issue_date', ''),
            'invoice': invoice.get('reference', ''),

            # Financial data
            'dpp': dpp_value,
            'ppn': ppn_value,
            'total': total_value,

            # Items (legacy)
            'nama_barang_jasa': ', '.join(item_descriptions) if item_descriptions else '',

            # Legacy fields (untuk backward compatibility)
            'nama': seller.get('name', ''),
            'npwp': seller.get('npwp', ''),
            'alamat': seller.get('address', ''),
        }

        return structured
    
    def _prepare_structured_fields(self, structured: Dict[str, Any], raw_text: str) -> Dict[str, Any]:
        base = structured if isinstance(structured, dict) else {}
        cleaned: Dict[str, Any] = dict(base)

        cleaned['nama'] = self._clean_company_name(base.get('nama'), raw_text)
        cleaned['alamat'] = self._clean_address(base.get('alamat'), raw_text)
        cleaned['npwp'] = self._extract_npwp(base.get('npwp'), raw_text)
        cleaned['nomor_faktur'] = self._extract_nomor_faktur(base.get('nomor_faktur'), raw_text)
        cleaned['tanggal'] = self._normalize_date(base.get('tanggal'), raw_text)
        cleaned['dpp'] = self._normalize_amount_field(base.get('dpp'), raw_text, ('DPP', 'Dasar Pengenaan Pajak', 'Harga Jual'))
        cleaned['ppn'] = self._normalize_amount_field(base.get('ppn'), raw_text, ('PPN', 'Pajak Pertambahan Nilai'))
        cleaned['total'] = self._normalize_amount_field(base.get('total'), raw_text, ('Total', 'Jumlah', 'Grand Total', 'Total Harga'))
        cleaned['nama_barang_jasa'] = self._format_barang_jasa(base.get('nama_barang_jasa'))
        if not cleaned['nama_barang_jasa'] or cleaned['nama_barang_jasa'] == "N/A":
            fallback_barang = self._extract_barang_jasa_from_text(raw_text)
            cleaned['nama_barang_jasa'] = self._format_barang_jasa(fallback_barang)
        cleaned['invoice'] = self._get_invoice_reference({**base, **cleaned})

        return cleaned
    
    def export_to_excel(self, result: Dict[str, Any], output_path: str) -> bool:
        """Export single Faktur Pajak to Excel with 9-column standardized format"""
        try:
            if not HAS_OPENPYXL:
                logger.error("❌ openpyxl not available for Excel export")
                return False
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Faktur Pajak"
            
            self._populate_excel_sheet(ws, result)
            
            wb.save(output_path)
            logger.info(f"✅ Faktur Pajak Excel export created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Faktur Pajak Excel export failed: {e}", exc_info=True)
            return False
    
    def _populate_excel_sheet(self, ws, result: Dict[str, Any]):
        """Helper to populate worksheet with structured Faktur Pajak data"""
        # Define styles
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        data_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
        border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Get structured data
        extracted_data = result.get('extracted_data', {})
        raw_text = extracted_data.get('raw_text', '')
        
        # PRIORITY 1: Use Smart Mapper data if available
        smart_mapped = extracted_data.get('smart_mapped', {})
        items = []  # Store items for nested table
        if smart_mapped:
            logger.info("✅ Using Smart Mapper GPT data for Excel export (CLEAN, NO POST-PROCESSING)")
            # Smart Mapper data is already clean from GPT-4o, no need for _prepare_structured_fields()
            structured = self._convert_smart_mapped_to_structured(smart_mapped)
            items = smart_mapped.get('items', [])  # Get items for nested table
        else:
            # FALLBACK: Use legacy structured_data
            logger.info("⚠️ Smart Mapper not available, using legacy parser with post-processing")
            structured = extracted_data.get('structured_data', {})

            # If no structured data, try to extract from text
            if not structured or not any(structured.values()):
                if raw_text:
                    from ai_processor import IndonesianTaxDocumentParser
                    temp_parser = IndonesianTaxDocumentParser()
                    structured = temp_parser.extract_structured_fields(raw_text, "faktur_pajak")

            # ONLY apply heavy post-processing for legacy data (not for Smart Mapper)
            structured = self._prepare_structured_fields(structured, raw_text)
        
        row = 1

        # ===== TITLE =====
        ws.merge_cells(f'A{row}:Q{row}')
        ws[f'A{row}'] = "📋 FAKTUR PAJAK - DATA TERSTRUKTUR"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_align
        ws[f'A{row}'].border = border_thin
        row += 1
        
        # ===== TABLE HEADERS =====
        for col_idx, header in enumerate(self.columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_thin
        row += 1
        
        # ===== DATA ROW =====
        data_row = [
            # Seller data (columns 1-3)
            structured.get('nama_seller') or 'N/A',
            structured.get('alamat_seller') or 'N/A',
            structured.get('npwp_seller') or 'N/A',

            # Buyer data (columns 4-7)
            structured.get('nama_buyer') or 'N/A',
            structured.get('alamat_buyer') or 'N/A',
            structured.get('npwp_buyer') or 'N/A',
            structured.get('email_buyer') or 'N/A',

            # Invoice & Financial data (columns 8-13)
            self._format_date(structured.get('tanggal')) or 'N/A',
            structured.get('nomor_faktur') or 'N/A',
            self._format_rupiah(structured.get('dpp')),
            self._format_rupiah(structured.get('ppn')),
            self._format_rupiah(structured.get('total')),
            structured.get('invoice') or 'N/A',
        ]

        # Write data row (columns 1-13)
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill = data_fill
            cell.alignment = left_align
            cell.border = border_thin
            cell.font = Font(size=10)

        # Column 14: Nama Barang/Jasa (descriptions only)
        if items and len(items) > 0:
            desc_text = self._create_items_description_list(items)
            cell = ws.cell(row=row, column=14, value=desc_text)
            cell.fill = data_fill
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = border_thin
            cell.font = Font(size=10)
        else:
            # No items from smart mapper, use fallback text
            cell = ws.cell(row=row, column=14, value=structured.get('nama_barang_jasa') or 'N/A')
            cell.fill = data_fill
            cell.alignment = left_align
            cell.border = border_thin
            cell.font = Font(size=10)

        # Column 15: Quantity (total)
        if items and len(items) > 0:
            qty_text = self._calculate_total_quantity(items)
        else:
            qty_text = '-'
        cell = ws.cell(row=row, column=15, value=qty_text)
        cell.fill = data_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
        cell.font = Font(size=10)

        # Column 16: Nilai Barang (satuan/unit price)
        if items and len(items) > 0:
            nilai_satuan_text = self._calculate_nilai_barang_satuan(items)
        else:
            nilai_satuan_text = '-'
        cell = ws.cell(row=row, column=16, value=nilai_satuan_text)
        cell.fill = data_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border_thin
        cell.font = Font(size=10)

        # Column 17: Total Nilai Barang (qty × unit_price)
        if items and len(items) > 0:
            total_nilai_text = self._calculate_total_nilai_barang(items)
        else:
            total_nilai_text = '-'
        cell = ws.cell(row=row, column=17, value=total_nilai_text)
        cell.fill = data_fill
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border_thin
        cell.font = Font(size=10)
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25  # Nama Seller
        ws.column_dimensions['B'].width = 35  # Alamat Seller
        ws.column_dimensions['C'].width = 20  # NPWP Seller
        ws.column_dimensions['D'].width = 25  # Nama Buyer
        ws.column_dimensions['E'].width = 35  # Alamat Buyer
        ws.column_dimensions['F'].width = 20  # NPWP Buyer
        ws.column_dimensions['G'].width = 30  # Email Buyer
        ws.column_dimensions['H'].width = 12  # Tgl
        ws.column_dimensions['I'].width = 18  # Nomor Faktur
        ws.column_dimensions['J'].width = 15  # DPP
        ws.column_dimensions['K'].width = 15  # PPN
        ws.column_dimensions['L'].width = 15  # Total
        ws.column_dimensions['M'].width = 18  # Invoice
        ws.column_dimensions['N'].width = 40  # Nama Barang/Jasa
        ws.column_dimensions['O'].width = 12  # Quantity
        ws.column_dimensions['P'].width = 18  # Nilai Barang (satuan)
        ws.column_dimensions['Q'].width = 20  # Total Nilai Barang

        # Set row height for data row (increase if multiple items)
        if items and len(items) > 1:
            # Calculate height based on number of items
            # Each item line ~15px, padding ~10px
            estimated_height = (len(items) * 15) + 10
            ws.row_dimensions[3].height = min(estimated_height, 150)  # Max 150
        else:
            ws.row_dimensions[3].height = 40
    
    def export_to_pdf(self, result: Dict[str, Any], output_path: str) -> bool:
        """
        Export single Faktur Pajak to professional PDF document
        Uses new document format (not table) with 2 pages: Keluaran and Masukan
        """
        # Use new document format
        return self.export_to_pdf_document(result, output_path)
    
    def batch_export_to_excel(self, batch_id: str, results: list, output_path: str) -> bool:
        """Export multiple Faktur Pajak documents to single Excel file with 9-column table"""
        try:
            if not HAS_OPENPYXL:
                logger.error("❌ openpyxl not available for batch Excel export")
                return False
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Batch Faktur Pajak"
            
            # Define styles
            header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
            header_font = Font(bold=True, size=11, color="FFFFFF")
            data_fill_1 = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
            data_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            border_thin = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            right_align = Alignment(horizontal='right', vertical='center')
            
            row = 1
            
            # ===== BATCH INFO =====
            ws.merge_cells(f'A{row}:Q{row}')
            ws[f'A{row}'] = f"📦 BATCH: {batch_id} | Total: {len(results)} Documents | Export: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = center_align
            ws[f'A{row}'].border = border_thin
            row += 1
            
            # ===== TABLE HEADERS =====
            for col_idx, header in enumerate(self.columns, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border_thin
            row += 1
            
            # ===== DATA ROWS (One row per document) =====
            from ai_processor import IndonesianTaxDocumentParser
            temp_parser = IndonesianTaxDocumentParser()
            
            for idx, result_dict in enumerate(results):
                # Get structured data
                extracted_data = result_dict.get('extracted_data', {})
                raw_text = extracted_data.get('raw_text', '')

                # PRIORITY: Use Smart Mapper data if available
                smart_mapped = extracted_data.get('smart_mapped', {})
                items = []
                if smart_mapped:
                    structured = self._convert_smart_mapped_to_structured(smart_mapped)
                    items = smart_mapped.get('items', [])

                    # DEBUG: Log items extraction untuk batch export
                    logger.info(f"🔍 BATCH EXPORT - Doc #{idx+1}: smart_mapped keys = {list(smart_mapped.keys())}")
                    logger.info(f"🔍 BATCH EXPORT - Doc #{idx+1}: items count = {len(items)}")
                    if items and len(items) > 0:
                        logger.info(f"🔍 BATCH EXPORT - Doc #{idx+1}: First item = {items[0]}")
                        logger.info(f"🔍 BATCH EXPORT - Doc #{idx+1}: Has quantity? {items[0].get('quantity', 'NO')}")
                        logger.info(f"🔍 BATCH EXPORT - Doc #{idx+1}: Has unit_price? {items[0].get('unit_price', 'NO')}")
                    else:
                        logger.warning(f"⚠️ BATCH EXPORT - Doc #{idx+1}: NO ITEMS FOUND in smart_mapped!")
                else:
                    logger.warning(f"⚠️ BATCH EXPORT - Doc #{idx+1}: No smart_mapped data, using fallback")
                    structured = extracted_data.get('structured_data', {})
                    # If no structured data, try to extract
                    if not structured or not any(structured.values()):
                        if raw_text:
                            structured = temp_parser.extract_structured_fields(raw_text, "faktur_pajak")
                    structured = self._prepare_structured_fields(structured, raw_text)

                # Data row (seller + buyer + invoice data)
                data_row = [
                    # Seller data (columns 1-3)
                    structured.get('nama_seller') or 'N/A',
                    structured.get('alamat_seller') or 'N/A',
                    structured.get('npwp_seller') or 'N/A',

                    # Buyer data (columns 4-7)
                    structured.get('nama_buyer') or 'N/A',
                    structured.get('alamat_buyer') or 'N/A',
                    structured.get('npwp_buyer') or 'N/A',
                    structured.get('email_buyer') or 'N/A',

                    # Invoice & Financial data (columns 8-13)
                    self._format_date(structured.get('tanggal')) or 'N/A',
                    structured.get('nomor_faktur') or 'N/A',
                    self._format_rupiah(structured.get('dpp')),
                    self._format_rupiah(structured.get('ppn')),
                    self._format_rupiah(structured.get('total')),
                    structured.get('invoice') or 'N/A',
                ]

                fill = data_fill_1 if idx % 2 == 0 else data_fill_2

                # Write data row (columns 1-13)
                for col_idx, value in enumerate(data_row, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.fill = fill
                    cell.alignment = left_align
                    cell.border = border_thin
                    cell.font = Font(size=10)

                # Column 14: Nama Barang/Jasa (descriptions only)
                if items and len(items) > 0:
                    desc_text = self._create_items_description_list(items)
                    cell = ws.cell(row=row, column=14, value=desc_text)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.border = border_thin
                    cell.font = Font(size=10)
                    # Set row height for multiple items
                    if len(items) > 1:
                        estimated_height = (len(items) * 15) + 10
                        ws.row_dimensions[row].height = min(estimated_height, 150)
                else:
                    cell = ws.cell(row=row, column=14, value=structured.get('nama_barang_jasa') or 'N/A')
                    cell.fill = fill
                    cell.alignment = left_align
                    cell.border = border_thin
                    cell.font = Font(size=10)

                # Column 15: Quantity (total)
                if items and len(items) > 0:
                    qty_text = self._calculate_total_quantity(items)
                else:
                    qty_text = '-'
                cell = ws.cell(row=row, column=15, value=qty_text)
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
                cell.font = Font(size=10)

                # Column 16: Nilai Barang (satuan/unit price)
                if items and len(items) > 0:
                    nilai_satuan_text = self._calculate_nilai_barang_satuan(items)
                else:
                    nilai_satuan_text = '-'
                cell = ws.cell(row=row, column=16, value=nilai_satuan_text)
                cell.fill = fill
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border_thin
                cell.font = Font(size=10)

                # Column 17: Total Nilai Barang (qty × unit_price)
                if items and len(items) > 0:
                    total_nilai_text = self._calculate_total_nilai_barang(items)
                else:
                    total_nilai_text = '-'
                cell = ws.cell(row=row, column=17, value=total_nilai_text)
                cell.fill = fill
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = border_thin
                cell.font = Font(size=10)

                row += 1

            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 25  # Nama Seller
            ws.column_dimensions['B'].width = 35  # Alamat Seller
            ws.column_dimensions['C'].width = 20  # NPWP Seller
            ws.column_dimensions['D'].width = 25  # Nama Buyer
            ws.column_dimensions['E'].width = 35  # Alamat Buyer
            ws.column_dimensions['F'].width = 20  # NPWP Buyer
            ws.column_dimensions['G'].width = 30  # Email Buyer
            ws.column_dimensions['H'].width = 12  # Tgl
            ws.column_dimensions['I'].width = 18  # Nomor Faktur
            ws.column_dimensions['J'].width = 15  # DPP
            ws.column_dimensions['K'].width = 15  # PPN
            ws.column_dimensions['L'].width = 15  # Total
            ws.column_dimensions['M'].width = 18  # Invoice
            ws.column_dimensions['N'].width = 40  # Nama Barang/Jasa
            ws.column_dimensions['O'].width = 12  # Quantity
            ws.column_dimensions['P'].width = 18  # Nilai Barang (satuan)
            ws.column_dimensions['Q'].width = 20  # Total Nilai Barang

            wb.save(output_path)
            logger.info(f"✅ Batch Faktur Pajak Excel export created: {output_path} with {len(results)} documents")
            return True
            
        except Exception as e:
            logger.error(f"❌ Batch Faktur Pajak Excel export failed: {e}", exc_info=True)
            return False
    
    def batch_export_to_pdf(self, batch_id: str, results: list, output_path: str) -> bool:
        """Export multiple Faktur Pajak documents to single PDF file with consolidated table"""
        try:
            if not HAS_REPORTLAB:
                logger.error("❌ reportlab not available for batch PDF export")
                return False

            doc = SimpleDocTemplate(output_path, pagesize=A4,
                                    topMargin=0.5*inch, bottomMargin=0.5*inch,
                                    leftMargin=0.5*inch, rightMargin=0.5*inch)
            styles = getSampleStyleSheet()
            story = []
            
            # Custom title style
            batch_title_style = ParagraphStyle(
                'BatchTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1e40af'),
                spaceAfter=15,
                alignment=1  # Center
            )

            # ===== BATCH TITLE =====
            story.append(Paragraph(f"📦 BATCH: {batch_id}", batch_title_style))
            story.append(Paragraph(f"Total: {len(results)} Documents | Export: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 
                                  ParagraphStyle('SubTitle', parent=styles['Normal'], fontSize=9, alignment=1, spaceAfter=15)))
            
            # ===== CONSOLIDATED TABLE (All documents in one table) =====
            table_data = [
                ["Nama", "Tgl", "NPWP", "Nomor Faktur", "Alamat", "DPP", "PPN", "Total", "Invoice", "Nama Barang/Jasa"]
            ]
            
            # Extract structured data for all documents
            from ai_processor import IndonesianTaxDocumentParser
            temp_parser = IndonesianTaxDocumentParser()
            
            for result_dict in results:
                # Get structured data
                extracted_data = result_dict.get('extracted_data', {})
                structured = extracted_data.get('structured_data', {})
                raw_text = extracted_data.get('raw_text', '')
                
                # If no structured data, try to extract
                if not structured or not any(structured.values()):
                    if raw_text:
                        structured = temp_parser.extract_structured_fields(raw_text, "faktur_pajak")
                
                structured = self._prepare_structured_fields(structured, raw_text)

                # Add data row with truncation for PDF
                table_data.append([
                    structured.get('nama', 'N/A')[:25] if len(structured.get('nama', 'N/A')) > 25 else structured.get('nama', 'N/A'),
                    structured.get('tanggal', 'N/A'),
                    structured.get('npwp', 'N/A'),
                    structured.get('nomor_faktur', 'N/A'),
                    structured.get('alamat', 'N/A')[:35] if len(structured.get('alamat', 'N/A')) > 35 else structured.get('alamat', 'N/A'),
                    structured.get('dpp', 'N/A'),
                    structured.get('ppn', 'N/A'),
                    structured.get('total', 'N/A'),
                    structured.get('invoice', 'N/A'),
                    (structured.get('nama_barang_jasa', 'N/A')[:60]
                     if structured.get('nama_barang_jasa', 'N/A') != "N/A" else "N/A")
                ])
            
            # Column widths (total = 7.5 inches for A4 with 0.5" margins)
            col_widths = [0.9*inch, 0.6*inch, 0.9*inch, 0.9*inch, 1.3*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.8*inch, 1.4*inch]
            
            main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Build styling with alternating row colors
            style_list = [
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                # Padding
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                # Data rows font
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ]
            
            # Alternating row colors for data rows
            for i in range(1, len(table_data)):
                if i % 2 == 1:
                    style_list.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#dbeafe')))
                else:
                    style_list.append(('BACKGROUND', (0, i), (-1, i), colors.white))
            
            main_table.setStyle(TableStyle(style_list))
            story.append(main_table)
            
            doc.build(story)
            logger.info(f"✅ Batch Faktur Pajak PDF export created: {output_path} with {len(results)} documents")
            return True
            
        except Exception as e:
            logger.error(f"❌ Batch Faktur Pajak PDF export failed: {e}", exc_info=True)
            return False
    
    def export_to_pdf_document(self, result: Dict[str, Any], output_path: str) -> bool:
        """
        Export Faktur Pajak to professional PDF document format (not table)
        Creates 2-page document: Page 1 = Faktur Keluaran, Page 2 = Faktur Masukan
        """
        try:
            if not HAS_REPORTLAB:
                logger.error("❌ reportlab not available for PDF export")
                return False
            
            from reportlab.platypus import PageBreak, Flowable
            from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
            
            doc = SimpleDocTemplate(output_path, pagesize=A4,
                                   topMargin=0.75*inch, bottomMargin=0.75*inch,
                                   leftMargin=0.75*inch, rightMargin=0.75*inch)
            
            story = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1e40af'),
                spaceAfter=6,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            section_style = ParagraphStyle(
                'Section',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#1e40af'),
                spaceAfter=12,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )
            
            label_style = ParagraphStyle(
                'Label',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#374151'),
                fontName='Helvetica-Bold',
                leading=14
            )
            
            value_style = ParagraphStyle(
                'Value',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.black,
                fontName='Helvetica',
                leading=14
            )
            
            # Get data
            extracted_data = result.get('extracted_data', {})
            
            # PRIORITY: Use Smart Mapper data
            smart_mapped = extracted_data.get('smart_mapped', {})
            if smart_mapped:
                logger.info("✅ Using Smart Mapper GPT data for PDF document export")
                structured = self._convert_smart_mapped_to_structured(smart_mapped)
                seller = smart_mapped.get('seller', {})
                buyer = smart_mapped.get('buyer', {})
                invoice = smart_mapped.get('invoice', {})
                financials = smart_mapped.get('financials', {})
                items = smart_mapped.get('items', [])
            else:
                logger.info("⚠️ Smart Mapper not available, using legacy parser")
                structured_data = extracted_data.get('structured_data', {})
                raw_text = extracted_data.get('raw_text', '')
                structured = self._prepare_structured_fields(structured_data, raw_text)
                seller = {'name': structured.get('nama', ''), 'npwp': structured.get('npwp', ''), 'address': structured.get('alamat', '')}
                buyer = {}
                invoice = {'number': structured.get('nomor_faktur', ''), 'issue_date': structured.get('tanggal', '')}
                financials = {'dpp': structured.get('dpp', ''), 'ppn': structured.get('ppn', ''), 'total': structured.get('total', '')}
                items = []
            
            # Date for export
            tgl_rekam = datetime.now().strftime("%d %B %Y")
            
            # ========== HALAMAN 1: FAKTUR PAJAK KELUARAN ==========
            story.append(Paragraph("FAKTUR PAJAK", title_style))
            story.append(Spacer(1, 12))
            story.append(Paragraph("A. Faktur Pajak Keluaran", section_style))
            story.append(Spacer(1, 8))
            
            # Create info table for Keluaran (with Paragraph for long text wrapping)
            keluaran_data = [
                ["Nama Pengusaha Kena Pajak", Paragraph(": " + (seller.get('name', '') or 'N/A'), value_style)],
                ["No Seri Faktur Pajak", Paragraph(": " + (invoice.get('number', '') or 'N/A'), value_style)],
                ["Alamat", Paragraph(": " + (seller.get('address', '') or 'N/A'), value_style)],
                ["NPWP", Paragraph(": " + (seller.get('npwp', '') or 'N/A'), value_style)],
                ["Tanggal Faktur", Paragraph(": " + (invoice.get('issue_date', '') or 'N/A'), value_style)],
                ["", ""],  # Spacer
                ["Keterangan Barang/Jasa", Paragraph(": " + self._format_items_list(items), value_style)],
                ["Quantity", Paragraph(": " + self._format_items_quantity(items), value_style)],
                ["Diskon", Paragraph(": " + (financials.get('discount', '') or '-'), value_style)],
                ["Harga Jual/Penggantian", Paragraph(": " + (financials.get('dpp', '') or 'N/A'), value_style)],
                ["Dasar Pengenaan Pajak (DPP)", Paragraph(": " + (financials.get('dpp', '') or 'N/A'), value_style)],
                ["Jumlah PPN", Paragraph(": " + (financials.get('ppn', '') or 'N/A'), value_style)],
                ["", ""],  # Spacer
                ["Keterangan Lain", Paragraph(": " + (structured.get('invoice', '') or '-'), value_style)],
                ["Tanggal Rekam", Paragraph(": " + tgl_rekam, value_style)],
            ]
            
            keluaran_table = Table(keluaran_data, colWidths=[2.5*inch, 4*inch])
            keluaran_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(keluaran_table)
            
            # Page break
            story.append(PageBreak())
            
            # ========== HALAMAN 2: FAKTUR PAJAK MASUKAN ==========
            story.append(Paragraph("FAKTUR PAJAK", title_style))
            story.append(Spacer(1, 12))
            story.append(Paragraph("B. Faktur Pajak Masukan", section_style))
            story.append(Spacer(1, 8))
            
            # Determine category
            kategori = self._determine_category(items)
            
            masukan_data = [
                ["Nama Pembeli BKP/Penerima JKP", Paragraph(": " + (buyer.get('name', '') or 'N/A'), value_style)],
                ["No Seri Faktur Pajak", Paragraph(": " + (invoice.get('number', '') or 'N/A'), value_style)],
                ["Alamat", Paragraph(": " + (buyer.get('address', '') or 'N/A'), value_style)],
                ["NPWP", Paragraph(": " + (buyer.get('npwp', '') or 'N/A'), value_style)],
                ["Email", Paragraph(": " + (buyer.get('email', '') or '-'), value_style)],
                ["Tanggal Faktur", Paragraph(": " + (invoice.get('issue_date', '') or 'N/A'), value_style)],
                ["", ""],  # Spacer
                ["Keterangan Barang/Jasa", Paragraph(": " + self._format_items_list(items), value_style)],
                ["Quantity", Paragraph(": " + self._format_items_quantity(items), value_style)],
                ["Diskon", Paragraph(": " + (financials.get('discount', '') or '-'), value_style)],
                ["Harga Jual/Penggantian", Paragraph(": " + (financials.get('dpp', '') or 'N/A'), value_style)],
                ["Dasar Pengenaan Pajak (DPP)", Paragraph(": " + (financials.get('dpp', '') or 'N/A'), value_style)],
                ["Jumlah PPN", Paragraph(": " + (financials.get('ppn', '') or 'N/A'), value_style)],
                ["", ""],  # Spacer
                ["Kategori Jasa/Barang", Paragraph(": " + kategori, value_style)],
                ["Keterangan Lain", Paragraph(": " + (structured.get('invoice', '') or '-'), value_style)],
                ["Tanggal Rekam", Paragraph(": " + tgl_rekam, value_style)],
            ]
            
            masukan_table = Table(masukan_data, colWidths=[2.5*inch, 4*inch])
            masukan_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(masukan_table)
            
            # Build PDF
            doc.build(story)
            logger.info(f"✅ Faktur Pajak PDF document created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Faktur Pajak PDF document export failed: {e}", exc_info=True)
            return False
    
    def _format_items_list(self, items: list) -> str:
        """Format items list as comma-separated string"""
        if not items:
            return "N/A"
        descriptions = []
        for item in items:
            if isinstance(item, dict):
                desc = item.get('description', '')
                if desc:
                    descriptions.append(desc)
        return ", ".join(descriptions) if descriptions else "N/A"
    
    def _format_items_quantity(self, items: list) -> str:
        """Format items quantity - sum all quantities"""
        if not items:
            return "N/A"
        
        total_quantity = 0
        has_quantity = False
        
        for item in items:
            if isinstance(item, dict):
                qty_str = item.get('quantity', '')
                if qty_str:
                    has_quantity = True
                    # Try to extract numeric value from quantity string
                    try:
                        # Remove any non-numeric characters except decimal point
                        import re
                        qty_clean = re.sub(r'[^\d.]', '', str(qty_str))
                        if qty_clean:
                            qty_num = float(qty_clean)
                            total_quantity += qty_num
                    except (ValueError, TypeError):
                        # If can't parse, skip this item
                        continue
        
        if not has_quantity:
            return "N/A"
        
        # Return integer if it's a whole number, otherwise return with decimal
        if total_quantity == int(total_quantity):
            return str(int(total_quantity))
        else:
            return f"{total_quantity:.2f}"
    
    def _determine_category(self, items: list) -> str:
        """Determine if items are Barang or Jasa based on descriptions"""
        if not items:
            return "N/A"
        
        # Keywords for Jasa (services)
        jasa_keywords = ['jasa', 'service', 'konsultasi', 'training', 'pelatihan', 'sewa', 'rent']
        
        descriptions = []
        for item in items:
            if isinstance(item, dict):
                desc = item.get('description', '').lower()
                if desc:
                    descriptions.append(desc)
        
        if not descriptions:
            return "N/A"
        
        # Check if any item contains jasa keywords
        has_jasa = any(keyword in desc for desc in descriptions for keyword in jasa_keywords)
        
        if has_jasa:
            return "Jasa (Service)"
        else:
            return "Barang (Goods)"
    
    def export_batch_to_pdf(self, batch_results: list, output_path: str) -> bool:
        """
        Export multiple Faktur Pajak documents to a single PDF with professional template
        Each document gets 2 pages (Keluaran + Masukan)
        """
        if not HAS_REPORTLAB:
            logger.error("❌ reportlab not installed - cannot create batch PDF")
            return False
        
        try:
            from reportlab.platypus import PageBreak
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            
            # Create PDF document
            doc = SimpleDocTemplate(output_path, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1e40af'),
                spaceAfter=6,
                alignment=1  # Center
            )
            section_style = ParagraphStyle(
                'CustomSection',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=colors.HexColor('#1e40af'),
                spaceAfter=6
            )
            value_style = ParagraphStyle(
                'ValueStyle',
                parent=styles['Normal'],
                fontSize=10,
                leading=14,
                wordWrap='LTR'
            )
            
            # Process each document
            for idx, result in enumerate(batch_results):
                # Add separator between documents (except first)
                if idx > 0:
                    story.append(PageBreak())
                
                # Convert smart_mapped data to structured format
                extracted_data = result.get('extracted_data', {})
                if 'smart_mapped' in extracted_data:
                    structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
                else:
                    structured = extracted_data
                
                # Extract data with case-insensitive lookup
                def get_value(data, key, default='N/A'):
                    if not isinstance(data, dict):
                        return default
                    # Try exact key first
                    if key in data:
                        val = data[key]
                        return val if val not in (None, '', 'N/A') else default
                    # Try case-insensitive
                    for k, v in data.items():
                        if k.lower() == key.lower():
                            return v if v not in (None, '', 'N/A') else default
                    return default
                
                # Extract smart_mapped fields
                smart_mapped = extracted_data.get('smart_mapped', {})
                seller = smart_mapped.get('seller', {}) if isinstance(smart_mapped, dict) else {}
                buyer = smart_mapped.get('buyer', {}) if isinstance(smart_mapped, dict) else {}
                invoice = smart_mapped.get('invoice', {}) if isinstance(smart_mapped, dict) else {}
                items = smart_mapped.get('items', []) if isinstance(smart_mapped, dict) else []
                
                # Financials with case-insensitive lookup
                financials_raw = smart_mapped.get('financials', {}) if isinstance(smart_mapped, dict) else {}
                financials = {
                    'dpp': get_value(financials_raw, 'dpp', 'N/A'),
                    'ppn': get_value(financials_raw, 'ppn', 'N/A'),
                    'total': get_value(financials_raw, 'total', 'N/A'),
                    'discount': get_value(financials_raw, 'discount', '-')
                }
                
                # Created date
                created_at = result.get('created_at', '')
                if created_at:
                    try:
                        from datetime import datetime
                        if isinstance(created_at, str):
                            dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        else:
                            dt = created_at
                        tgl_rekam = dt.strftime('%d/%m/%Y %H:%M')
                    except:
                        tgl_rekam = str(created_at)
                else:
                    tgl_rekam = 'N/A'
                
                # Document filename
                filename = result.get('filename', result.get('original_filename', 'Unknown'))
                
                # ========== HALAMAN 1: FAKTUR PAJAK KELUARAN ==========
                story.append(Paragraph("FAKTUR PAJAK", title_style))
                story.append(Spacer(1, 6))
                story.append(Paragraph(f"<i>Document: {filename}</i>", value_style))
                story.append(Spacer(1, 12))
                story.append(Paragraph("A. Faktur Pajak Keluaran", section_style))
                story.append(Spacer(1, 8))
                
                # Create info table for Keluaran
                keluaran_data = [
                    ["Nama Pengusaha Kena Pajak", Paragraph(": " + (seller.get('name', '') or 'N/A'), value_style)],
                    ["No Seri Faktur Pajak", Paragraph(": " + (invoice.get('number', '') or 'N/A'), value_style)],
                    ["Alamat", Paragraph(": " + (seller.get('address', '') or 'N/A'), value_style)],
                    ["NPWP", Paragraph(": " + (seller.get('npwp', '') or 'N/A'), value_style)],
                    ["Tanggal Faktur", Paragraph(": " + (invoice.get('issue_date', '') or 'N/A'), value_style)],
                    ["", ""],
                    ["Keterangan Barang/Jasa", Paragraph(": " + self._format_items_list(items), value_style)],
                    ["Quantity", Paragraph(": " + self._format_items_quantity(items), value_style)],
                    ["Diskon", Paragraph(": " + (financials.get('discount', '') or '-'), value_style)],
                    ["Harga Jual/Penggantian", Paragraph(": " + (financials.get('dpp', '') or 'N/A'), value_style)],
                    ["Dasar Pengenaan Pajak (DPP)", Paragraph(": " + (financials.get('dpp', '') or 'N/A'), value_style)],
                    ["Jumlah PPN", Paragraph(": " + (financials.get('ppn', '') or 'N/A'), value_style)],
                    ["", ""],
                    ["Keterangan Lain", Paragraph(": " + (structured.get('invoice', '') or '-'), value_style)],
                    ["Tanggal Rekam", Paragraph(": " + tgl_rekam, value_style)],
                ]
                
                keluaran_table = Table(keluaran_data, colWidths=[2.5*inch, 4*inch])
                keluaran_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                story.append(keluaran_table)
                
                # Page break between Keluaran and Masukan
                story.append(PageBreak())
                
                # ========== HALAMAN 2: FAKTUR PAJAK MASUKAN ==========
                story.append(Paragraph("FAKTUR PAJAK", title_style))
                story.append(Spacer(1, 6))
                story.append(Paragraph(f"<i>Document: {filename}</i>", value_style))
                story.append(Spacer(1, 12))
                story.append(Paragraph("B. Faktur Pajak Masukan", section_style))
                story.append(Spacer(1, 8))
                
                # Determine category
                kategori = self._determine_category(items)
                
                masukan_data = [
                    ["Nama Pembeli BKP/Penerima JKP", Paragraph(": " + (buyer.get('name', '') or 'N/A'), value_style)],
                    ["No Seri Faktur Pajak", Paragraph(": " + (invoice.get('number', '') or 'N/A'), value_style)],
                    ["Alamat", Paragraph(": " + (buyer.get('address', '') or 'N/A'), value_style)],
                    ["NPWP", Paragraph(": " + (buyer.get('npwp', '') or 'N/A'), value_style)],
                    ["Email", Paragraph(": " + (buyer.get('email', '') or '-'), value_style)],
                    ["Tanggal Faktur", Paragraph(": " + (invoice.get('issue_date', '') or 'N/A'), value_style)],
                    ["", ""],
                    ["Keterangan Barang/Jasa", Paragraph(": " + self._format_items_list(items), value_style)],
                    ["Quantity", Paragraph(": " + self._format_items_quantity(items), value_style)],
                    ["Diskon", Paragraph(": " + (financials.get('discount', '') or '-'), value_style)],
                    ["Harga Jual/Penggantian", Paragraph(": " + (financials.get('dpp', '') or 'N/A'), value_style)],
                    ["Dasar Pengenaan Pajak (DPP)", Paragraph(": " + (financials.get('dpp', '') or 'N/A'), value_style)],
                    ["Jumlah PPN", Paragraph(": " + (financials.get('ppn', '') or 'N/A'), value_style)],
                    ["", ""],
                    ["Kategori Jasa/Barang", Paragraph(": " + kategori, value_style)],
                    ["Keterangan Lain", Paragraph(": " + (structured.get('invoice', '') or '-'), value_style)],
                    ["Tanggal Rekam", Paragraph(": " + tgl_rekam, value_style)],
                ]
                
                masukan_table = Table(masukan_data, colWidths=[2.5*inch, 4*inch])
                masukan_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                story.append(masukan_table)
            
            # Build PDF with all documents
            doc.build(story)
            logger.info(f"✅ Batch Faktur Pajak PDF created: {output_path} ({len(batch_results)} documents)")
            return True
            
        except Exception as e:
            logger.error(f"❌ Batch Faktur Pajak PDF export failed: {e}", exc_info=True)
            return False
