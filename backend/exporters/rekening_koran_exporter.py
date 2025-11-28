"""
Rekening Koran Exporter
Handles Excel and PDF export specifically for Bank Statement (Rekening Koran) documents
"""

from typing import Dict, Any
from datetime import datetime
import logging

from .base_exporter import BaseExporter

logger = logging.getLogger(__name__)

# Check for required libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("⚠️ openpyxl not installed - Excel export disabled")

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    logger.warning("⚠️ reportlab not installed - PDF export disabled")


class RekeningKoranExporter(BaseExporter):
    """Exporter for Bank Statement (Rekening Koran) documents"""

    def __init__(self):
        super().__init__("rekening_koran")
        # Columns specific to Rekening Koran (with Quality indicator)
        self.columns = [
            "Tanggal",
            "Nilai Uang Masuk",
            "Nilai Uang Keluar",
            "Saldo",
            "Sumber Uang Masuk",
            "Tujuan Uang Keluar",
            "Keterangan",
            "Quality"  # NEW: Data quality indicator
        ]

    def _format_rupiah(self, value) -> str:
        """
        Format number to Indonesian Rupiah format without decimals
        Example: 1234567.89 -> "Rp 1.234.567"
        """
        if not value or value in ['-', 'N/A', '', '0', 0, 'None', 'null']:
            return '-'

        try:
            # If already a number (int or float), convert to int first
            if isinstance(value, (int, float)):
                value_int = int(value)
            else:
                # If string, clean and parse
                value_str = str(value).replace('Rp', '').replace('IDR', '').replace('.', '').replace(',', '').strip()
                if not value_str or value_str == '-':
                    return '-'
                value_int = int(float(value_str))

            if value_int == 0:
                return '-'

            # Format with thousand separator
            formatted = f"Rp {value_int:,}"
            # Replace comma with dot for Indonesian format
            formatted = formatted.replace(',', '.')

            return formatted
        except (ValueError, AttributeError):
            logger.warning(f"⚠️ Failed to format rupiah value: {value}")
            return str(value) if value else '-'

    def _format_title_case(self, text: str) -> str:
        """
        Format text to proper title case
        Example: "TRUCKING UNICER S1232" -> "Trucking Unicer S1232"
        """
        if not text or not isinstance(text, str):
            return text

        # Convert to lowercase first, then capitalize each word
        words = text.lower().split()
        formatted_words = []

        for word in words:
            # Keep numbers and codes as-is
            if word.isdigit() or any(char.isdigit() for char in word):
                formatted_words.append(word.upper() if len(word) <= 4 else word.capitalize())
            # Keep short uppercase words (likely abbreviations)
            elif len(word) <= 3:
                formatted_words.append(word.upper())
            else:
                formatted_words.append(word.capitalize())

        return ' '.join(formatted_words)

    def _clean_prefix_keterangan(self, keterangan: str) -> str:
        """
        Remove specific prefixes, reference codes, and amount numbers from keterangan
        Removes:
        - TRSF E-BANKING CR, TRSF E-BANKING DB, BI-FAST DB, BI-FAST CR, etc
        - Reference codes like: 0107/FTSCY/WS95051, 0207/ADSCY/WS95051, etc
        - Amount numbers like: 8800000.00, 277533104.00, 1350000.00, etc
        """
        if not keterangan or not isinstance(keterangan, str):
            return keterangan

        import re

        # Remove specific prefixes (case insensitive)
        prefixes_to_remove = [
            r'^TRSF E-BANKING CR\s*',
            r'^TRSF E-BANKING DB\s*',
            r'^BI-FAST DB\s*',
            r'^BI-FAST CR\s*',
            r'^BYR VIA E-BANKING\s*',
            r'^KR OTOMATIS\s*',
        ]

        cleaned = keterangan
        for pattern in prefixes_to_remove:
            cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)

        # Remove reference codes pattern: DDMM/XXXXX/WSXXXXX or similar
        # Pattern: 4 digits / 5-6 letters / 5-8 alphanumeric
        # Examples: 0107/FTSCY/WS95051, 1407/ADSCY/WS95051, 2807/FTLLG/0000100
        reference_patterns = [
            r'\b\d{4}/[A-Z]{5,6}/[A-Z0-9]{5,8}\b\s*',  # 0107/FTSCY/WS95051
            r'\b\d{4}/[A-Z]{5,6}/\d+\b\s*',             # 1407/ADSCY/0000100
            r'\b\d{4}/[A-Z]+/[A-Z0-9]+\b\s*',           # Generic pattern
        ]

        for pattern in reference_patterns:
            cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)

        # Remove amount numbers (standalone numbers with optional decimals)
        # Pattern: Numbers like 8800000.00, 277533104.00, 1350000, etc
        # Match numbers with 4+ digits (to avoid removing dates, codes, etc)
        amount_patterns = [
            r'\b\d{4,}\.?\d{0,2}\b\s*',  # Match: 8800000.00, 277533104, 1350000
            r'\b\d{1,3}(?:,\d{3})+(?:\.\d{2})?\b\s*',  # Match: 1,350,000.00 (with commas)
        ]

        for pattern in amount_patterns:
            cleaned = re.sub(pattern, '', cleaned)

        # Remove multiple spaces and trim
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()

        # If cleaning removed everything, return original
        if not cleaned:
            return keterangan

        return cleaned

    def _extract_year_from_periode(self, periode: str) -> str:
        """
        Extract year from periode string.
        Examples:
        - "Januari 2025" → "2025"
        - "01/2025" → "2025"
        - "01/01/2025 - 31/01/2025" → "2025"
        - "Jan 2025" → "2025"
        """
        if not periode or not isinstance(periode, str):
            return None

        import re
        # Look for 4-digit year
        year_match = re.search(r'\b(20\d{2})\b', periode)
        if year_match:
            return year_match.group(1)

        return None

    def _complete_date_with_year(self, date_str: str, year: str) -> str:
        """
        Complete date and standardize to DD/MM/YYYY format.
        Handles ALL common date formats from Indonesian banks:
        - "01/01" + "2025" → "01/01/2025"
        - "15/03" + "2025" → "15/03/2025"
        - "01 JAN" + "2025" → "01/01/2025"
        - "15 agustus 2025" → "15/08/2025"
        - "15/08/25" → "15/08/2025"
        - "01/01/2025" → "01/01/2025" (no change)
        - "1/3" + "2025" → "01/03/2025" (with zero padding)
        - "2025-08-15" → "15/08/2025" (ISO format)
        """
        if not date_str or not isinstance(date_str, str):
            return date_str

        original_date = date_str
        date_str = date_str.strip()

        import re
        from datetime import datetime

        # Provide default year if not provided
        if not year:
            year = str(datetime.now().year)
            logger.warning(f"⚠️ No year provided for date completion, using current year: {year}")

        # Month name mapping (Indonesian + English, case-insensitive)
        month_map = {
            'JAN': '01', 'JANUARI': '01', 'JANUARY': '01',
            'FEB': '02', 'FEBRUARI': '02', 'FEBRUARY': '02',
            'MAR': '03', 'MARET': '03', 'MARCH': '03',
            'APR': '04', 'APRIL': '04',
            'MAY': '05', 'MEI': '05',
            'JUN': '06', 'JUNI': '06', 'JUNE': '06',
            'JUL': '07', 'JULI': '07', 'JULY': '07',
            'AUG': '08', 'AGUSTUS': '08', 'AUGUST': '08', 'AGT': '08',
            'SEP': '09', 'SEPTEMBER': '09', 'SEPT': '09',
            'OCT': '10', 'OKTOBER': '10', 'OCTOBER': '10', 'OKT': '10',
            'NOV': '11', 'NOVEMBER': '11',
            'DEC': '12', 'DESEMBER': '12', 'DECEMBER': '12', 'DES': '12'
        }

        # Pattern 1: Already complete DD/MM/YYYY
        match = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', date_str)
        if match:
            day = match.group(1).zfill(2)
            month = match.group(2).zfill(2)
            year_full = match.group(3)
            completed = f"{day}/{month}/{year_full}"
            if original_date != completed:
                logger.debug(f"📅 Date standardized: '{original_date}' → '{completed}'")
            return completed

        # Pattern 2: DD/MM/YY (2-digit year) → convert to DD/MM/YYYY
        match = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2})$', date_str)
        if match:
            day = match.group(1).zfill(2)
            month = match.group(2).zfill(2)
            year_short = match.group(3)
            # Assume 20XX for years 00-99
            year_full = f"20{year_short}"
            completed = f"{day}/{month}/{year_full}"
            logger.debug(f"📅 Date converted from 2-digit year: '{original_date}' → '{completed}'")
            return completed

        # Pattern 3: DD/MM (no year) → add year
        match = re.match(r'^(\d{1,2})/(\d{1,2})$', date_str)
        if match:
            day = match.group(1).zfill(2)
            month = match.group(2).zfill(2)
            completed = f"{day}/{month}/{year}"
            logger.debug(f"📅 Date completed: '{original_date}' → '{completed}'")
            return completed

        # Pattern 4: DD-MM-YYYY or DD-MM-YY → convert to DD/MM/YYYY
        match = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{2,4})$', date_str)
        if match:
            day = match.group(1).zfill(2)
            month = match.group(2).zfill(2)
            year_part = match.group(3)
            year_full = f"20{year_part}" if len(year_part) == 2 else year_part
            completed = f"{day}/{month}/{year_full}"
            logger.debug(f"📅 Date converted from dash format: '{original_date}' → '{completed}'")
            return completed

        # Pattern 5: DD-MM (no year, dash separator) → add year and convert to slash
        match = re.match(r'^(\d{1,2})-(\d{1,2})$', date_str)
        if match:
            day = match.group(1).zfill(2)
            month = match.group(2).zfill(2)
            completed = f"{day}/{month}/{year}"
            logger.debug(f"📅 Date completed from dash: '{original_date}' → '{completed}'")
            return completed

        # Pattern 6: ISO format YYYY-MM-DD → DD/MM/YYYY
        match = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', date_str)
        if match:
            year_iso = match.group(1)
            month = match.group(2).zfill(2)
            day = match.group(3).zfill(2)
            completed = f"{day}/{month}/{year_iso}"
            logger.debug(f"📅 Date converted from ISO: '{original_date}' → '{completed}'")
            return completed

        # Pattern 7: "DD MONTH_NAME YYYY" (e.g., "15 agustus 2025", "1 JANUARI 2025")
        match = re.match(r'^(\d{1,2})\s+([A-Z]+)\s+(\d{4})$', date_str.upper())
        if match:
            day = match.group(1).zfill(2)
            month_name = match.group(2)
            year_full = match.group(3)
            month = month_map.get(month_name, '01')
            completed = f"{day}/{month}/{year_full}"
            logger.debug(f"📅 Date converted from full name: '{original_date}' → '{completed}'")
            return completed

        # Pattern 8: "DD MONTH_NAME" (no year, e.g., "15 agustus", "1 JAN")
        match = re.match(r'^(\d{1,2})\s+([A-Z]+)$', date_str.upper())
        if match:
            day = match.group(1).zfill(2)
            month_name = match.group(2)
            month = month_map.get(month_name, '01')
            completed = f"{day}/{month}/{year}"
            logger.debug(f"📅 Date completed from month name: '{original_date}' → '{completed}'")
            return completed

        # Pattern 9: "MONTH_NAME DD, YYYY" (English format, e.g., "August 15, 2025")
        match = re.match(r'^([A-Z]+)\s+(\d{1,2}),?\s+(\d{4})$', date_str.upper())
        if match:
            month_name = match.group(1)
            day = match.group(2).zfill(2)
            year_full = match.group(3)
            month = month_map.get(month_name, '01')
            completed = f"{day}/{month}/{year_full}"
            logger.debug(f"📅 Date converted from English format: '{original_date}' → '{completed}'")
            return completed

        # If no pattern matched, return as-is
        logger.debug(f"⚠️ Date format not recognized, keeping as-is: '{date_str}'")
        return date_str

    def _validate_transaction(self, trans: dict, prev_saldo: float = None) -> dict:
        """
        Validate transaction data quality and return quality score
        
        Returns:
            {
                'score': float (0.0 to 1.0),
                'issues': list of issue codes,
                'label': str ('High', 'Medium', 'Low')
            }
        """
        issues = []
        score = 1.0
        
        # Check 1: Valid date
        tanggal = trans.get('tanggal', '')
        if not tanggal or tanggal == 'N/A':
            issues.append('missing_date')
            score -= 0.3
        else:
            parsed = self._parse_date_for_sorting(tanggal)
            if parsed == (9999, 12, 31):  # Invalid date
                issues.append('invalid_date')
                score -= 0.2
        
        # Check 2: Valid saldo
        saldo_str = trans.get('saldo', '')
        saldo_val = None
        if not saldo_str or saldo_str in ['N/A', '-', '']:
            issues.append('missing_saldo')
            score -= 0.2
        else:
            try:
                saldo_clean = self._fix_misread_amount(saldo_str)
                saldo_val = float(saldo_clean)
                if saldo_val < 0:
                    issues.append('negative_saldo')
                    score -= 0.1
            except (ValueError, TypeError):
                issues.append('invalid_saldo_format')
                score -= 0.2
        
        # Check 3: Kredit/Debet sanity (not both present)
        kredit = trans.get('kredit', '')
        debet = trans.get('debet', '')
        
        if kredit and debet and kredit not in ['-', '', 'N/A', '0', 0] and debet not in ['-', '', 'N/A', '0', 0]:
            issues.append('both_kredit_debet')
            score -= 0.3
        
        # Check 4: At least one of kredit/debet should exist
        if (not kredit or kredit in ['-', '', 'N/A', '0', 0]) and (not debet or debet in ['-', '', 'N/A', '0', 0]):
            issues.append('no_mutation')
            score -= 0.2
        
        # Check 5: Saldo consistency with previous transaction
        if prev_saldo is not None and saldo_val is not None:
            try:
                kredit_val = float(self._fix_misread_amount(kredit)) if kredit and kredit not in ['-', '', 'N/A'] else 0
                debet_val = float(self._fix_misread_amount(debet)) if debet and debet not in ['-', '', 'N/A'] else 0
                
                expected_saldo = prev_saldo + kredit_val - debet_val
                diff = abs(expected_saldo - saldo_val)
                
                # Allow up to 1 Rp rounding error
                if diff > 1:
                    issues.append('saldo_mismatch')
                    score -= 0.15
            except (ValueError, TypeError):
                pass
        
        # Check 6: Keterangan exists
        keterangan = trans.get('keterangan', '')
        if not keterangan or keterangan == 'N/A':
            issues.append('missing_description')
            score -= 0.05
        
        # Ensure score is between 0 and 1
        score = max(0.0, min(1.0, score))
        
        # Determine label
        if score >= 0.8:
            label = '✅ High'
        elif score >= 0.5:
            label = '⚠️ Medium'
        else:
            label = '❌ Low'
        
        return {
            'score': score,
            'issues': issues,
            'label': label
        }
    
    def _remove_duplicates(self, transaksi: list) -> list:
        """
        Remove duplicate transactions based on fingerprint
        
        Args:
            transaksi: List of transaction dicts
            
        Returns:
            List with duplicates removed
        """
        if not transaksi or not isinstance(transaksi, list):
            return transaksi
        
        seen = set()
        unique = []
        duplicates_removed = 0
        
        for trans in transaksi:
            if not isinstance(trans, dict):
                continue
            
            # Create fingerprint: date + kredit + debet + saldo
            tanggal = trans.get('tanggal', 'N/A')
            kredit = self._fix_misread_amount(trans.get('kredit', '')) or '0'
            debet = self._fix_misread_amount(trans.get('debet', '')) or '0'
            saldo = self._fix_misread_amount(trans.get('saldo', '')) or '0'
            
            fingerprint = f"{tanggal}_{kredit}_{debet}_{saldo}"
            
            if fingerprint not in seen:
                seen.add(fingerprint)
                unique.append(trans)
            else:
                duplicates_removed += 1
                logger.warning(f"⚠️ Duplicate transaction removed: {tanggal} | Kredit: {kredit} | Debet: {debet}")
        
        if duplicates_removed > 0:
            logger.info(f"✅ Removed {duplicates_removed} duplicate transactions")
        
        return unique
    
    def _calculate_summary_statistics(self, transaksi: list) -> dict:
        """
        Calculate summary statistics for transactions
        
        Returns:
            {
                'total_transactions': int,
                'total_kredit': float,
                'total_debet': float,
                'net_change': float,
                'avg_quality': float,
                'high_quality_pct': float
            }
        """
        if not transaksi or not isinstance(transaksi, list):
            return {
                'total_transactions': 0,
                'total_kredit': 0,
                'total_debet': 0,
                'net_change': 0,
                'avg_quality': 0,
                'high_quality_pct': 0
            }
        
        total_kredit = 0
        total_debet = 0
        quality_scores = []
        high_quality_count = 0
        
        for trans in transaksi:
            if not isinstance(trans, dict):
                continue
            
            # Sum kredit
            kredit = trans.get('kredit', '')
            if kredit and kredit not in ['-', '', 'N/A', '0', 0]:
                try:
                    kredit_val = float(self._fix_misread_amount(kredit))
                    total_kredit += kredit_val
                except (ValueError, TypeError):
                    pass
            
            # Sum debet
            debet = trans.get('debet', '')
            if debet and debet not in ['-', '', 'N/A', '0', 0]:
                try:
                    debet_val = float(self._fix_misread_amount(debet))
                    total_debet += debet_val
                except (ValueError, TypeError):
                    pass
            
            # Collect quality scores
            quality_info = trans.get('_quality', {})
            if quality_info and 'score' in quality_info:
                quality_scores.append(quality_info['score'])
                if quality_info['score'] >= 0.8:
                    high_quality_count += 1
        
        net_change = total_kredit - total_debet
        avg_quality = sum(quality_scores) / len(quality_scores) if quality_scores else 0
        high_quality_pct = (high_quality_count / len(transaksi) * 100) if transaksi else 0
        
        return {
            'total_transactions': len(transaksi),
            'total_kredit': total_kredit,
            'total_debet': total_debet,
            'net_change': net_change,
            'avg_quality': avg_quality,
            'high_quality_pct': high_quality_pct
        }

    def _parse_date_for_sorting(self, date_str: str) -> tuple:
        """
        Parse date string to tuple (year, month, day) for sorting.
        Handles various formats:
        - "01/01/2025" → (2025, 1, 1)
        - "01/01" → (9999, 1, 1)  # Put incomplete dates at end
        - "01 JAN" → (9999, 1, 1)
        - Invalid → (9999, 12, 31)  # Put invalid at very end

        Returns tuple for easy sorting
        """
        if not date_str or not isinstance(date_str, str):
            return (9999, 12, 31)

        date_str = date_str.strip()

        import re
        from datetime import datetime

        # Try DD/MM/YYYY format
        match = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', date_str)
        if match:
            try:
                day = int(match.group(1))
                month = int(match.group(2))
                year = int(match.group(3))
                return (year, month, day)
            except:
                pass

        # Try DD/MM format (incomplete - no year)
        match = re.match(r'^(\d{1,2})/(\d{1,2})$', date_str)
        if match:
            try:
                day = int(match.group(1))
                month = int(match.group(2))
                # Return with placeholder year 9999 to put at end
                # But keep month/day for sorting within incomplete dates
                return (9999, month, day)
            except:
                pass

        # Try DD-MM format
        match = re.match(r'^(\d{1,2})-(\d{1,2})$', date_str)
        if match:
            try:
                day = int(match.group(1))
                month = int(match.group(2))
                return (9999, month, day)
            except:
                pass

        # Try DD MMM or DD MONTH format
        month_map = {
            'JAN': 1, 'JANUARI': 1, 'JANUARY': 1,
            'FEB': 2, 'FEBRUARI': 2, 'FEBRUARY': 2,
            'MAR': 3, 'MARET': 3, 'MARCH': 3,
            'APR': 4, 'APRIL': 4,
            'MAY': 5, 'MEI': 5,
            'JUN': 6, 'JUNI': 6, 'JUNE': 6,
            'JUL': 7, 'JULI': 7, 'JULY': 7,
            'AUG': 8, 'AGUSTUS': 8, 'AUGUST': 8,
            'SEP': 9, 'SEPTEMBER': 9,
            'OCT': 10, 'OKTOBER': 10, 'OCTOBER': 10,
            'NOV': 11, 'NOVEMBER': 11,
            'DEC': 12, 'DESEMBER': 12, 'DECEMBER': 12
        }

        match = re.match(r'^(\d{1,2})\s+([A-Z]+)(?:\s+(\d{4}))?$', date_str.upper())
        if match:
            try:
                day = int(match.group(1))
                month_name = match.group(2)
                year_str = match.group(3)

                month = month_map.get(month_name, 1)
                year = int(year_str) if year_str else 9999

                return (year, month, day)
            except:
                pass

        # Default: invalid date - put at very end
        return (9999, 12, 31)

    def _sort_transactions_by_month(self, transaksi: list) -> list:
        """
        Sort transactions by date, prioritizing complete dates first,
        then sorting by month (lowest first).

        Args:
            transaksi: List of transaction dicts with 'tanggal' field

        Returns:
            Sorted list of transactions
        """
        if not transaksi or not isinstance(transaksi, list):
            return transaksi

        logger.info(f"📅 Sorting {len(transaksi)} transactions by date...")

        # Create list of (transaction, sort_key) tuples
        trans_with_keys = []
        for trans in transaksi:
            if not isinstance(trans, dict):
                continue

            tanggal = trans.get('tanggal', '')
            sort_key = self._parse_date_for_sorting(tanggal)
            trans_with_keys.append((trans, sort_key))

        # Sort by the tuple (year, month, day)
        trans_with_keys.sort(key=lambda x: x[1])

        # Extract sorted transactions
        sorted_transaksi = [trans for trans, _ in trans_with_keys]

        # Log sample of sorted dates
        if sorted_transaksi:
            sample_dates = [t.get('tanggal', 'N/A') for t in sorted_transaksi[:5]]
            logger.info(f"✅ Transactions sorted. First 5 dates: {sample_dates}")

        return sorted_transaksi

    def _fix_misread_amount(self, value: str) -> str:
        """
        Fix amounts misread by GPT when it confuses decimal separators.

        Example: GPT returns "953,628,332.69" (US format with decimals)
        Solution: Remove currency, then split by '.' and take only integer part

        Logic:
        - If has comma AND dot: "953,628,332.69" → remove decimals → "953,628,332"
        - Return cleaned number (remove Rp, commas, etc. for storage)
        """
        if not value or not isinstance(value, str):
            return value

        logger.info(f"🔍 _fix_misread_amount INPUT: '{value}' (type: {type(value).__name__})")
        original = value
        # Remove currency symbols first
        value = value.replace('Rp', '').replace('IDR', '').strip()

        # If has both comma and dot (US format with decimals like "953,628,332.69")
        # Split by dot and take only the integer part (before dot)
        if ',' in value and '.' in value:
            # Check which comes last (dot should be last for decimals)
            last_comma = value.rfind(',')
            last_dot = value.rfind('.')

            if last_dot > last_comma:
                # Dot is decimal separator, remove everything after it
                value = value.split('.')[0]  # "953,628,332.69" → "953,628,332"
                logger.info(f"🔧 Removed decimals: {original} → {value}")

        # Now clean for storage (remove commas and dots for pure number)
        cleaned = value.replace('.', '').replace(',', '').strip()

        logger.info(f"🔍 _fix_misread_amount OUTPUT: '{cleaned}'")
        return cleaned if cleaned else value

    def _convert_smart_mapped_to_structured(self, smart_mapped: dict) -> dict:
        """Convert Smart Mapper output to structured format for Rekening Koran"""
        if not smart_mapped or not isinstance(smart_mapped, dict):
            return {}

        structured = {}

        # Extract bank info (with more field variations)
        bank_info = smart_mapped.get('bank_info', {}) or {}
        structured['nama_bank'] = bank_info.get('nama_bank') or bank_info.get('bank_name') or 'N/A'
        structured['nomor_rekening'] = bank_info.get('nomor_rekening') or bank_info.get('account_number') or 'N/A'
        structured['nama_pemilik'] = bank_info.get('nama_pemilik') or bank_info.get('account_holder') or bank_info.get('nama') or 'N/A'
        structured['periode'] = bank_info.get('periode') or bank_info.get('period') or 'N/A'
        structured['jenis_rekening'] = bank_info.get('jenis_rekening') or bank_info.get('account_type') or ''
        structured['cabang'] = bank_info.get('cabang') or bank_info.get('branch') or ''
        structured['alamat'] = bank_info.get('alamat') or bank_info.get('address') or ''

        # Extract year from periode for completing dates
        year = self._extract_year_from_periode(structured['periode'])
        if year:
            logger.info(f"✅ Extracted year from periode: {year}")

        # Extract saldo info (with misread amount fix)
        saldo_info = smart_mapped.get('saldo_info', {}) or {}
        structured['saldo_awal'] = self._fix_misread_amount(saldo_info.get('saldo_awal') or saldo_info.get('opening_balance') or '')
        structured['saldo_akhir'] = self._fix_misread_amount(saldo_info.get('saldo_akhir') or saldo_info.get('closing_balance') or saldo_info.get('ending_balance') or '')
        structured['saldo'] = structured['saldo_akhir']  # Alias
        structured['total_kredit'] = self._fix_misread_amount(saldo_info.get('total_kredit') or saldo_info.get('total_credit') or '')
        structured['total_debet'] = self._fix_misread_amount(saldo_info.get('total_debet') or saldo_info.get('total_debit') or '')

        # Extract transactions with support for 'mutasi' field
        transactions = smart_mapped.get('transactions', []) or smart_mapped.get('transaksi', [])
        if transactions and isinstance(transactions, list):
            structured['transaksi'] = []
            for trans in transactions:
                if isinstance(trans, dict):
                    # Handle 'mutasi' field (some banks use single column with +/-)
                    mutasi = trans.get('mutasi') or trans.get('mutation') or ''
                    kredit = trans.get('kredit') or trans.get('credit') or ''
                    debet = trans.get('debet') or trans.get('debit') or ''

                    # If mutasi exists and kredit/debet are empty, parse mutasi
                    if mutasi and not kredit and not debet:
                        # Remove currency symbols and clean
                        mutasi_clean = str(mutasi).replace('Rp', '').replace('IDR', '').replace(',', '').replace('.', '').strip()
                        if mutasi_clean.startswith('+'):
                            kredit = mutasi_clean.replace('+', '')
                        elif mutasi_clean.startswith('-'):
                            debet = mutasi_clean.replace('-', '')
                        elif mutasi_clean:
                            # Try to determine from value (negative = debet)
                            try:
                                val = float(mutasi_clean)
                                if val > 0:
                                    kredit = mutasi_clean
                                else:
                                    debet = str(abs(val))
                            except:
                                # If parsing fails, store as-is in kredit
                                kredit = mutasi

                    # Get transaction date and complete with year if needed
                    tanggal_raw = trans.get('tanggal') or trans.get('date') or trans.get('transaction_date') or 'N/A'
                    tanggal_complete = self._complete_date_with_year(tanggal_raw, year) if year else tanggal_raw

                    # Fix misread amounts in transactions
                    transaction_item = {
                        'tanggal': tanggal_complete,
                        'keterangan': trans.get('keterangan') or trans.get('description') or trans.get('remarks') or trans.get('details') or 'N/A',
                        'kredit': self._fix_misread_amount(kredit) if kredit else '',
                        'debet': self._fix_misread_amount(debet) if debet else '',
                        'saldo': self._fix_misread_amount(trans.get('saldo') or trans.get('balance') or trans.get('running_balance') or ''),
                        'referensi': trans.get('referensi') or trans.get('reference') or trans.get('ref') or '',
                        'cabang': trans.get('cabang') or trans.get('branch') or ''
                    }
                    structured['transaksi'].append(transaction_item)

            # Sort transactions by date (month ascending)
            if structured.get('transaksi'):
                structured['transaksi'] = self._sort_transactions_by_month(structured['transaksi'])
                
                # Remove duplicates
                original_count = len(structured['transaksi'])
                structured['transaksi'] = self._remove_duplicates(structured['transaksi'])
                if len(structured['transaksi']) < original_count:
                    logger.info(f"📊 Removed {original_count - len(structured['transaksi'])} duplicates")
                
                # Validate transactions and add quality scores
                prev_saldo = None
                for trans in structured['transaksi']:
                    # Get current saldo for next iteration
                    try:
                        current_saldo = float(self._fix_misread_amount(trans.get('saldo', 0)))
                    except:
                        current_saldo = None
                    
                    # Validate transaction
                    quality_info = self._validate_transaction(trans, prev_saldo)
                    trans['_quality'] = quality_info
                    
                    # Update prev_saldo for next iteration
                    prev_saldo = current_saldo
                
                # Calculate summary
                summary = self._calculate_summary_statistics(structured['transaksi'])
                logger.info(f"📊 Quality Summary: Avg={summary['avg_quality']:.1%}, High Quality={summary['high_quality_pct']:.1f}%")

        logger.info(f"✅ Rekening Koran Smart Mapper data converted to structured format with {len(structured.get('transaksi', []))} transactions")

        return structured
    
    def export_to_excel(self, result: Dict[str, Any], output_path: str) -> bool:
        """Export single Rekening Koran to Excel with structured table format"""
        try:
            if not HAS_OPENPYXL:
                logger.error("❌ openpyxl not available for Excel export")
                return False
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Rekening Koran"
            
            self._populate_excel_sheet(ws, result)
            
            wb.save(output_path)
            logger.info(f"✅ Rekening Koran Excel export created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Rekening Koran Excel export failed: {e}", exc_info=True)
            return False
    
    def _populate_excel_sheet(self, ws, result: Dict[str, Any]):
        """Helper to populate worksheet with structured Rekening Koran data"""
        # Define styles
        header_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")  # Purple
        header_font = Font(bold=True, size=11, color="FFFFFF")
        data_fill = PatternFill(start_color="ede9fe", end_color="ede9fe", fill_type="solid")  # Light purple
        border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=False)

        # Get structured data with Smart Mapper priority
        extracted_data = result.get('extracted_data', {})

        # DEBUG: Log what keys are available
        logger.info(f"🔍 Excel Export - Available keys in extracted_data: {list(extracted_data.keys())}")
        logger.info(f"🔍 Excel Export - Has smart_mapped: {'smart_mapped' in extracted_data}")
        logger.info(f"🔍 Excel Export - Has structured_data: {'structured_data' in extracted_data}")

        if 'smart_mapped' in extracted_data:
            smart_mapped_data = extracted_data['smart_mapped']
            logger.info(f"🔍 Excel Export - smart_mapped type: {type(smart_mapped_data)}")
            if isinstance(smart_mapped_data, dict):
                logger.info(f"🔍 Excel Export - smart_mapped keys: {list(smart_mapped_data.keys())}")
                if 'transactions' in smart_mapped_data:
                    logger.info(f"🔍 Excel Export - smart_mapped transactions count: {len(smart_mapped_data.get('transactions', []))}")

        # Priority: smart_mapped > flat structure > structured_data > extracted_data itself
        if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
            structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
            logger.info("✅ Using smart_mapped data for Rekening Koran Excel")
        elif 'transactions' in extracted_data and isinstance(extracted_data.get('transactions'), list):
            # Handle flat structure from enhanced_bank_processor
            # Extract fields from nested bank_info and saldo_info to root level
            structured = extracted_data.copy()
            bank_info = extracted_data.get('bank_info', {})
            saldo_info = extracted_data.get('saldo_info', {})

            # Flatten bank_info fields to root level
            structured['nama_bank'] = bank_info.get('nama_bank', '') or bank_info.get('bank_name', '')
            structured['nomor_rekening'] = bank_info.get('nomor_rekening', '') or bank_info.get('account_number', '')
            structured['nama_pemilik'] = bank_info.get('nama_pemilik', '') or bank_info.get('account_holder', '')
            structured['periode'] = bank_info.get('periode', '') or bank_info.get('period', '')
            structured['jenis_rekening'] = bank_info.get('jenis_rekening', '') or bank_info.get('account_type', '')
            structured['cabang'] = bank_info.get('cabang', '') or bank_info.get('branch', '')
            structured['alamat'] = bank_info.get('alamat', '') or bank_info.get('address', '')

            # Flatten saldo_info fields to root level
            structured['saldo_awal'] = saldo_info.get('saldo_awal', '') or saldo_info.get('opening_balance', '')
            structured['saldo_akhir'] = saldo_info.get('saldo_akhir', '') or saldo_info.get('closing_balance', '') or saldo_info.get('ending_balance', '')
            structured['saldo'] = structured['saldo_akhir']  # Alias
            structured['total_kredit'] = saldo_info.get('total_kredit', '') or saldo_info.get('total_credit', '')
            structured['total_debet'] = saldo_info.get('total_debet', '') or saldo_info.get('total_debit', '')
            structured['mata_uang'] = saldo_info.get('mata_uang', '') or saldo_info.get('currency', '')

            # Use 'transactions' as 'transaksi' for compatibility
            structured['transaksi'] = extracted_data.get('transactions', [])

            logger.info("✅ Using flat structure (enhanced processor) for Rekening Koran Excel")
            logger.info(f"   📝 Flattened fields: nama_bank='{structured.get('nama_bank')}', transactions={len(structured.get('transaksi', []))}")
        elif 'structured_data' in extracted_data:
            structured = extracted_data['structured_data']
            logger.info("✅ Using structured_data for Rekening Koran Excel")
        else:
            structured = extracted_data
            logger.info("⚠️ Using raw extracted_data for Rekening Koran Excel")

        # DEBUG: Log structured data after conversion
        logger.info(f"🔍 Excel Export - After conversion, transaksi count: {len(structured.get('transaksi', []))}")
        logger.info(f"🔍 Excel Export - saldo_awal: {structured.get('saldo_awal', 'NOT FOUND')}")
        logger.info(f"🔍 Excel Export - saldo_akhir: {structured.get('saldo_akhir', 'NOT FOUND')}")
        logger.info(f"🔍 Excel Export - saldo: {structured.get('saldo', 'NOT FOUND')}")

        row = 1

        # ===== TITLE =====
        ws.merge_cells(f'A{row}:G{row}')
        nama_bank = structured.get('nama_bank', 'N/A')
        ws[f'A{row}'] = f"🏦 REKENING KORAN - {nama_bank.upper()}"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_align
        ws[f'A{row}'].border = border_thin
        row += 1

        # ===== BANK INFO =====
        info_fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")

        # Row 1: Nomor Rekening & Nama Pemilik
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = f"Nomor Rekening: {structured.get('nomor_rekening', 'N/A')}"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = info_fill
        ws[f'A{row}'].alignment = left_align
        ws[f'A{row}'].border = border_thin

        ws.merge_cells(f'D{row}:G{row}')
        ws[f'D{row}'] = f"Nama: {structured.get('nama_pemilik', 'N/A')}"
        ws[f'D{row}'].font = Font(bold=True, size=10)
        ws[f'D{row}'].fill = info_fill
        ws[f'D{row}'].alignment = left_align
        ws[f'D{row}'].border = border_thin
        row += 1

        # Row 2: Periode & Saldo
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = f"Periode: {structured.get('periode', 'N/A')}"
        ws[f'A{row}'].font = Font(size=10)
        ws[f'A{row}'].fill = info_fill
        ws[f'A{row}'].alignment = left_align
        ws[f'A{row}'].border = border_thin

        saldo_akhir = structured.get('saldo_akhir', structured.get('saldo', 'N/A'))
        ws.merge_cells(f'D{row}:G{row}')
        ws[f'D{row}'] = f"Saldo Akhir: {self._format_rupiah(saldo_akhir)}"
        ws[f'D{row}'].font = Font(bold=True, size=10, color="7c3aed")
        ws[f'D{row}'].fill = info_fill
        ws[f'D{row}'].alignment = left_align
        ws[f'D{row}'].border = border_thin
        row += 1

        # ===== TABLE HEADERS =====
        for col_idx, header in enumerate(self.columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_thin
        row += 1

        # ===== DATA ROWS =====
        # Handle array of transactions if exists
        # Support both 'transaksi' (old format) and 'transactions' (enhanced processor format)
        transaksi = structured.get('transaksi', []) or structured.get('transactions', [])

        # ✅ FIX: Extract year from periode for date completion (same as batch export)
        periode = structured.get('periode', '')
        year_for_completion = self._extract_year_from_periode(periode)
        if not year_for_completion:
            year_for_completion = str(datetime.now().year)
        logger.info(f"📅 Using year {year_for_completion} for date completion in single export")

        if isinstance(transaksi, list) and transaksi:
            # Multiple transactions - iterate through array
            logger.info(f"📊 Processing {len(transaksi)} transactions for Excel export")

            for idx, trans in enumerate(transaksi):
                if not isinstance(trans, dict):
                    logger.warning(f"⚠️ Transaction {idx} is not a dict, skipping")
                    continue

                # Get kredit/debet - handle multiple field names and mutasi
                kredit = trans.get('kredit', trans.get('credit', ''))
                debet = trans.get('debet', trans.get('debit', ''))

                # Handle 'mutasi' field (some banks use single column with +/-)
                # This should already be handled by _convert_smart_mapped_to_structured
                # but add fallback just in case
                if not kredit and not debet:
                    mutasi = trans.get('mutasi', trans.get('mutation', ''))
                    if mutasi:
                        mutasi_str = str(mutasi).strip()
                        # Remove Rp, IDR, etc
                        mutasi_clean = mutasi_str.replace('Rp', '').replace('IDR', '').replace('.', '').replace(',', '').strip()
                        if mutasi_clean.startswith('+') or (mutasi_clean and not mutasi_clean.startswith('-')):
                            kredit = mutasi_clean.replace('+', '')
                        elif mutasi_clean.startswith('-'):
                            debet = mutasi_clean.replace('-', '')

                # Get other fields with fallbacks
                tanggal_raw = trans.get('tanggal', trans.get('date', trans.get('transaction_date', 'N/A')))
                # ✅ FIX: Complete date with year (same as batch export)
                tanggal = self._complete_date_with_year(tanggal_raw, year_for_completion)
                keterangan = trans.get('keterangan', trans.get('description', trans.get('remarks', trans.get('details', 'N/A'))))
                saldo = trans.get('saldo', trans.get('balance', trans.get('running_balance', 'N/A')))

                # Format kredit, debet, saldo to rupiah (no decimals)
                kredit_formatted = self._format_rupiah(kredit)
                debet_formatted = self._format_rupiah(debet)
                saldo_formatted = self._format_rupiah(saldo)

                # Clean and format keterangan: remove prefixes and apply title case
                keterangan_cleaned = self._clean_prefix_keterangan(keterangan)
                keterangan_formatted = self._format_title_case(keterangan_cleaned)

                # Determine sumber/tujuan based on transaction type
                sumber_masuk = keterangan_formatted if kredit_formatted != '-' else '-'
                tujuan_keluar = keterangan_formatted if debet_formatted != '-' else '-'
                
                # Get quality indicator
                quality_info = trans.get('_quality', {})
                quality_label = quality_info.get('label', '⚠️ Medium')

                data_row = [
                    tanggal,
                    kredit_formatted,
                    debet_formatted,
                    saldo_formatted,
                    sumber_masuk,
                    tujuan_keluar,
                    keterangan_formatted,
                    quality_label  # NEW: Quality indicator
                ]

                for col_idx, value in enumerate(data_row, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.fill = data_fill
                    # Right align for numeric columns (Nilai Uang Masuk, Keluar, Saldo)
                    if col_idx in [2, 3, 4]:
                        cell.alignment = right_align
                    else:
                        cell.alignment = left_align
                    cell.border = border_thin
                    
                    # Color code quality column
                    if col_idx == 8:  # Quality column
                        cell.alignment = center_align
                        if '✅' in str(value):
                            cell.font = Font(size=10, bold=True, color="70AD47")  # Green
                        elif '⚠️' in str(value):
                            cell.font = Font(size=10, bold=True, color="FFC000")  # Orange
                        else:
                            cell.font = Font(size=10, bold=True, color="C00000")  # Red
                    else:
                        cell.font = Font(size=10)

                row += 1

            logger.info(f"✅ Exported {len(transaksi)} transactions to Excel")
            
            # Add summary statistics section
            row += 2  # Add spacing
            
            # Summary header
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = "📊 RINGKASAN STATISTIK"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = center_align
            ws[f'A{row}'].border = border_thin
            row += 1
            
            # Calculate summary
            summary = self._calculate_summary_statistics(transaksi)
            
            # Summary data
            summary_fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
            
            summary_items = [
                ("Total Transaksi:", f"{summary['total_transactions']} transaksi"),
                ("Total Uang Masuk:", self._format_rupiah(summary['total_kredit'])),
                ("Total Uang Keluar:", self._format_rupiah(summary['total_debet'])),
                ("Net Change:", self._format_rupiah(summary['net_change'])),
                ("Kualitas Data Rata-rata:", f"{summary['avg_quality']:.1%}"),
                ("Transaksi Berkualitas Tinggi:", f"{summary['high_quality_pct']:.1f}%")
            ]
            
            for label, value in summary_items:
                ws.merge_cells(f'A{row}:C{row}')
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True, size=10)
                ws[f'A{row}'].fill = summary_fill
                ws[f'A{row}'].alignment = left_align
                ws[f'A{row}'].border = border_thin
                
                ws.merge_cells(f'D{row}:H{row}')
                ws[f'D{row}'] = value
                ws[f'D{row}'].font = Font(size=10, bold=True if 'Total' in label else False)
                ws[f'D{row}'].fill = summary_fill
                ws[f'D{row}'].alignment = left_align
                ws[f'D{row}'].border = border_thin
                row += 1

        elif structured.get('transaksi') is not None and not isinstance(structured.get('transaksi'), list):
            logger.warning(f"⚠️ 'transaksi' field exists but is not a list: {type(structured.get('transaksi'))}")
        else:
            # Single transaction - legacy support
            kredit = structured.get('kredit', structured.get('credit', 'N/A'))
            debet = structured.get('debet', structured.get('debit', 'N/A'))
            keterangan = structured.get('keterangan', structured.get('description', 'N/A'))
            saldo = structured.get('saldo', 'N/A')

            # ✅ FIX: Complete date with year for legacy single transaction
            tanggal_raw = structured.get('tanggal', 'N/A')
            tanggal = self._complete_date_with_year(tanggal_raw, year_for_completion)

            # Format to rupiah
            kredit_formatted = self._format_rupiah(kredit)
            debet_formatted = self._format_rupiah(debet)
            saldo_formatted = self._format_rupiah(saldo)

            # Clean and format keterangan
            keterangan_cleaned = self._clean_prefix_keterangan(keterangan)
            keterangan_formatted = self._format_title_case(keterangan_cleaned)

            sumber_masuk = keterangan_formatted if kredit_formatted != '-' else '-'
            tujuan_keluar = keterangan_formatted if debet_formatted != '-' else '-'
            
            # Validate single transaction
            single_trans = {
                'tanggal': tanggal,
                'kredit': kredit,
                'debet': debet,
                'saldo': saldo,
                'keterangan': keterangan
            }
            quality_info = self._validate_transaction(single_trans)
            quality_label = quality_info.get('label', '⚠️ Medium')

            data_row = [
                tanggal,  # ✅ Use completed date instead of raw
                kredit_formatted,
                debet_formatted,
                saldo_formatted,
                sumber_masuk,
                tujuan_keluar,
                keterangan_formatted,
                quality_label
            ]

            for col_idx, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.fill = data_fill
                if col_idx in [2, 3, 4]:
                    cell.alignment = right_align
                elif col_idx == 8:  # Quality column
                    cell.alignment = center_align
                    if '✅' in str(value):
                        cell.font = Font(size=10, bold=True, color="70AD47")
                    elif '⚠️' in str(value):
                        cell.font = Font(size=10, bold=True, color="FFC000")
                    else:
                        cell.font = Font(size=10, bold=True, color="C00000")
                else:
                    cell.alignment = left_align
                cell.border = border_thin
                if col_idx != 8:
                    cell.font = Font(size=10)

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 12  # Tanggal
        ws.column_dimensions['B'].width = 18  # Nilai Uang Masuk
        ws.column_dimensions['C'].width = 18  # Nilai Uang Keluar
        ws.column_dimensions['D'].width = 18  # Saldo
        ws.column_dimensions['E'].width = 30  # Sumber Uang Masuk
        ws.column_dimensions['F'].width = 30  # Tujuan Uang Keluar
        ws.column_dimensions['G'].width = 30  # Keterangan
        ws.column_dimensions['H'].width = 14  # Quality (NEW)
        
        ws.row_dimensions[3].height = 40
    
    def export_to_pdf(self, result: Dict[str, Any], output_path: str) -> bool:
        """Export single Rekening Koran to formatted PDF (simple format, no table)"""
        try:
            if not HAS_REPORTLAB:
                logger.error("❌ reportlab not available for PDF export")
                return False

            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import Paragraph, Spacer
            from reportlab.lib.enums import TA_CENTER
            from datetime import datetime, timezone

            # Get structured data with Smart Mapper priority
            extracted_data = result.get('extracted_data', {})

            # Priority: smart_mapped > flat structure > structured_data > extracted_data itself
            if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
                structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
                logger.info("✅ Using smart_mapped data for Rekening Koran PDF")
            elif 'transactions' in extracted_data and isinstance(extracted_data.get('transactions'), list):
                # Handle flat structure from enhanced_bank_processor
                # Extract fields from nested bank_info and saldo_info to root level
                structured = extracted_data.copy()
                bank_info = extracted_data.get('bank_info', {})
                saldo_info = extracted_data.get('saldo_info', {})

                # Flatten bank_info fields to root level
                structured['nama_bank'] = bank_info.get('nama_bank', '') or bank_info.get('bank_name', '')
                structured['nomor_rekening'] = bank_info.get('nomor_rekening', '') or bank_info.get('account_number', '')
                structured['nama_pemilik'] = bank_info.get('nama_pemilik', '') or bank_info.get('account_holder', '')
                structured['periode'] = bank_info.get('periode', '') or bank_info.get('period', '')
                structured['jenis_rekening'] = bank_info.get('jenis_rekening', '') or bank_info.get('account_type', '')
                structured['cabang'] = bank_info.get('cabang', '') or bank_info.get('branch', '')
                structured['alamat'] = bank_info.get('alamat', '') or bank_info.get('address', '')

                # Flatten saldo_info fields to root level
                structured['saldo_awal'] = saldo_info.get('saldo_awal', '') or saldo_info.get('opening_balance', '')
                structured['saldo_akhir'] = saldo_info.get('saldo_akhir', '') or saldo_info.get('closing_balance', '') or saldo_info.get('ending_balance', '')
                structured['saldo'] = structured['saldo_akhir']  # Alias
                structured['total_kredit'] = saldo_info.get('total_kredit', '') or saldo_info.get('total_credit', '')
                structured['total_debet'] = saldo_info.get('total_debet', '') or saldo_info.get('total_debit', '')
                structured['mata_uang'] = saldo_info.get('mata_uang', '') or saldo_info.get('currency', '')

                # Use 'transactions' as 'transaksi' for compatibility
                structured['transaksi'] = extracted_data.get('transactions', [])

                logger.info("✅ Using flat structure (enhanced processor) for Rekening Koran PDF")
                logger.info(f"   📝 Flattened fields: nama_bank='{structured.get('nama_bank')}', transactions={len(structured.get('transaksi', []))}")
            elif 'structured_data' in extracted_data:
                structured = extracted_data['structured_data']
                logger.info("✅ Using structured_data for Rekening Koran PDF")
            else:
                structured = extracted_data
                logger.info("⚠️ Using raw extracted_data for Rekening Koran PDF")

            # Create PDF document
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=20*mm,
                leftMargin=20*mm,
                topMargin=15*mm,
                bottomMargin=15*mm
            )

            story = []
            styles = getSampleStyleSheet()

            # Title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=4*mm,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            # Section header style
            section_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Heading2'],
                fontSize=11,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=2*mm,
                spaceBefore=6*mm,
                fontName='Helvetica-Bold'
            )

            # Field style
            field_style = ParagraphStyle(
                'Field',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#333333'),
                spaceAfter=2*mm,
                leading=12
            )

            # Highlight style for amounts
            amount_style = ParagraphStyle(
                'Amount',
                parent=field_style,
                fontSize=10,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1F4E78')
            )

            # Add title
            story.append(Paragraph("REKENING KORAN - MUTASI BANK", title_style))
            story.append(Paragraph("Bank Statement Summary",
                                 ParagraphStyle('Subtitle', parent=styles['Normal'],
                                              fontSize=9, alignment=TA_CENTER,
                                              textColor=colors.HexColor('#666666'))))
            story.append(Spacer(1, 10*mm))

            # Section 1: Informasi Rekening
            story.append(Paragraph("INFORMASI REKENING", section_style))
            story.append(Paragraph(f"<b>Nomor Rekening:</b> {structured.get('nomor_rekening', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>Nama Pemilik:</b> {structured.get('nama_pemilik', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>Periode:</b> {structured.get('periode', 'N/A')}", field_style))

            # Section 2: Ringkasan Saldo
            story.append(Paragraph("RINGKASAN SALDO", section_style))
            story.append(Paragraph(f"<b>Saldo Awal:</b> Rp {structured.get('saldo_awal', 'N/A')}", field_style))

            # Calculate totals if transactions available
            # Support both 'transaksi' and 'transactions' field names
            transaksi = structured.get('transaksi', []) or structured.get('transactions', [])
            if transaksi and isinstance(transaksi, list):
                try:
                    total_masuk = 0
                    total_keluar = 0

                    for t in transaksi:
                        # Safely convert kredit to float
                        kredit = t.get('kredit', 0)
                        if kredit and kredit not in ['', 'N/A', '-', 0, '0']:
                            try:
                                kredit_str = str(kredit).replace(',', '').replace('Rp', '').replace('.', '').strip()
                                if kredit_str and kredit_str.replace('.', '').isdigit():
                                    total_masuk += float(kredit_str)
                            except (ValueError, AttributeError):
                                logger.warning(f"⚠️ Failed to convert kredit value: {kredit}")

                        # Safely convert debet to float
                        debet = t.get('debet', 0)
                        if debet and debet not in ['', 'N/A', '-', 0, '0']:
                            try:
                                debet_str = str(debet).replace(',', '').replace('Rp', '').replace('.', '').strip()
                                if debet_str and debet_str.replace('.', '').isdigit():
                                    total_keluar += float(debet_str)
                            except (ValueError, AttributeError):
                                logger.warning(f"⚠️ Failed to convert debet value: {debet}")

                    story.append(Paragraph(f"<b>Total Pemasukan:</b> Rp {total_masuk:,.0f}", field_style))
                    story.append(Paragraph(f"<b>Total Pengeluaran:</b> Rp {total_keluar:,.0f}", field_style))
                except Exception as e:
                    logger.error(f"❌ Failed to calculate totals: {e}")
                    story.append(Paragraph(f"<b>Total Pemasukan:</b> N/A", field_style))
                    story.append(Paragraph(f"<b>Total Pengeluaran:</b> N/A", field_style))

            story.append(Paragraph(f"<b>Saldo Akhir:</b> Rp {structured.get('saldo_akhir', structured.get('saldo', 'N/A'))}", amount_style))

            # Section 3: Transaksi
            if transaksi and isinstance(transaksi, list) and len(transaksi) > 0:
                story.append(Paragraph("TRANSAKSI", section_style))

                for idx, trans in enumerate(transaksi[:20], start=1):  # Limit to 20 transactions
                    story.append(Spacer(1, 3*mm))
                    story.append(Paragraph(f"<b>Transaksi {idx} - {trans.get('tanggal', 'N/A')}</b>", field_style))

                    # Check if it's credit (uang masuk) or debit (uang keluar)
                    kredit = trans.get('kredit', trans.get('credit', ''))
                    debet = trans.get('debet', trans.get('debit', ''))
                    keterangan = trans.get('keterangan', 'N/A')

                    if kredit and kredit not in ['', '0', 0, 'N/A']:
                        story.append(Paragraph(f"<b>Nilai Uang Masuk:</b> Rp {kredit}", amount_style))
                        story.append(Paragraph(f"<b>Sumber:</b> {keterangan}", field_style))
                    elif debet and debet not in ['', '0', 0, 'N/A']:
                        story.append(Paragraph(f"<b>Nilai Uang Keluar:</b> Rp {debet}", amount_style))
                        story.append(Paragraph(f"<b>Tujuan:</b> {keterangan}", field_style))

                    story.append(Paragraph(f"<b>Saldo:</b> Rp {trans.get('saldo', 'N/A')}", field_style))

                    if idx < len(transaksi[:20]):
                        story.append(Spacer(1, 2*mm))
            else:
                # Single transaction format (if not array)
                story.append(Paragraph("TRANSAKSI", section_style))
                story.append(Paragraph(f"<b>Tanggal:</b> {structured.get('tanggal', 'N/A')}", field_style))

                kredit = structured.get('kredit', structured.get('credit', ''))
                debet = structured.get('debet', structured.get('debit', ''))
                keterangan = structured.get('keterangan', 'N/A')

                if kredit and kredit not in ['', '0', 0, 'N/A']:
                    story.append(Paragraph(f"<b>Nilai Uang Masuk:</b> Rp {kredit}", amount_style))
                    story.append(Paragraph(f"<b>Sumber:</b> {keterangan}", field_style))
                elif debet and debet not in ['', '0', 0, 'N/A']:
                    story.append(Paragraph(f"<b>Nilai Uang Keluar:</b> Rp {debet}", amount_style))
                    story.append(Paragraph(f"<b>Tujuan:</b> {keterangan}", field_style))

                story.append(Paragraph(f"<b>Saldo:</b> Rp {structured.get('saldo', 'N/A')}", field_style))

            # Footer
            story.append(Spacer(1, 10*mm))
            footer_text = f"<i>Dokumen ini dibuat secara otomatis oleh Doc-Scan AI System pada {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M:%S UTC')}</i>"
            story.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Normal'],
                                                              fontSize=8, alignment=TA_CENTER,
                                                              textColor=colors.HexColor('#999999'))))

            # Build PDF
            doc.build(story)
            logger.info(f"✅ Rekening Koran PDF export created: {output_path}")
            return True

        except Exception as e:
            logger.error(f"❌ Rekening Koran PDF export failed: {e}", exc_info=True)
            return False
    
    def _extract_year_from_batch(self, results: list) -> str:
        """
        Extract year from batch of results by analyzing all periode fields.
        Returns the most common year found, or current year as fallback.
        """
        from collections import Counter
        from datetime import datetime

        years = []

        for result_dict in results:
            extracted_data = result_dict.get('extracted_data', {})

            # Try smart_mapped first
            if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
                bank_info = extracted_data['smart_mapped'].get('bank_info', {})
                periode = bank_info.get('periode') or bank_info.get('period', '')
            elif 'structured_data' in extracted_data:
                periode = extracted_data['structured_data'].get('periode', '')
            else:
                periode = extracted_data.get('periode', '')

            # Extract year from periode
            if periode:
                year = self._extract_year_from_periode(periode)
                if year:
                    years.append(year)

        if years:
            # Return most common year
            most_common_year = Counter(years).most_common(1)[0][0]
            logger.info(f"✅ Extracted batch year: {most_common_year} (from {len(years)} documents)")
            return most_common_year
        else:
            # Fallback to current year
            current_year = str(datetime.now().year)
            logger.warning(f"⚠️ No year found in batch, using current year: {current_year}")
            return current_year

    def batch_export_to_excel(self, batch_id: str, results: list, output_path: str) -> bool:
        """Export multiple Rekening Koran entries to single Excel file"""
        try:
            if not HAS_OPENPYXL:
                logger.error("❌ openpyxl not available for batch Excel export")
                return False

            wb = Workbook()
            ws = wb.active
            ws.title = "Batch Mutasi Bank"

            # Extract year from batch for completing incomplete dates
            batch_year = self._extract_year_from_batch(results)
            
            # Styles
            header_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
            header_font = Font(bold=True, size=11, color="FFFFFF")
            data_fill_1 = PatternFill(start_color="ede9fe", end_color="ede9fe", fill_type="solid")
            data_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            border_thin = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            right_align = Alignment(horizontal='right', vertical='center', wrap_text=False)
            
            row = 1
            
            # Batch info
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = f"🏦 BATCH MUTASI BANK: {batch_id} | Total: {len(results)} | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = center_align
            ws[f'A{row}'].border = border_thin
            row += 1
            
            # Headers
            for col_idx, header in enumerate(self.columns, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border_thin
            row += 1
            
            # Data rows - handle multiple documents and their transactions
            transaction_row_idx = 0  # For alternating colors

            # Collect ALL transactions from ALL documents for global sorting
            all_transactions = []

            for doc_idx, result_dict in enumerate(results):
                extracted_data = result_dict.get('extracted_data', {})

                # Priority: smart_mapped > flat structure > structured_data > extracted_data itself
                if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
                    structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
                    logger.info(f"✅ Using smart_mapped data for Rekening Koran batch item {doc_idx}")
                elif 'transactions' in extracted_data and isinstance(extracted_data.get('transactions'), list):
                    # Handle flat structure from enhanced_bank_processor
                    # Extract fields from nested bank_info and saldo_info to root level
                    structured = extracted_data.copy()
                    bank_info = extracted_data.get('bank_info', {})
                    saldo_info = extracted_data.get('saldo_info', {})

                    # Flatten bank_info fields to root level
                    structured['nama_bank'] = bank_info.get('nama_bank', '') or bank_info.get('bank_name', '')
                    structured['nomor_rekening'] = bank_info.get('nomor_rekening', '') or bank_info.get('account_number', '')
                    structured['nama_pemilik'] = bank_info.get('nama_pemilik', '') or bank_info.get('account_holder', '')
                    structured['periode'] = bank_info.get('periode', '') or bank_info.get('period', '')
                    structured['jenis_rekening'] = bank_info.get('jenis_rekening', '') or bank_info.get('account_type', '')
                    structured['cabang'] = bank_info.get('cabang', '') or bank_info.get('branch', '')
                    structured['alamat'] = bank_info.get('alamat', '') or bank_info.get('address', '')

                    # Flatten saldo_info fields to root level
                    structured['saldo_awal'] = saldo_info.get('saldo_awal', '') or saldo_info.get('opening_balance', '')
                    structured['saldo_akhir'] = saldo_info.get('saldo_akhir', '') or saldo_info.get('closing_balance', '') or saldo_info.get('ending_balance', '')
                    structured['saldo'] = structured['saldo_akhir']  # Alias
                    structured['total_kredit'] = saldo_info.get('total_kredit', '') or saldo_info.get('total_credit', '')
                    structured['total_debet'] = saldo_info.get('total_debet', '') or saldo_info.get('total_debit', '')
                    structured['mata_uang'] = saldo_info.get('mata_uang', '') or saldo_info.get('currency', '')

                    # Use 'transactions' as 'transaksi' for compatibility
                    structured['transaksi'] = extracted_data.get('transactions', [])

                    logger.info(f"✅ Using flat structure (enhanced processor) for Rekening Koran batch item {doc_idx}")
                    logger.info(f"   📝 Flattened fields: nama_bank='{structured.get('nama_bank')}', transactions={len(structured.get('transaksi', []))}")
                elif 'structured_data' in extracted_data:
                    structured = extracted_data['structured_data']
                    logger.info(f"✅ Using structured_data for Rekening Koran batch item {doc_idx}")
                else:
                    structured = extracted_data
                    logger.info(f"⚠️ Using raw extracted_data for Rekening Koran batch item {doc_idx}")

                # Get transactions array
                # Support both 'transaksi' and 'transactions' field names
                transaksi = structured.get('transaksi', []) or structured.get('transactions', [])

                if isinstance(transaksi, list) and transaksi:
                    # Multiple transactions from this document
                    logger.info(f"📊 Collecting {len(transaksi)} transactions from document {doc_idx+1}")

                    # Complete any incomplete dates with batch_year
                    for trans in transaksi:
                        if isinstance(trans, dict) and trans.get('tanggal'):
                            tanggal_raw = trans['tanggal']
                            # Re-complete date with batch_year (in case it wasn't completed before)
                            tanggal_complete = self._complete_date_with_year(tanggal_raw, batch_year)
                            trans['tanggal'] = tanggal_complete

                    # Add all transactions to global list
                    all_transactions.extend(transaksi)

                else:
                    # Legacy: Single transaction per document
                    tanggal_raw = structured.get('tanggal', 'N/A')
                    tanggal_complete = self._complete_date_with_year(tanggal_raw, batch_year)

                    single_trans = {
                        'tanggal': tanggal_complete,
                        'kredit': structured.get('kredit', structured.get('credit', '')),
                        'debet': structured.get('debet', structured.get('debit', '')),
                        'keterangan': structured.get('keterangan', structured.get('description', 'N/A')),
                        'saldo': structured.get('saldo', 'N/A')
                    }
                    all_transactions.append(single_trans)

            # Sort ALL transactions globally by date
            logger.info(f"📅 Sorting {len(all_transactions)} total transactions globally...")
            all_transactions = self._sort_transactions_by_month(all_transactions)

            # Now render all sorted transactions
            for trans in all_transactions:
                if not isinstance(trans, dict):
                    continue

                # Get kredit/debet with mutasi handling
                kredit = trans.get('kredit', trans.get('credit', ''))
                debet = trans.get('debet', trans.get('debit', ''))

                # Fallback mutasi handling
                if not kredit and not debet:
                    mutasi = trans.get('mutasi', trans.get('mutation', ''))
                    if mutasi:
                        mutasi_str = str(mutasi).strip()
                        mutasi_clean = mutasi_str.replace('Rp', '').replace('IDR', '').replace('.', '').replace(',', '').strip()
                        if mutasi_clean.startswith('+') or (mutasi_clean and not mutasi_clean.startswith('-')):
                            kredit = mutasi_clean.replace('+', '')
                        elif mutasi_clean.startswith('-'):
                            debet = mutasi_clean.replace('-', '')

                # Get other fields
                tanggal = trans.get('tanggal', trans.get('date', 'N/A'))
                keterangan = trans.get('keterangan', trans.get('description', trans.get('remarks', 'N/A')))
                saldo = trans.get('saldo', trans.get('balance', 'N/A'))

                # Format to rupiah
                kredit_formatted = self._format_rupiah(kredit)
                debet_formatted = self._format_rupiah(debet)
                saldo_formatted = self._format_rupiah(saldo)

                # Clean and format keterangan
                keterangan_cleaned = self._clean_prefix_keterangan(keterangan)
                keterangan_formatted = self._format_title_case(keterangan_cleaned)

                # Determine sumber/tujuan
                sumber_masuk = keterangan_formatted if kredit_formatted != '-' else '-'
                tujuan_keluar = keterangan_formatted if debet_formatted != '-' else '-'
                
                # Get quality indicator
                quality_info = trans.get('_quality', {})
                quality_label = quality_info.get('label', '⚠️ Medium')

                data_row = [
                    tanggal,
                    kredit_formatted,
                    debet_formatted,
                    saldo_formatted,
                    sumber_masuk,
                    tujuan_keluar,
                    keterangan_formatted,
                    quality_label  # NEW: Quality indicator
                ]

                fill = data_fill_1 if transaction_row_idx % 2 == 0 else data_fill_2

                for col_idx, value in enumerate(data_row, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.fill = fill
                    # Right align for numeric columns
                    if col_idx in [2, 3, 4]:
                        cell.alignment = right_align
                    elif col_idx == 8:  # Quality column
                        cell.alignment = center_align
                        if '✅' in str(value):
                            cell.font = Font(size=10, bold=True, color="70AD47")
                        elif '⚠️' in str(value):
                            cell.font = Font(size=10, bold=True, color="FFC000")
                        else:
                            cell.font = Font(size=10, bold=True, color="C00000")
                    else:
                        cell.alignment = left_align
                    cell.border = border_thin
                    if col_idx != 8:
                        cell.font = Font(size=10)

                row += 1
                transaction_row_idx += 1

            logger.info(f"✅ Exported {transaction_row_idx} total transactions from {len(results)} documents")

            # Column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 30
            ws.column_dimensions['H'].width = 14  # Quality column

            wb.save(output_path)
            logger.info(f"✅ Batch Rekening Koran Excel export created: {output_path} with {len(results)} entries")
            return True
            
        except Exception as e:
            logger.error(f"❌ Batch Rekening Koran Excel export failed: {e}", exc_info=True)
            return False
    
    def batch_export_to_pdf(self, batch_id: str, results: list, output_path: str) -> bool:
        """Export multiple Rekening Koran entries to formatted PDF (simple format, no table)"""
        try:
            if not HAS_REPORTLAB:
                logger.error("❌ reportlab not available for batch PDF export")
                return False

            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import Paragraph, Spacer, PageBreak
            from reportlab.lib.enums import TA_CENTER
            from datetime import datetime, timezone

            # Create PDF document
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=20*mm,
                leftMargin=20*mm,
                topMargin=15*mm,
                bottomMargin=15*mm
            )

            story = []
            styles = getSampleStyleSheet()

            # Title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=4*mm,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            # Section header style
            section_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Heading2'],
                fontSize=11,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=2*mm,
                spaceBefore=6*mm,
                fontName='Helvetica-Bold'
            )

            # Field style
            field_style = ParagraphStyle(
                'Field',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#333333'),
                spaceAfter=2*mm,
                leading=12
            )

            # Amount style
            amount_style = ParagraphStyle(
                'Amount',
                parent=field_style,
                fontSize=10,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1F4E78')
            )

            # Document separator style (simple - no background)
            doc_separator_style = ParagraphStyle(
                'DocSeparator',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=4*mm,
                spaceBefore=8*mm,
                fontName='Helvetica-Bold',
                alignment=TA_CENTER,
                borderWidth=0,
                borderPadding=0
            )

            # Add batch title
            story.append(Paragraph("REKENING KORAN - MUTASI BANK", title_style))
            story.append(Paragraph(f"BATCH EXPORT - {len(results)} Dokumen",
                                 ParagraphStyle('Subtitle', parent=styles['Normal'],
                                              fontSize=10, alignment=TA_CENTER,
                                              textColor=colors.HexColor('#666666'))))
            story.append(Spacer(1, 10*mm))

            # Process each result
            for idx, result_dict in enumerate(results, start=1):
                extracted_data = result_dict.get('extracted_data', {})

                # Priority: smart_mapped > flat structure > structured_data > extracted_data itself
                if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
                    structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
                    logger.info(f"✅ Using smart_mapped data for Rekening Koran batch PDF item {idx}")
                elif 'transactions' in extracted_data and isinstance(extracted_data.get('transactions'), list):
                    # Handle flat structure from enhanced_bank_processor
                    # Extract fields from nested bank_info and saldo_info to root level
                    structured = extracted_data.copy()
                    bank_info = extracted_data.get('bank_info', {})
                    saldo_info = extracted_data.get('saldo_info', {})

                    # Flatten bank_info fields to root level
                    structured['nama_bank'] = bank_info.get('nama_bank', '') or bank_info.get('bank_name', '')
                    structured['nomor_rekening'] = bank_info.get('nomor_rekening', '') or bank_info.get('account_number', '')
                    structured['nama_pemilik'] = bank_info.get('nama_pemilik', '') or bank_info.get('account_holder', '')
                    structured['periode'] = bank_info.get('periode', '') or bank_info.get('period', '')
                    structured['jenis_rekening'] = bank_info.get('jenis_rekening', '') or bank_info.get('account_type', '')
                    structured['cabang'] = bank_info.get('cabang', '') or bank_info.get('branch', '')
                    structured['alamat'] = bank_info.get('alamat', '') or bank_info.get('address', '')

                    # Flatten saldo_info fields to root level
                    structured['saldo_awal'] = saldo_info.get('saldo_awal', '') or saldo_info.get('opening_balance', '')
                    structured['saldo_akhir'] = saldo_info.get('saldo_akhir', '') or saldo_info.get('closing_balance', '') or saldo_info.get('ending_balance', '')
                    structured['saldo'] = structured['saldo_akhir']  # Alias
                    structured['total_kredit'] = saldo_info.get('total_kredit', '') or saldo_info.get('total_credit', '')
                    structured['total_debet'] = saldo_info.get('total_debet', '') or saldo_info.get('total_debit', '')
                    structured['mata_uang'] = saldo_info.get('mata_uang', '') or saldo_info.get('currency', '')

                    # Use 'transactions' as 'transaksi' for compatibility
                    structured['transaksi'] = extracted_data.get('transactions', [])

                    logger.info(f"✅ Using flat structure (enhanced processor) for Rekening Koran batch PDF item {idx}")
                    logger.info(f"   📝 Flattened fields: nama_bank='{structured.get('nama_bank')}', transactions={len(structured.get('transaksi', []))}")
                elif 'structured_data' in extracted_data:
                    structured = extracted_data['structured_data']
                    logger.info(f"✅ Using structured_data for Rekening Koran batch PDF item {idx}")
                else:
                    structured = extracted_data
                    logger.info(f"⚠️ Using raw extracted_data for Rekening Koran batch PDF item {idx}")

                # Document separator
                story.append(Paragraph(f"REKENING {idx} dari {len(results)}", doc_separator_style))
                story.append(Spacer(1, 3*mm))

                # Compact format for batch
                story.append(Paragraph("INFORMASI REKENING", section_style))
                story.append(Paragraph(f"<b>Nomor Rekening:</b> {structured.get('nomor_rekening', 'N/A')}", field_style))
                story.append(Paragraph(f"<b>Nama Pemilik:</b> {structured.get('nama_pemilik', 'N/A')}", field_style))
                story.append(Paragraph(f"<b>Periode:</b> {structured.get('periode', 'N/A')}", field_style))

                # Ringkasan (compact)
                story.append(Paragraph("RINGKASAN", section_style))
                story.append(Paragraph(f"<b>Saldo Awal:</b> Rp {structured.get('saldo_awal', 'N/A')}", field_style))
                story.append(Paragraph(f"<b>Saldo Akhir:</b> Rp {structured.get('saldo_akhir', structured.get('saldo', 'N/A'))}", amount_style))

                # Transaksi (show limited data for batch)
                # Support both 'transaksi' and 'transactions' field names
                transaksi = structured.get('transaksi', []) or structured.get('transactions', [])
                if transaksi and isinstance(transaksi, list) and len(transaksi) > 0:
                    story.append(Paragraph(f"TRANSAKSI ({len(transaksi)} entries)", section_style))

                    # Show only first 5 transactions per account
                    for tidx, trans in enumerate(transaksi[:5], start=1):
                        story.append(Paragraph(f"<b>{tidx}. {trans.get('tanggal', 'N/A')}</b>", field_style))

                        kredit = trans.get('kredit', trans.get('credit', ''))
                        debet = trans.get('debet', trans.get('debit', ''))

                        if kredit and kredit not in ['', '0', 0, 'N/A']:
                            story.append(Paragraph(f"Masuk: Rp {kredit} | Saldo: Rp {trans.get('saldo', 'N/A')}", field_style))
                        elif debet and debet not in ['', '0', 0, 'N/A']:
                            story.append(Paragraph(f"Keluar: Rp {debet} | Saldo: Rp {trans.get('saldo', 'N/A')}", field_style))
                else:
                    # Single transaction
                    story.append(Paragraph("TRANSAKSI", section_style))
                    kredit = structured.get('kredit', structured.get('credit', ''))
                    debet = structured.get('debet', structured.get('debit', ''))

                    if kredit and kredit not in ['', '0', 0, 'N/A']:
                        story.append(Paragraph(f"<b>Uang Masuk:</b> Rp {kredit}", amount_style))
                    elif debet and debet not in ['', '0', 0, 'N/A']:
                        story.append(Paragraph(f"<b>Uang Keluar:</b> Rp {debet}", amount_style))

                    story.append(Paragraph(f"<b>Saldo:</b> Rp {structured.get('saldo', 'N/A')}", field_style))

                # Add page break between accounts (except last one)
                if idx < len(results):
                    story.append(PageBreak())

            # Footer on last page
            story.append(Spacer(1, 10*mm))
            footer_text = f"<i>Batch export {len(results)} rekening koran dibuat oleh Doc-Scan AI pada {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M:%S UTC')}</i>"
            story.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Normal'],
                                                              fontSize=8, alignment=TA_CENTER,
                                                              textColor=colors.HexColor('#999999'))))

            # Build PDF
            doc.build(story)
            logger.info(f"✅ Batch Rekening Koran PDF export created: {output_path} with {len(results)} entries")
            return True

        except Exception as e:
            logger.error(f"❌ Batch Rekening Koran PDF export failed: {e}", exc_info=True)
            return False
