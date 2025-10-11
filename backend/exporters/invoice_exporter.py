"""
Invoice Exporter
Handles Excel and PDF export specifically for Standard Invoice documents
"""

from typing import Dict, Any
from datetime import datetime
import logging

from .base_exporter import BaseExporter

logger = logging.getLogger(__name__)

# Check for required libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("⚠️ openpyxl not installed - Excel export disabled")

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    logger.warning("⚠️ reportlab not installed - PDF export disabled")


class InvoiceExporter(BaseExporter):
    """Exporter for Standard Invoice documents"""

    def __init__(self):
        super().__init__("invoice")
        # Columns specific to Invoice
        self.columns = [
            "PO", "Tgl PO", "Tgl Invoice", "Keterangan", "Nilai", "Tanggal"
        ]

    def _convert_smart_mapped_to_structured(self, smart_mapped: dict) -> dict:
        """Convert Smart Mapper output to structured format for Invoice"""
        if not smart_mapped or not isinstance(smart_mapped, dict):
            return {}

        structured = {}

        # Extract main sections from smart_mapped
        purchase_order = smart_mapped.get('purchase_order', {}) or {}
        invoice_info = smart_mapped.get('invoice_info', {}) or {}
        customer = smart_mapped.get('customer', {}) or {}
        vendor = smart_mapped.get('vendor', {}) or {}
        summary = smart_mapped.get('summary', {}) or {}
        payment = smart_mapped.get('payment', {}) or {}
        additional = smart_mapped.get('additional', {}) or {}

        # PO Information
        structured['po_number'] = purchase_order.get('po_number') or purchase_order.get('po') or 'N/A'
        structured['po'] = structured['po_number']  # Alias
        structured['tgl_po'] = purchase_order.get('tgl_po') or purchase_order.get('tanggal_po') or 'N/A'
        structured['tanggal_po'] = structured['tgl_po']  # Alias

        # Invoice Information
        structured['invoice_number'] = invoice_info.get('invoice_number') or invoice_info.get('nomor') or 'N/A'
        structured['invoice_type'] = invoice_info.get('invoice_type') or invoice_info.get('type') or ''
        structured['tgl_invoice'] = invoice_info.get('tgl_invoice') or invoice_info.get('tanggal_invoice') or invoice_info.get('tanggal') or 'N/A'
        structured['tanggal_invoice'] = structured['tgl_invoice']  # Alias
        structured['tanggal'] = invoice_info.get('tanggal') or structured['tgl_invoice']
        structured['due_date'] = invoice_info.get('due_date') or invoice_info.get('tanggal_jatuh_tempo') or ''
        structured['currency'] = invoice_info.get('currency') or invoice_info.get('mata_uang') or 'IDR'
        structured['payment_terms'] = invoice_info.get('payment_terms') or invoice_info.get('syarat_pembayaran') or ''
        structured['reference'] = invoice_info.get('reference') or invoice_info.get('referensi') or ''

        # Customer Information
        structured['customer_nama'] = customer.get('nama') or customer.get('name') or 'N/A'
        structured['customer_alamat'] = customer.get('alamat') or customer.get('address') or 'N/A'
        structured['customer_kota'] = customer.get('kota') or customer.get('city') or ''
        structured['customer_negara'] = customer.get('negara') or customer.get('country') or ''
        structured['customer_telepon'] = customer.get('telepon') or customer.get('phone') or ''
        structured['customer_email'] = customer.get('email') or ''
        structured['customer_npwp'] = customer.get('npwp') or customer.get('tax_id') or 'N/A'

        # Vendor Information
        structured['vendor_nama'] = vendor.get('nama') or vendor.get('name') or 'N/A'
        structured['vendor_alamat'] = vendor.get('alamat') or vendor.get('address') or 'N/A'
        structured['vendor_kota'] = vendor.get('kota') or vendor.get('city') or ''
        structured['vendor_negara'] = vendor.get('negara') or vendor.get('country') or ''
        structured['vendor_telepon'] = vendor.get('telepon') or vendor.get('phone') or ''
        structured['vendor_email'] = vendor.get('email') or ''
        structured['vendor_npwp'] = vendor.get('npwp') or vendor.get('tax_id') or 'N/A'

        # Extract items array (CRITICAL for multi-item invoices)
        items = smart_mapped.get('items', []) or []
        if items and isinstance(items, list):
            structured['items'] = []
            for item in items:
                if isinstance(item, dict):
                    item_dict = {
                        'no': item.get('no') or '',
                        'kode': item.get('kode') or item.get('sku') or '',
                        'deskripsi': item.get('deskripsi') or item.get('description') or item.get('nama') or 'N/A',
                        'keterangan': item.get('keterangan') or item.get('notes') or '',
                        'qty': item.get('qty') or item.get('quantity') or item.get('jumlah') or '',
                        'satuan': item.get('satuan') or item.get('unit') or '',
                        'harga_satuan': item.get('harga_satuan') or item.get('unit_price') or item.get('price') or '',
                        'diskon': item.get('diskon') or item.get('discount') or '',
                        'subtotal': item.get('subtotal') or item.get('amount') or item.get('total') or '',
                        'pajak': item.get('pajak') or item.get('tax') or ''
                    }
                    structured['items'].append(item_dict)
            logger.info(f"✅ Extracted {len(structured['items'])} items from Smart Mapper")

        # Summary/Total Information
        structured['subtotal'] = summary.get('subtotal') or summary.get('sub_total') or ''
        structured['diskon_total'] = summary.get('diskon_total') or summary.get('total_discount') or ''
        structured['pajak_persen'] = summary.get('pajak_persen') or summary.get('tax_percent') or ''
        structured['pajak_nilai'] = summary.get('pajak_nilai') or summary.get('tax_amount') or ''
        structured['biaya_lain'] = summary.get('biaya_lain') or summary.get('other_charges') or ''
        structured['total'] = summary.get('total') or summary.get('nilai') or summary.get('grand_total') or 'N/A'
        structured['nilai'] = structured['total']  # Alias
        structured['terbilang'] = summary.get('terbilang') or summary.get('amount_in_words') or ''

        # Payment Information
        structured['payment_metode'] = payment.get('metode') or payment.get('method') or ''
        structured['payment_bank'] = payment.get('bank_name') or ''
        structured['payment_account'] = payment.get('account_number') or ''
        structured['payment_account_name'] = payment.get('account_name') or ''
        structured['payment_swift'] = payment.get('swift_code') or ''

        # Additional / Notes
        structured['notes'] = additional.get('notes') or additional.get('catatan') or ''
        structured['keterangan'] = additional.get('keterangan') or additional.get('remarks') or ''

        # Backward compatibility: If no items array but has old 'detail' section, use it
        if not structured.get('items'):
            detail = smart_mapped.get('detail', {}) or {}
            if detail:
                structured['keterangan'] = detail.get('keterangan') or detail.get('deskripsi') or detail.get('description') or structured.get('keterangan', 'N/A')
                structured['deskripsi'] = structured['keterangan']
                # If detail has nilai, it overrides summary total (for backward compatibility)
                if detail.get('nilai') or detail.get('total'):
                    structured['nilai'] = detail.get('nilai') or detail.get('total') or structured.get('nilai', 'N/A')
                    structured['total'] = structured['nilai']

        logger.info(f"✅ Invoice Smart Mapper data converted to structured format with {len(structured.get('items', []))} items")

        return structured
    
    def export_to_excel(self, result: Dict[str, Any], output_path: str) -> bool:
        """Export single Invoice to Excel with structured table format"""
        try:
            if not HAS_OPENPYXL:
                logger.error("❌ openpyxl not available for Excel export")
                return False
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoice"
            
            self._populate_excel_sheet(ws, result)
            
            wb.save(output_path)
            logger.info(f"✅ Invoice Excel export created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Invoice Excel export failed: {e}", exc_info=True)
            return False
    
    def _populate_excel_sheet(self, ws, result: Dict[str, Any]):
        """Helper to populate worksheet with structured Invoice data"""
        # Define styles
        header_fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")  # Red
        header_font = Font(bold=True, size=11, color="FFFFFF")
        data_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")  # Light red
        border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=False)
        
        # Get structured data
        extracted_data = result.get('extracted_data', {})

        # DEBUG: Log what keys are available
        logger.info(f"🔍 Invoice Excel Export - Available keys in extracted_data: {list(extracted_data.keys())}")
        logger.info(f"🔍 Invoice Excel Export - Has smart_mapped: {'smart_mapped' in extracted_data}")
        logger.info(f"🔍 Invoice Excel Export - Has structured_data: {'structured_data' in extracted_data}")

        if 'smart_mapped' in extracted_data:
            smart_mapped_data = extracted_data['smart_mapped']
            logger.info(f"🔍 Invoice Excel Export - smart_mapped type: {type(smart_mapped_data)}")
            if isinstance(smart_mapped_data, dict):
                logger.info(f"🔍 Invoice Excel Export - smart_mapped keys: {list(smart_mapped_data.keys())}")
                if 'items' in smart_mapped_data:
                    logger.info(f"🔍 Invoice Excel Export - smart_mapped items count: {len(smart_mapped_data.get('items', []))}")

        # Priority: smart_mapped > structured_data > extracted_data itself
        if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
            structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
            logger.info("✅ Using smart_mapped data for Invoice Excel export")
        elif 'structured_data' in extracted_data:
            structured = extracted_data['structured_data']
            logger.info("✅ Using structured_data for Invoice Excel export")
        else:
            structured = extracted_data
            logger.info("⚠️ Using raw extracted_data for Invoice Excel export")

        # DEBUG: Log structured data after conversion
        logger.info(f"🔍 Invoice Excel Export - After conversion, items count: {len(structured.get('items', []))}")
        
        row = 1
        
        # ===== TITLE =====
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "📄 INVOICE - TAGIHAN"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_align
        ws[f'A{row}'].border = border_thin
        row += 1
        
        # ===== TABLE HEADERS =====
        for col_idx, header in enumerate(self.columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_thin
        row += 1
        
        # ===== DATA ROW =====
        data_row = [
            structured.get('po_number', structured.get('po', 'N/A')),
            structured.get('tgl_po', structured.get('tanggal_po', 'N/A')),
            structured.get('tgl_invoice', structured.get('tanggal_invoice', structured.get('tanggal', 'N/A'))),
            structured.get('keterangan', structured.get('deskripsi', 'N/A')),
            structured.get('nilai', structured.get('total', 'N/A')),
            structured.get('tanggal', 'N/A')
        ]

        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill = data_fill
            # Right align for numeric columns (Nilai)
            if col_idx == 5:
                cell.alignment = right_align
            else:
                cell.alignment = left_align
            cell.border = border_thin
            cell.font = Font(size=10)
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 18  # PO
        ws.column_dimensions['B'].width = 14  # Tgl PO
        ws.column_dimensions['C'].width = 14  # Tgl Invoice
        ws.column_dimensions['D'].width = 40  # Keterangan
        ws.column_dimensions['E'].width = 18  # Nilai
        ws.column_dimensions['F'].width = 14  # Tanggal

        ws.row_dimensions[3].height = 40
    
    def export_to_pdf(self, result: Dict[str, Any], output_path: str) -> bool:
        """Export single Invoice to formatted PDF (formal, without table)"""
        try:
            if not HAS_REPORTLAB:
                logger.error("❌ reportlab not available for PDF export")
                return False

            from reportlab.lib.units import mm
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            from datetime import datetime, timezone

            # Get structured data
            extracted_data = result.get('extracted_data', {})

            # Priority: smart_mapped > structured_data > extracted_data itself
            if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
                structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
                logger.info("✅ Using smart_mapped data for Invoice PDF export")
            elif 'structured_data' in extracted_data:
                structured = extracted_data['structured_data']
                logger.info("✅ Using structured_data for Invoice PDF export")
            else:
                structured = extracted_data
                logger.info("⚠️ Using raw extracted_data for Invoice PDF export")

            # Create PDF document
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=20*mm,
                leftMargin=20*mm,
                topMargin=15*mm,
                bottomMargin=15*mm
            )

            # Container for PDF elements
            story = []

            # Setup styles
            styles = getSampleStyleSheet()

            # Title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#dc2626'),
                spaceAfter=4*mm,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            # Section header style
            section_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Heading2'],
                fontSize=11,
                textColor=colors.HexColor('#dc2626'),
                spaceAfter=2*mm,
                spaceBefore=6*mm,
                fontName='Helvetica-Bold'
            )

            # Field style
            field_style = ParagraphStyle(
                'Field',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#333333'),
                spaceAfter=2*mm,
                leading=14
            )

            # Add title
            story.append(Paragraph("INVOICE", title_style))
            story.append(Paragraph("TAGIHAN",
                                 ParagraphStyle('Subtitle', parent=styles['Normal'],
                                              fontSize=10, alignment=TA_CENTER,
                                              textColor=colors.HexColor('#666666'))))
            story.append(Spacer(1, 10*mm))

            # Section 1: Informasi Purchase Order
            story.append(Paragraph("INFORMASI PURCHASE ORDER", section_style))
            story.append(Paragraph(f"<b>Nomor PO:</b> {structured.get('po_number', structured.get('po', 'N/A'))}", field_style))
            story.append(Paragraph(f"<b>Tanggal PO:</b> {structured.get('tgl_po', structured.get('tanggal_po', 'N/A'))}", field_style))

            # Section 2: Informasi Invoice
            story.append(Paragraph("INFORMASI INVOICE", section_style))
            story.append(Paragraph(f"<b>Tanggal Invoice:</b> {structured.get('tgl_invoice', structured.get('tanggal_invoice', structured.get('tanggal', 'N/A')))}", field_style))
            story.append(Paragraph(f"<b>Tanggal:</b> {structured.get('tanggal', 'N/A')}", field_style))

            # Section 3: Detail Tagihan
            story.append(Paragraph("DETAIL TAGIHAN", section_style))
            story.append(Paragraph(f"<b>Keterangan:</b> {structured.get('keterangan', structured.get('deskripsi', 'N/A'))}", field_style))

            # Nilai - highlighted
            nilai_style = ParagraphStyle(
                'Nilai',
                parent=field_style,
                fontSize=12,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#dc2626'),
                spaceAfter=4*mm
            )
            story.append(Paragraph(f"<b>Nilai:</b> Rp {structured.get('nilai', structured.get('total', 'N/A'))}", nilai_style))

            # Footer
            story.append(Spacer(1, 10*mm))
            footer_text = f"<i>Dokumen ini dibuat secara otomatis oleh Doc-Scan AI System pada {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M:%S UTC')}</i>"
            story.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Normal'],
                                                              fontSize=8, alignment=TA_CENTER,
                                                              textColor=colors.HexColor('#999999'))))

            # Build PDF
            doc.build(story)
            logger.info(f"✅ Invoice PDF export created: {output_path}")
            return True

        except Exception as e:
            logger.error(f"❌ Invoice PDF export failed: {e}", exc_info=True)
            return False
    
    def batch_export_to_excel(self, batch_id: str, results: list, output_path: str) -> bool:
        """Export multiple Invoices to single Excel file"""
        try:
            if not HAS_OPENPYXL:
                logger.error("❌ openpyxl not available for batch Excel export")
                return False
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Batch Invoices"
            
            # Styles
            header_fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")
            header_font = Font(bold=True, size=11, color="FFFFFF")
            data_fill_1 = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
            data_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            border_thin = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            right_align = Alignment(horizontal='right', vertical='center', wrap_text=False)
            
            row = 1
            
            # Batch info
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = f"📄 BATCH INVOICES: {batch_id} | Total: {len(results)} | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = center_align
            ws[f'A{row}'].border = border_thin
            row += 1
            
            # Headers
            for col_idx, header in enumerate(self.columns, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border_thin
            row += 1
            
            # Data rows
            for idx, result_dict in enumerate(results):
                extracted_data = result_dict.get('extracted_data', {})

                # Priority: smart_mapped > structured_data
                if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
                    structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
                elif 'structured_data' in extracted_data:
                    structured = extracted_data['structured_data']
                else:
                    structured = extracted_data
                
                data_row = [
                    structured.get('po_number', structured.get('po', 'N/A')),
                    structured.get('tgl_po', structured.get('tanggal_po', 'N/A')),
                    structured.get('tgl_invoice', structured.get('tanggal_invoice', structured.get('tanggal', 'N/A'))),
                    structured.get('keterangan', structured.get('deskripsi', 'N/A')),
                    structured.get('nilai', structured.get('total', 'N/A')),
                    structured.get('tanggal', 'N/A')
                ]

                fill = data_fill_1 if idx % 2 == 0 else data_fill_2

                for col_idx, value in enumerate(data_row, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.fill = fill
                    # Right align for numeric columns (Nilai)
                    if col_idx == 5:
                        cell.alignment = right_align
                    else:
                        cell.alignment = left_align
                    cell.border = border_thin
                    cell.font = Font(size=10)
                
                row += 1
            
            # Column widths
            ws.column_dimensions['A'].width = 18  # PO
            ws.column_dimensions['B'].width = 14  # Tgl PO
            ws.column_dimensions['C'].width = 14  # Tgl Invoice
            ws.column_dimensions['D'].width = 40  # Keterangan
            ws.column_dimensions['E'].width = 18  # Nilai
            ws.column_dimensions['F'].width = 14  # Tanggal

            wb.save(output_path)
            logger.info(f"✅ Batch Invoice Excel export created: {output_path} with {len(results)} invoices")
            return True
            
        except Exception as e:
            logger.error(f"❌ Batch Invoice Excel export failed: {e}", exc_info=True)
            return False
    
    def batch_export_to_pdf(self, batch_id: str, results: list, output_path: str) -> bool:
        """Export multiple Invoices to formatted PDF (formal, without table)"""
        try:
            if not HAS_REPORTLAB:
                logger.error("❌ reportlab not available for batch PDF export")
                return False

            from reportlab.lib.units import mm
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            from reportlab.platypus import PageBreak
            from datetime import datetime, timezone

            # Create PDF document
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=20*mm,
                leftMargin=20*mm,
                topMargin=15*mm,
                bottomMargin=15*mm
            )

            # Container for PDF elements
            story = []

            # Setup styles
            styles = getSampleStyleSheet()

            # Title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#dc2626'),
                spaceAfter=4*mm,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            # Section header style
            section_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Heading2'],
                fontSize=11,
                textColor=colors.HexColor('#dc2626'),
                spaceAfter=2*mm,
                spaceBefore=6*mm,
                fontName='Helvetica-Bold'
            )

            # Field style
            field_style = ParagraphStyle(
                'Field',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#333333'),
                spaceAfter=2*mm,
                leading=14
            )

            # Document number style
            doc_number_style = ParagraphStyle(
                'DocNumber',
                parent=styles['Heading2'],
                fontSize=13,
                textColor=colors.HexColor('#dc2626'),
                spaceAfter=4*mm,
                spaceBefore=8*mm,
                fontName='Helvetica-Bold',
                alignment=TA_CENTER
            )

            # Add batch title
            story.append(Paragraph("BATCH INVOICES", title_style))
            story.append(Paragraph(f"BATCH EXPORT - {len(results)} Dokumen",
                                 ParagraphStyle('Subtitle', parent=styles['Normal'],
                                              fontSize=10, alignment=TA_CENTER,
                                              textColor=colors.HexColor('#666666'))))
            story.append(Spacer(1, 8*mm))

            # Process each result
            for idx, result_dict in enumerate(results, start=1):
                # Get structured data
                extracted_data = result_dict.get('extracted_data', {})

                # Priority: smart_mapped > structured_data
                if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
                    structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
                    logger.info(f"✅ Using smart_mapped data for Invoice batch item {idx}")
                elif 'structured_data' in extracted_data:
                    structured = extracted_data['structured_data']
                    logger.info(f"✅ Using structured_data for Invoice batch item {idx}")
                else:
                    structured = extracted_data
                    logger.info(f"⚠️ Using raw extracted_data for Invoice batch item {idx}")

                # Document separator
                story.append(Paragraph(f"INVOICE {idx} dari {len(results)}", doc_number_style))
                story.append(Spacer(1, 3*mm))

                # Section 1: Informasi Purchase Order
                story.append(Paragraph("INFORMASI PURCHASE ORDER", section_style))
                story.append(Paragraph(f"<b>Nomor PO:</b> {structured.get('po_number', structured.get('po', 'N/A'))}", field_style))
                story.append(Paragraph(f"<b>Tanggal PO:</b> {structured.get('tgl_po', structured.get('tanggal_po', 'N/A'))}", field_style))

                # Section 2: Informasi Invoice
                story.append(Paragraph("INFORMASI INVOICE", section_style))
                story.append(Paragraph(f"<b>Tanggal Invoice:</b> {structured.get('tgl_invoice', structured.get('tanggal_invoice', structured.get('tanggal', 'N/A')))}", field_style))
                story.append(Paragraph(f"<b>Tanggal:</b> {structured.get('tanggal', 'N/A')}", field_style))

                # Section 3: Detail Tagihan (Compact)
                story.append(Paragraph("DETAIL TAGIHAN", section_style))
                story.append(Paragraph(f"<b>Keterangan:</b> {structured.get('keterangan', structured.get('deskripsi', 'N/A'))}", field_style))

                # Nilai - highlighted
                nilai_highlight = ParagraphStyle(
                    'Nilai',
                    parent=field_style,
                    fontSize=11,
                    fontName='Helvetica-Bold',
                    textColor=colors.HexColor('#dc2626'),
                    spaceAfter=4*mm
                )
                story.append(Paragraph(f"<b>Nilai:</b> Rp {structured.get('nilai', structured.get('total', 'N/A'))}", nilai_highlight))

                # Add page break between documents (except last one)
                if idx < len(results):
                    story.append(PageBreak())

            # Footer on last page
            story.append(Spacer(1, 10*mm))
            footer_text = f"<i>Batch export {len(results)} invoice dibuat oleh Doc-Scan AI pada {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M:%S UTC')}</i>"
            story.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Normal'],
                                                              fontSize=8, alignment=TA_CENTER,
                                                              textColor=colors.HexColor('#999999'))))

            # Build PDF
            doc.build(story)
            logger.info(f"✅ Batch Invoice PDF export created: {output_path} ({len(results)} documents)")
            return True

        except Exception as e:
            logger.error(f"❌ Batch Invoice PDF export failed: {e}", exc_info=True)
            return False
