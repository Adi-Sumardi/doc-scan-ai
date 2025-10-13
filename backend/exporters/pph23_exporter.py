"""
PPh 23 Exporter
Handles Excel export specifically for Indonesian Tax Withholding Certificate (Bukti Potong PPh 23)
"""

from typing import Dict, Any
import logging

from .base_exporter import BaseExporter

logger = logging.getLogger(__name__)

# Check for required libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("⚠️ openpyxl not installed - Excel export disabled")


class PPh23Exporter(BaseExporter):
    """Exporter for PPh 23 (Indonesian Tax Withholding Certificate) documents"""

    def __init__(self):
        super().__init__("pph23")
        self.columns = [
            "Nomor",
            "Masa Pajak",
            "Sifat Pemotongan",
            "Status Bukti Pemotongan",
            "NPWP/NIK Penerima",
            "Nama Penerima",
            "NITKU Penerima",
            "Jenis PPh",
            "Kode Objek Pajak",
            "Objek Pajak",
            "DPP",
            "Tarif (%)",
            "PPh",
            "Jenis Dokumen Dasar",
            "Tanggal Dokumen Dasar",
            "NPWP/NIK Pemotong",
            "NITKU/Subunit Pemotong",
            "Nama Pemotong",
            "Tanggal Potong",
            "Nama Penandatangan"
        ]

    def _format_rupiah(self, value) -> str:
        """Format number to Rupiah without decimals"""
        if not value or value in ['-', 'N/A', '', '0', 0, 'None', 'null']:
            return '-'
        try:
            # If already a number (int or float), convert to int first
            if isinstance(value, (int, float)):
                value_int = int(value)
            else:
                # If string, clean and parse
                value_str = str(value).replace('Rp', '').replace('IDR', '').replace('.', '').replace(',', '').strip()
                if not value_str or value_str == '-':
                    return '-'
                value_int = int(float(value_str))

            if value_int == 0:
                return '-'
            formatted = f"Rp {value_int:,}"
            formatted = formatted.replace(',', '.')
            return formatted
        except (ValueError, AttributeError):
            logger.warning(f"⚠️ Failed to format rupiah value: {value}")
            return str(value) if value else '-'

    def _format_date(self, date_str) -> str:
        """Format date to DD/MM/YYYY"""
        if not date_str or date_str in ['N/A', '', 'None', 'null']:
            return 'N/A'

        import re
        from datetime import datetime

        date_str = str(date_str).strip()

        # Try various date formats
        formats = [
            '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d %B %Y', '%d %b %Y', '%d-%m-%y', '%d/%m/%y'
        ]

        for fmt in formats:
            try:
                date_obj = datetime.strptime(date_str, fmt)
                return date_obj.strftime('%d/%m/%Y')
            except:
                continue

        # If no format worked, try to extract DD, MM, YYYY
        match = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', date_str)
        if match:
            return f"{match.group(1).zfill(2)}/{match.group(2).zfill(2)}/{match.group(3)}"

        return date_str  # Return as-is if can't format

    def _convert_smart_mapped_to_structured(self, smart_mapped: dict) -> dict:
        """Convert Smart Mapper output to structured format for PPh 23"""
        if not smart_mapped or not isinstance(smart_mapped, dict):
            return {}
        
        structured = {}
        
        # Extract main sections from smart_mapped
        penerima = smart_mapped.get('penerima', {}) or {}
        pemotong = smart_mapped.get('pemotong', {}) or {}
        dokumen = smart_mapped.get('dokumen', {}) or {}
        objek_pajak = smart_mapped.get('objek_pajak', {}) or {}
        dokumen_dasar = smart_mapped.get('dokumen_dasar', {}) or {}
        
        # Nomor
        structured['nomor'] = dokumen.get('nomor') or dokumen.get('number') or 'N/A'
        
        # Masa Pajak
        structured['masa_pajak'] = dokumen.get('masa_pajak') or dokumen.get('tax_period') or 'N/A'
        
        # Sifat Pemotongan
        structured['sifat_pemotongan'] = dokumen.get('sifat_pemotongan') or dokumen.get('nature') or 'N/A'
        
        # Status Bukti Pemotongan
        structured['status'] = dokumen.get('status') or 'N/A'
        
        # Identitas Penerima Penghasilan
        structured['penerima_npwp'] = penerima.get('npwp') or penerima.get('nik') or 'N/A'
        structured['penerima_nama'] = penerima.get('nama') or penerima.get('name') or 'N/A'
        # Try multiple keys for NITKU
        structured['penerima_nitku'] = (
            penerima.get('nitku') or 
            penerima.get('NITKU') or 
            penerima.get('nomor_identitas') or 
            penerima.get('nomor_identitas_tempat_kegiatan_usaha') or 
            'N/A'
        )
        
        # Jenis PPh
        structured['jenis_pph'] = objek_pajak.get('jenis_pph') or 'PPh 23'
        
        # Kode Objek Pajak
        structured['kode_objek'] = objek_pajak.get('kode') or objek_pajak.get('code') or 'N/A'
        
        # Objek Pajak
        structured['objek_pajak'] = objek_pajak.get('objek') or objek_pajak.get('description') or 'N/A'
        
        # Financials with case-insensitive lookup
        financials = smart_mapped.get('financials', {}) or {}
        
        def get_financial_value(key_variants):
            """Try multiple key variants case-insensitively"""
            if not isinstance(financials, dict):
                return 'N/A'
            for key in key_variants:
                # Try exact match first
                if key in financials:
                    val = financials[key]
                    return val if val not in (None, '', 'N/A') else 'N/A'
                # Try case-insensitive
                for k, v in financials.items():
                    if k.lower() == key.lower():
                        return v if v not in (None, '', 'N/A') else 'N/A'
            return 'N/A'
        
        # Remove 'Rp' and non-digit characters from DPP
        raw_dpp = get_financial_value(['dpp', 'DPP', 'bruto', 'gross_amount'])
        if isinstance(raw_dpp, str):
            cleaned_dpp = ''.join(c for c in raw_dpp if c.isdigit() or c == '.')
            structured['dpp'] = cleaned_dpp if cleaned_dpp else raw_dpp
        else:
            structured['dpp'] = raw_dpp
        structured['tarif'] = get_financial_value(['tarif', 'rate', 'tax_rate'])
        structured['pph'] = get_financial_value(['pph', 'PPh', 'tax_amount', 'withheld'])
        
        # Dasar Dokumen
        structured['dokumen_dasar_jenis'] = dokumen_dasar.get('jenis') or dokumen_dasar.get('type') or 'N/A'
        structured['dokumen_dasar_tanggal'] = dokumen_dasar.get('tanggal') or dokumen_dasar.get('date') or 'N/A'
        
        # Identitas Pemotong
        structured['pemotong_npwp'] = pemotong.get('npwp') or pemotong.get('nik') or 'N/A'
        # Try multiple keys for NITKU/Subunit
        structured['pemotong_nitku'] = (
            pemotong.get('nitku') or 
            pemotong.get('NITKU') or 
            pemotong.get('subunit') or 
            pemotong.get('nomor_identitas') or
            pemotong.get('nomor_identitas_tempat_kegiatan_usaha') or 
            'N/A'
        )
        structured['pemotong_nama'] = pemotong.get('nama') or pemotong.get('name') or 'N/A'
        structured['tanggal_potong'] = dokumen.get('tanggal') or dokumen.get('date') or 'N/A'
        structured['penandatangan'] = pemotong.get('penandatangan') or pemotong.get('signatory') or 'N/A'
        
        logger.info(f"✅ PPh 23 Smart Mapper data converted to structured format")
        
        return structured
    
    def export_to_excel(self, result: dict, output_path: str) -> bool:
        """Export single PPh 23 document to Excel"""
        if not HAS_OPENPYXL:
            logger.error("❌ openpyxl not installed - cannot create Excel")
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "PPh 23"
            
            # Get structured data from result
            extracted_data = result.get('extracted_data', {})

            # Priority: smart_mapped > structured_data > extracted_data itself
            # Note: For PPh23, smart_mapped is preferred as it has better AI mapping
            if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
                structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
                logger.info("✅ Using smart_mapped data for PPh 23 Excel export")
            elif 'structured_data' in extracted_data:
                structured = extracted_data['structured_data']
                logger.info("✅ Using structured_data for PPh 23 Excel export")
            else:
                structured = extracted_data
                logger.info("⚠️ Using raw extracted_data for PPh 23 Excel export")
            
            # Styles
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            for col_idx, column in enumerate(self.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Write data with formatting
            row_data = [
                structured.get('nomor', 'N/A'),
                structured.get('masa_pajak', 'N/A'),
                structured.get('sifat_pemotongan', 'N/A'),
                structured.get('status', 'N/A'),
                structured.get('penerima_npwp', 'N/A'),
                structured.get('penerima_nama', 'N/A'),
                structured.get('penerima_nitku', 'N/A'),
                structured.get('jenis_pph', 'PPh 23'),
                structured.get('kode_objek', 'N/A'),
                structured.get('objek_pajak', 'N/A'),
                self._format_rupiah(structured.get('dpp', 'N/A')),
                structured.get('tarif', 'N/A'),
                self._format_rupiah(structured.get('pph', 'N/A')),
                structured.get('dokumen_dasar_jenis', 'N/A'),
                self._format_date(structured.get('dokumen_dasar_tanggal', 'N/A')),
                structured.get('pemotong_npwp', 'N/A'),
                structured.get('pemotong_nitku', 'N/A'),
                structured.get('pemotong_nama', 'N/A'),
                self._format_date(structured.get('tanggal_potong', 'N/A')),
                structured.get('penandatangan', 'N/A'),
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=2, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Adjust column widths
            column_widths = {
                'A': 15, 'B': 15, 'C': 20, 'D': 20, 'E': 20,
                'F': 25, 'G': 20, 'H': 12, 'I': 15, 'J': 30,
                'K': 15, 'L': 10, 'M': 15, 'N': 20, 'O': 15,
                'P': 20, 'Q': 20, 'R': 25, 'S': 15, 'T': 25,
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"✅ PPh 23 Excel export created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ PPh 23 Excel export failed: {e}", exc_info=True)
            return False
    
    def export_to_pdf(self, result: dict, output_path: str) -> bool:
        """Export single PPh 23 document to formatted PDF (not table)"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, mm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            from datetime import datetime, timezone

            # Get structured data from result
            extracted_data = result.get('extracted_data', {})

            # Priority: smart_mapped > structured_data > extracted_data itself
            # Note: For PPh23, smart_mapped is preferred as it has better AI mapping
            if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
                structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
                logger.info("✅ Using smart_mapped data for PPh 23 PDF export")
            elif 'structured_data' in extracted_data:
                structured = extracted_data['structured_data']
                logger.info("✅ Using structured_data for PPh 23 PDF export")
            else:
                structured = extracted_data
                logger.info("⚠️ Using raw extracted_data for PPh 23 PDF export")

            # Create PDF document
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=20*mm,
                leftMargin=20*mm,
                topMargin=15*mm,
                bottomMargin=15*mm
            )

            # Container for PDF elements
            story = []

            # Setup styles
            styles = getSampleStyleSheet()

            # Title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=4*mm,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            # Section header style
            section_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Heading2'],
                fontSize=11,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=2*mm,
                spaceBefore=6*mm,
                fontName='Helvetica-Bold',
                borderWidth=0,
                borderPadding=0
            )

            # Field style (Label: Value on same line)
            field_style = ParagraphStyle(
                'Field',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#333333'),
                spaceAfter=2*mm,
                leading=12
            )

            # Add title
            story.append(Paragraph("BUKTI PEMOTONGAN PPh PASAL 23", title_style))
            story.append(Paragraph("INCOME TAX ARTICLE 23 WITHHOLDING CERTIFICATE",
                                 ParagraphStyle('Subtitle', parent=styles['Normal'],
                                              fontSize=9, alignment=TA_CENTER,
                                              textColor=colors.HexColor('#666666'))))
            story.append(Spacer(1, 10*mm))

            # Section 1: Informasi Dokumen
            story.append(Paragraph("INFORMASI DOKUMEN", section_style))
            story.append(Paragraph(f"<b>Nomor Bukti Potong:</b> {structured.get('nomor', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>Masa Pajak:</b> {structured.get('masa_pajak', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>Tanggal Pemotongan:</b> {structured.get('tanggal_potong', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>Sifat Pemotongan:</b> {structured.get('sifat_pemotongan', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>Status:</b> {structured.get('status', 'N/A')}", field_style))

            # Section 2: Identitas Penerima Penghasilan
            story.append(Paragraph("IDENTITAS PENERIMA PENGHASILAN", section_style))
            story.append(Paragraph(f"<b>Nama:</b> {structured.get('penerima_nama', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>NPWP/NIK:</b> {structured.get('penerima_npwp', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>NITKU:</b> {structured.get('penerima_nitku', 'N/A')}", field_style))

            # Section 3: Objek Pajak
            story.append(Paragraph("OBJEK PAJAK", section_style))
            story.append(Paragraph(f"<b>Jenis PPh:</b> {structured.get('jenis_pph', 'PPh 23')}", field_style))
            story.append(Paragraph(f"<b>Kode Objek Pajak:</b> {structured.get('kode_objek', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>Uraian Objek Pajak:</b> {structured.get('objek_pajak', 'N/A')}", field_style))

            # Section 4: Perhitungan Pajak
            story.append(Paragraph("PERHITUNGAN PAJAK", section_style))
            story.append(Paragraph(f"<b>Dasar Pengenaan Pajak (DPP):</b> Rp {structured.get('dpp', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>Tarif:</b> {structured.get('tarif', 'N/A')}", field_style))

            # PPh Dipotong - highlighted
            pph_style = ParagraphStyle(
                'PPh',
                parent=field_style,
                fontSize=10,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=4*mm
            )
            story.append(Paragraph(f"<b>PPh Dipotong:</b> Rp {structured.get('pph', 'N/A')}", pph_style))

            # Section 5: Dokumen Dasar
            story.append(Paragraph("DOKUMEN DASAR PEMOTONGAN", section_style))
            story.append(Paragraph(f"<b>Jenis Dokumen:</b> {structured.get('dokumen_dasar_jenis', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>Tanggal Dokumen:</b> {structured.get('dokumen_dasar_tanggal', 'N/A')}", field_style))

            # Section 6: Identitas Pemotong
            story.append(Paragraph("IDENTITAS PEMOTONG PAJAK", section_style))
            story.append(Paragraph(f"<b>Nama:</b> {structured.get('pemotong_nama', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>NPWP/NIK:</b> {structured.get('pemotong_npwp', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>NITKU/Subunit:</b> {structured.get('pemotong_nitku', 'N/A')}", field_style))
            story.append(Paragraph(f"<b>Nama Penandatangan:</b> {structured.get('penandatangan', 'N/A')}", field_style))

            # Footer
            story.append(Spacer(1, 10*mm))
            footer_text = f"<i>Dokumen ini dibuat secara otomatis oleh Doc-Scan AI System pada {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M:%S UTC')}</i>"
            story.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Normal'],
                                                              fontSize=8, alignment=TA_CENTER,
                                                              textColor=colors.HexColor('#999999'))))

            # Build PDF
            doc.build(story)
            logger.info(f"✅ PPh 23 PDF export created: {output_path}")
            return True

        except Exception as e:
            logger.error(f"❌ PPh 23 PDF export failed: {e}", exc_info=True)
            return False
    
    def batch_export_to_excel(self, batch_id: str, batch_results: list, output_path: str) -> bool:
        """Export multiple PPh 23 documents to a single Excel file"""
        if not HAS_OPENPYXL:
            logger.error("❌ openpyxl not installed - cannot create Excel")
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "PPh 23 Batch"
            
            # Styles
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            for col_idx, column in enumerate(self.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Convert and collect all structured data
            structured_list = []
            for idx, result in enumerate(batch_results, start=1):
                extracted_data = result.get('extracted_data', {})
                if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
                    structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
                    logger.info(f"✅ Using smart_mapped data for PPh 23 batch Excel row {idx}")
                elif 'structured_data' in extracted_data:
                    structured = extracted_data['structured_data']
                    logger.info(f"✅ Using structured_data for PPh 23 batch Excel row {idx}")
                else:
                    structured = extracted_data
                    logger.info(f"⚠️ Using raw extracted_data for PPh 23 batch Excel row {idx}")
                structured_list.append(structured)

            # Sort by Tanggal Dokumen Dasar
            from datetime import datetime
            def parse_date_for_sort(date_str):
                """Parse date for sorting, return datetime or min date if can't parse"""
                try:
                    formatted = self._format_date(date_str)
                    if formatted and formatted != 'N/A':
                        return datetime.strptime(formatted, '%d/%m/%Y')
                except:
                    pass
                return datetime.min

            structured_list.sort(key=lambda x: parse_date_for_sort(x.get('dokumen_dasar_tanggal', 'N/A')))

            # Write sorted data
            current_row = 2
            for structured in structured_list:
                row_data = [
                    structured.get('nomor', 'N/A'),
                    structured.get('masa_pajak', 'N/A'),
                    structured.get('sifat_pemotongan', 'N/A'),
                    structured.get('status', 'N/A'),
                    structured.get('penerima_npwp', 'N/A'),
                    structured.get('penerima_nama', 'N/A'),
                    structured.get('penerima_nitku', 'N/A'),
                    structured.get('jenis_pph', 'PPh 23'),
                    structured.get('kode_objek', 'N/A'),
                    structured.get('objek_pajak', 'N/A'),
                    self._format_rupiah(structured.get('dpp', 'N/A')),
                    structured.get('tarif', 'N/A'),
                    self._format_rupiah(structured.get('pph', 'N/A')),
                    structured.get('dokumen_dasar_jenis', 'N/A'),
                    self._format_date(structured.get('dokumen_dasar_tanggal', 'N/A')),
                    structured.get('pemotong_npwp', 'N/A'),
                    structured.get('pemotong_nitku', 'N/A'),
                    structured.get('pemotong_nama', 'N/A'),
                    self._format_date(structured.get('tanggal_potong', 'N/A')),
                    structured.get('penandatangan', 'N/A'),
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                current_row += 1
            
            # Adjust column widths
            column_widths = {
                'A': 15, 'B': 15, 'C': 20, 'D': 20, 'E': 20,
                'F': 25, 'G': 20, 'H': 12, 'I': 15, 'J': 30,
                'K': 15, 'L': 10, 'M': 15, 'N': 20, 'O': 15,
                'P': 20, 'Q': 20, 'R': 25, 'S': 15, 'T': 25,
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"✅ PPh 23 Batch Excel export created: {output_path} ({len(batch_results)} documents)")
            return True
            
        except Exception as e:
            logger.error(f"❌ PPh 23 Batch Excel export failed: {e}", exc_info=True)
            return False
    
    def batch_export_to_pdf(self, batch_id: str, results: list, output_path: str) -> bool:
        """Export multiple PPh 23 documents to formatted PDF (not table)"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, mm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            from datetime import datetime, timezone

            # Create PDF document
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=20*mm,
                leftMargin=20*mm,
                topMargin=15*mm,
                bottomMargin=15*mm
            )

            # Container for PDF elements
            story = []

            # Setup styles
            styles = getSampleStyleSheet()

            # Title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=4*mm,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            # Section header style (simple - no background)
            section_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Heading2'],
                fontSize=11,
                textColor=colors.HexColor('#1F4E78'),  # Blue text, no background
                spaceAfter=2*mm,
                spaceBefore=6*mm,
                fontName='Helvetica-Bold',
                borderWidth=0,
                borderPadding=0
            )

            # Field style (inline label + value)
            field_style = ParagraphStyle(
                'Field',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#333333'),
                spaceAfter=2*mm,
                leading=12
            )

            # Document number style (simple - no background)
            doc_number_style = ParagraphStyle(
                'DocNumber',
                parent=styles['Heading2'],
                fontSize=13,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=4*mm,
                spaceBefore=8*mm,
                fontName='Helvetica-Bold',
                alignment=TA_CENTER,
                borderWidth=0,
                borderPadding=0
            )

            # Add batch title
            story.append(Paragraph("BUKTI PEMOTONGAN PPh PASAL 23", title_style))
            story.append(Paragraph(f"BATCH EXPORT - {len(results)} Dokumen",
                                 ParagraphStyle('Subtitle', parent=styles['Normal'],
                                              fontSize=10, alignment=TA_CENTER,
                                              textColor=colors.HexColor('#666666'))))
            story.append(Spacer(1, 8*mm))

            # Process each result
            for idx, result in enumerate(results, start=1):
                # Get structured data from result
                extracted_data = result.get('extracted_data', {})

                # Priority: smart_mapped > structured_data > extracted_data itself
                if 'smart_mapped' in extracted_data and extracted_data['smart_mapped']:
                    structured = self._convert_smart_mapped_to_structured(extracted_data['smart_mapped'])
                    logger.info(f"✅ Using smart_mapped data for PPh 23 batch item {idx}")
                elif 'structured_data' in extracted_data:
                    structured = extracted_data['structured_data']
                    logger.info(f"✅ Using structured_data for PPh 23 batch item {idx}")
                else:
                    structured = extracted_data
                    logger.info(f"⚠️ Using raw extracted_data for PPh 23 batch item {idx}")

                # Document separator
                story.append(Paragraph(f"DOKUMEN {idx} dari {len(results)}", doc_number_style))
                story.append(Spacer(1, 3*mm))

                # Section 1: Informasi Dokumen
                story.append(Paragraph("INFORMASI DOKUMEN", section_style))
                story.append(Paragraph(f"<b>Nomor Bukti Potong:</b> {structured.get('nomor', 'N/A')}", field_style))
                story.append(Paragraph(f"<b>Masa Pajak:</b> {structured.get('masa_pajak', 'N/A')}", field_style))
                story.append(Paragraph(f"<b>Tanggal Pemotongan:</b> {structured.get('tanggal_potong', 'N/A')}", field_style))

                # Section 2: Identitas Penerima
                story.append(Paragraph("PENERIMA PENGHASILAN", section_style))
                story.append(Paragraph(f"<b>Nama:</b> {structured.get('penerima_nama', 'N/A')}", field_style))
                story.append(Paragraph(f"<b>NPWP/NIK:</b> {structured.get('penerima_npwp', 'N/A')}", field_style))

                # Section 3: Objek Pajak
                story.append(Paragraph("OBJEK PAJAK", section_style))
                story.append(Paragraph(f"<b>Kode:</b> {structured.get('kode_objek', 'N/A')}", field_style))
                story.append(Paragraph(f"<b>Uraian:</b> {structured.get('objek_pajak', 'N/A')}", field_style))

                # Section 4: Perhitungan Pajak (simple format - no table)
                story.append(Paragraph("PERHITUNGAN PAJAK", section_style))
                story.append(Paragraph(f"<b>Dasar Pengenaan Pajak (DPP):</b> Rp {structured.get('dpp', 'N/A')}", field_style))
                story.append(Paragraph(f"<b>Tarif:</b> {structured.get('tarif', 'N/A')}", field_style))

                # PPh Dipotong - highlighted
                pph_highlight = ParagraphStyle(
                    'PPh',
                    parent=field_style,
                    fontSize=10,
                    fontName='Helvetica-Bold',
                    textColor=colors.HexColor('#1F4E78'),
                    spaceAfter=4*mm
                )
                story.append(Paragraph(f"<b>PPh Dipotong:</b> Rp {structured.get('pph', 'N/A')}", pph_highlight))

                # Section 5: Pemotong (Compact)
                story.append(Paragraph("PEMOTONG PAJAK", section_style))
                story.append(Paragraph(f"<b>Nama:</b> {structured.get('pemotong_nama', 'N/A')}", field_style))
                story.append(Paragraph(f"<b>NPWP:</b> {structured.get('pemotong_npwp', 'N/A')}", field_style))

                # Add page break between documents (except last one)
                if idx < len(results):
                    story.append(PageBreak())

            # Footer on last page
            story.append(Spacer(1, 10*mm))
            footer_text = f"<i>Batch export {len(results)} dokumen PPh 23 dibuat oleh Doc-Scan AI pada {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M:%S UTC')}</i>"
            story.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Normal'],
                                                              fontSize=8, alignment=TA_CENTER,
                                                              textColor=colors.HexColor('#999999'))))

            # Build PDF
            doc.build(story)
            logger.info(f"✅ PPh 23 Batch PDF export created: {output_path} ({len(results)} documents)")
            return True

        except Exception as e:
            logger.error(f"❌ PPh 23 Batch PDF export failed: {e}", exc_info=True)
            return False
