"""
Excel Reader Service for Reconciliation Module
Reads and parses Excel exports from OCR scanning results

This service provides:
1. List available Excel exports by document type
2. Parse Faktur Pajak Excel files
3. Parse Rekening Koran Excel files
4. Parse PPh 21/23 Excel files
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from dataclasses import dataclass
from decimal import Decimal
import re

logger = logging.getLogger(__name__)

# Import MatchCandidate from matching_engine
try:
    from matching_engine import MatchCandidate
except ImportError:
    # Fallback if import fails (for testing)
    logger.warning("Could not import MatchCandidate from matching_engine")


# ==================== Data Classes ====================

@dataclass
class ExcelFileInfo:
    """Metadata about an Excel export file"""
    filename: str
    filepath: str
    document_type: str
    row_count: int
    file_size: int  # bytes
    created_at: datetime
    sheet_names: List[str]
    date_range: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            'filename': self.filename,
            'filepath': self.filepath,
            'document_type': self.document_type,
            'row_count': self.row_count,
            'file_size': self.file_size,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'sheet_names': self.sheet_names,
            'date_range': self.date_range
        }


@dataclass
class FakturPajakData:
    """Parsed data from Faktur Pajak Excel with Buyer & Seller information"""
    row_number: int
    tanggal: str
    nomor_faktur: str
    dpp: Decimal
    ppn: Decimal
    total: Decimal
    # Seller (Penjual/PKP yang menjual)
    nama_seller: Optional[str] = None
    alamat_seller: Optional[str] = None
    npwp_seller: Optional[str] = None
    # Buyer (Pembeli/PKP yang membeli)
    nama_buyer: Optional[str] = None
    alamat_buyer: Optional[str] = None
    npwp_buyer: Optional[str] = None
    # Additional fields
    invoice_number: Optional[str] = None
    # Legacy fields for backward compatibility
    nama: Optional[str] = None
    npwp: Optional[str] = None
    alamat: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            'row_number': self.row_number,
            'tanggal': self.tanggal,
            'nomor_faktur': self.nomor_faktur,
            'dpp': float(self.dpp) if self.dpp else 0.0,
            'ppn': float(self.ppn) if self.ppn else 0.0,
            'total': float(self.total) if self.total else 0.0,
            'nama_seller': self.nama_seller,
            'alamat_seller': self.alamat_seller,
            'npwp_seller': self.npwp_seller,
            'nama_buyer': self.nama_buyer,
            'alamat_buyer': self.alamat_buyer,
            'npwp_buyer': self.npwp_buyer,
            'invoice_number': self.invoice_number,
            # Legacy fields for backward compatibility
            'nama': self.nama or self.nama_seller,
            'npwp': self.npwp or self.npwp_seller,
            'alamat': self.alamat or self.alamat_seller
        }


@dataclass
class RekeningKoranData:
    """Parsed data from Rekening Koran Excel"""
    row_number: int
    tanggal: str
    nilai_masuk: Decimal  # Kredit
    nilai_keluar: Decimal  # Debit
    saldo: Decimal
    keterangan: str
    sumber_masuk: Optional[str] = None
    tujuan_keluar: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            'row_number': self.row_number,
            'tanggal': self.tanggal,
            'nilai_masuk': float(self.nilai_masuk) if self.nilai_masuk else 0.0,
            'nilai_keluar': float(self.nilai_keluar) if self.nilai_keluar else 0.0,
            'saldo': float(self.saldo) if self.saldo else 0.0,
            'keterangan': self.keterangan,
            'sumber_masuk': self.sumber_masuk,
            'tujuan_keluar': self.tujuan_keluar
        }


@dataclass
class PPhData:
    """Parsed data from PPh 21/23 Excel"""
    row_number: int
    nomor: str
    masa_pajak: str
    npwp_nik: str
    nama_penerima: str
    jumlah_bruto: Decimal
    pph_dipotong: Decimal

    def to_dict(self) -> Dict[str, Any]:
        return {
            'row_number': self.row_number,
            'nomor': self.nomor,
            'masa_pajak': self.masa_pajak,
            'npwp_nik': self.npwp_nik,
            'nama_penerima': self.nama_penerima,
            'jumlah_bruto': float(self.jumlah_bruto) if self.jumlah_bruto else 0.0,
            'pph_dipotong': float(self.pph_dipotong) if self.pph_dipotong else 0.0
        }


# ==================== Excel Reader Service ====================

class ExcelReaderService:
    """Service to read and parse Excel exports from OCR results"""

    def __init__(self, exports_dir: Optional[str] = None):
        # Use absolute path from current file location if not specified
        if exports_dir is None:
            self.exports_dir = Path(__file__).parent / "exports"
        else:
            self.exports_dir = Path(exports_dir)

        if not self.exports_dir.exists():
            logger.warning(f"Exports directory not found: {self.exports_dir}")
            self.exports_dir.mkdir(parents=True, exist_ok=True)

    # ==================== List Available Exports ====================

    def list_available_exports(self, document_type: Optional[str] = None) -> List[ExcelFileInfo]:
        """
        List all available Excel exports, optionally filtered by document type

        Args:
            document_type: Filter by type (faktur_pajak, rekening_koran, pph21, pph23, batch)

        Returns:
            List of ExcelFileInfo objects
        """
        excel_files = []

        for excel_file in self.exports_dir.glob("*.xlsx"):
            # Skip temporary Excel files
            if excel_file.name.startswith('~'):
                continue

            try:
                # Determine document type from filename
                detected_type = self._detect_document_type(excel_file.name)

                # Filter if document_type specified
                if document_type and detected_type != document_type:
                    continue

                # Get file metadata
                file_info = self._get_file_metadata(excel_file, detected_type)
                excel_files.append(file_info)

            except Exception as e:
                logger.warning(f"Error reading {excel_file.name}: {e}")
                continue

        # Sort by created_at descending (newest first)
        excel_files.sort(key=lambda x: x.created_at, reverse=True)

        logger.info(f"Found {len(excel_files)} Excel exports" +
                   (f" for type '{document_type}'" if document_type else ""))

        return excel_files

    def _detect_document_type(self, filename: str) -> str:
        """Detect document type from filename"""
        filename_lower = filename.lower()

        if 'batch' in filename_lower:
            return 'batch'
        elif 'faktur' in filename_lower or 'fp' in filename_lower:
            return 'faktur_pajak'
        elif any(keyword in filename_lower for keyword in ['rek', 'bank', 'estat', 'bni', 'bca', 'mandiri', 'permata', 'bri']):
            return 'rekening_koran'
        elif 'pph 21' in filename_lower or 'pph21' in filename_lower:
            return 'pph21'
        elif 'pph 23' in filename_lower or 'pph23' in filename_lower:
            return 'pph23'
        else:
            return 'unknown'

    def _get_file_metadata(self, filepath: Path, document_type: str) -> ExcelFileInfo:
        """Extract metadata from Excel file"""
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        # Get sheet names
        sheet_names = wb.sheetnames

        # Count rows in first sheet
        ws = wb[sheet_names[0]]
        row_count = ws.max_row

        # Get file stats
        file_stats = filepath.stat()

        wb.close()

        return ExcelFileInfo(
            filename=filepath.name,
            filepath=str(filepath),
            document_type=document_type,
            row_count=row_count,
            file_size=file_stats.st_size,
            created_at=datetime.fromtimestamp(file_stats.st_mtime),
            sheet_names=sheet_names,
            date_range=None  # Will be determined when parsing
        )

    # ==================== Parse Faktur Pajak ====================

    def read_faktur_pajak(self, excel_path: str) -> List[FakturPajakData]:
        """
        Parse Faktur Pajak Excel file

        Expected structure:
        - Sheet: "Faktur Pajak" or "Batch Faktur Pajak"
        - Columns: Nama, Tgl, NPWP, Nomor Faktur, Alamat, DPP, PPN, Total, Invoice

        Args:
            excel_path: Path to Excel file

        Returns:
            List of parsed FakturPajakData objects
        """
        logger.info(f"📄 Reading Faktur Pajak: {excel_path}")

        filepath = Path(excel_path)
        if not filepath.exists():
            # Try relative to exports_dir
            filepath = self.exports_dir / excel_path

        if not filepath.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = openpyxl.load_workbook(filepath, data_only=True)

        # Find the correct sheet
        sheet_name = None
        for name in wb.sheetnames:
            if 'Faktur Pajak' in name:
                sheet_name = name
                break

        if not sheet_name:
            wb.close()
            raise ValueError(f"Sheet 'Faktur Pajak' not found in {filepath.name}")

        ws = wb[sheet_name]

        # Find header row (look for row with multiple filled cells)
        header_row_idx = self._find_header_row(ws)

        if not header_row_idx:
            wb.close()
            raise ValueError("Could not find header row in Excel")

        # Parse data
        fakturs = []

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row = ws[row_idx]

            # Skip empty rows
            if all(cell.value is None for cell in row):
                continue

            try:
                # Extract data (assuming standard column order)
                faktur = self._parse_faktur_row(row, row_idx)
                if faktur:
                    fakturs.append(faktur)

            except Exception as e:
                logger.warning(f"Error parsing row {row_idx}: {e}")
                continue

        wb.close()

        logger.info(f"✅ Parsed {len(fakturs)} Faktur Pajak records")
        return fakturs

    def _find_header_row(self, worksheet) -> Optional[int]:
        """Find header row in worksheet"""
        for row_idx in range(1, min(10, worksheet.max_row + 1)):
            row_values = [cell.value for cell in worksheet[row_idx]]
            non_none = sum(1 for v in row_values if v)

            # Header should have at least 5 non-empty cells
            if non_none >= 5:
                return row_idx

        return None

    def _parse_faktur_row(self, row, row_number: int) -> Optional[FakturPajakData]:
        """Parse single Faktur Pajak row with buyer & seller support"""
        # Try new format first: Nama Seller, Alamat Seller, NPWP Seller, Nama Buyer, Alamat Buyer, NPWP Buyer, Tgl, Nomor Faktur, DPP, PPN, Total
        # Fallback to old format: Nama, Tgl, NPWP, Nomor Faktur, Alamat, DPP, PPN, Total, Invoice
        
        # Check if this is new format (has buyer/seller columns) by looking at row length
        has_buyer_seller = len(row) >= 11
        
        if has_buyer_seller:
            # New format with buyer & seller
            nama_seller = str(row[0].value or '').strip()
            alamat_seller = str(row[1].value or '').strip()
            npwp_seller = str(row[2].value or '').strip()
            nama_buyer = str(row[3].value or '').strip()
            alamat_buyer = str(row[4].value or '').strip()
            npwp_buyer = str(row[5].value or '').strip()
            tanggal = str(row[6].value or '').strip()
            nomor_faktur = str(row[7].value or '').strip()
            dpp = self._parse_amount(row[8].value if len(row) > 8 else None)
            ppn = self._parse_amount(row[9].value if len(row) > 9 else None)
            total = self._parse_amount(row[10].value if len(row) > 10 else None)
            invoice_number = str(row[11].value or '').strip() if len(row) > 11 else ''
            
            # Skip if essential fields are missing
            if not tanggal and not nomor_faktur:
                return None
            
            return FakturPajakData(
                row_number=row_number,
                tanggal=tanggal,
                nomor_faktur=nomor_faktur,
                dpp=dpp,
                ppn=ppn,
                total=total,
                nama_seller=nama_seller,
                alamat_seller=alamat_seller,
                npwp_seller=npwp_seller,
                nama_buyer=nama_buyer,
                alamat_buyer=alamat_buyer,
                npwp_buyer=npwp_buyer,
                invoice_number=invoice_number,
                # Legacy fields
                nama=nama_seller,
                npwp=npwp_seller,
                alamat=alamat_seller
            )
        else:
            # Old format (backward compatibility)
            nama = str(row[0].value or '').strip()
            tanggal = str(row[1].value or '').strip()
            npwp = str(row[2].value or '').strip()
            nomor_faktur = str(row[3].value or '').strip()
            alamat = str(row[4].value or '').strip() if len(row) > 4 else ''
            dpp = self._parse_amount(row[5].value if len(row) > 5 else None)
            ppn = self._parse_amount(row[6].value if len(row) > 6 else None)
            total = self._parse_amount(row[7].value if len(row) > 7 else None)
            invoice_number = str(row[8].value or '').strip() if len(row) > 8 else ''
            
            # Skip if essential fields are missing
            if not nama and not nomor_faktur:
                return None
            
            return FakturPajakData(
                row_number=row_number,
                tanggal=tanggal,
                nomor_faktur=nomor_faktur,
                dpp=dpp,
                ppn=ppn,
                total=total,
                nama_seller=nama,
                alamat_seller=alamat,
                npwp_seller=npwp,
                invoice_number=invoice_number,
                # Legacy fields
                nama=nama,
                npwp=npwp,
                alamat=alamat
            )

    def _parse_amount(self, value: Any) -> Decimal:
        """Parse amount from Excel cell (handles 'Rp', thousands separators)"""
        if value is None:
            return Decimal('0')

        # If already a number
        if isinstance(value, (int, float)):
            return Decimal(str(value))

        # If string, clean it
        if isinstance(value, str):
            # Remove "Rp", spaces, and dots (thousands separator)
            cleaned = value.replace('Rp', '').replace(' ', '').replace('.', '').replace(',', '.')
            cleaned = cleaned.strip()

            if not cleaned or cleaned == '-':
                return Decimal('0')

            try:
                return Decimal(cleaned)
            except:
                return Decimal('0')

        return Decimal('0')

    # ==================== Parse Rekening Koran ====================

    def read_rekening_koran(self, excel_path: str) -> List[RekeningKoranData]:
        """
        Parse Rekening Koran Excel file

        Expected structure:
        - Sheet: "Rekening Koran"
        - Columns: Tanggal, Nilai Uang Masuk, Nilai Uang Keluar, Saldo, Keterangan

        Args:
            excel_path: Path to Excel file

        Returns:
            List of parsed RekeningKoranData objects
        """
        logger.info(f"📄 Reading Rekening Koran: {excel_path}")

        filepath = Path(excel_path)
        if not filepath.exists():
            filepath = self.exports_dir / excel_path

        if not filepath.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = openpyxl.load_workbook(filepath, data_only=True)

        # Find Rekening Koran sheet
        sheet_name = None
        for name in wb.sheetnames:
            if 'Rekening Koran' in name:
                sheet_name = name
                break

        if not sheet_name:
            wb.close()
            raise ValueError(f"Sheet 'Rekening Koran' not found")

        ws = wb[sheet_name]

        # Find header row
        header_row_idx = self._find_header_row(ws)

        if not header_row_idx:
            wb.close()
            raise ValueError("Could not find header row")

        # Parse data
        transactions = []

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row = ws[row_idx]

            # Skip empty rows
            if all(cell.value is None for cell in row):
                continue

            try:
                transaction = self._parse_rekening_row(row, row_idx)
                if transaction:
                    transactions.append(transaction)

            except Exception as e:
                logger.warning(f"Error parsing row {row_idx}: {e}")
                continue

        wb.close()

        logger.info(f"✅ Parsed {len(transactions)} Rekening Koran transactions")
        return transactions

    def _parse_rekening_row(self, row, row_number: int) -> Optional[RekeningKoranData]:
        """Parse single Rekening Koran row"""
        # Expected: Tanggal, Nilai Uang Masuk, Nilai Uang Keluar, Saldo, Sumber, Tujuan, Keterangan

        tanggal = str(row[0].value or '').strip()
        nilai_masuk = self._parse_amount(row[1].value if len(row) > 1 else None)
        nilai_keluar = self._parse_amount(row[2].value if len(row) > 2 else None)
        saldo = self._parse_amount(row[3].value if len(row) > 3 else None)
        sumber_masuk = str(row[4].value or '').strip() if len(row) > 4 else ''
        tujuan_keluar = str(row[5].value or '').strip() if len(row) > 5 else ''
        keterangan = str(row[6].value or '').strip() if len(row) > 6 else ''

        # Skip if no date
        if not tanggal:
            return None

        return RekeningKoranData(
            row_number=row_number,
            tanggal=tanggal,
            nilai_masuk=nilai_masuk,
            nilai_keluar=nilai_keluar,
            saldo=saldo,
            keterangan=keterangan,
            sumber_masuk=sumber_masuk,
            tujuan_keluar=tujuan_keluar
        )

    # ==================== Parse PPh ====================

    def read_pph(self, excel_path: str, pph_type: str = 'pph21') -> List[PPhData]:
        """
        Parse PPh 21/23 Excel file

        Expected structure:
        - Sheet: "PPh 21" or "PPh 23"
        - Columns: Nomor, Masa Pajak, NPWP/NIK Penerima, Nama, Jumlah Bruto, PPh Dipotong

        Args:
            excel_path: Path to Excel file
            pph_type: 'pph21' or 'pph23'

        Returns:
            List of parsed PPhData objects
        """
        logger.info(f"📄 Reading {pph_type.upper()}: {excel_path}")

        filepath = Path(excel_path)
        if not filepath.exists():
            filepath = self.exports_dir / excel_path

        if not filepath.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = openpyxl.load_workbook(filepath, data_only=True)

        # Find PPh sheet
        sheet_name = None
        search_name = "PPh 21" if pph_type == 'pph21' else "PPh 23"

        for name in wb.sheetnames:
            if search_name in name:
                sheet_name = name
                break

        if not sheet_name:
            wb.close()
            raise ValueError(f"Sheet '{search_name}' not found")

        ws = wb[sheet_name]

        # Find header row
        header_row_idx = self._find_header_row(ws)

        if not header_row_idx:
            wb.close()
            raise ValueError("Could not find header row")

        # Parse data
        pph_records = []

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row = ws[row_idx]

            if all(cell.value is None for cell in row):
                continue

            try:
                pph_record = self._parse_pph_row(row, row_idx)
                if pph_record:
                    pph_records.append(pph_record)

            except Exception as e:
                logger.warning(f"Error parsing row {row_idx}: {e}")
                continue

        wb.close()

        logger.info(f"✅ Parsed {len(pph_records)} {pph_type.upper()} records")
        return pph_records

    def _parse_pph_row(self, row, row_number: int) -> Optional[PPhData]:
        """Parse single PPh row"""
        # Expected: Nomor, Masa Pajak, Sifat, Status, NPWP/NIK, Nama, ..., Jumlah Bruto, PPh

        nomor = str(row[0].value or '').strip()
        masa_pajak = str(row[1].value or '').strip()
        npwp_nik = str(row[4].value or '').strip() if len(row) > 4 else ''
        nama_penerima = str(row[5].value or '').strip() if len(row) > 5 else ''

        # Amounts might be in different columns depending on PPh type
        # Try to find them
        jumlah_bruto = Decimal('0')
        pph_dipotong = Decimal('0')

        # Look for amount columns (usually near the end)
        for i in range(len(row) - 1, max(5, len(row) - 10), -1):
            if row[i].value:
                amount = self._parse_amount(row[i].value)
                if amount > 0:
                    if pph_dipotong == 0:
                        pph_dipotong = amount
                    elif jumlah_bruto == 0:
                        jumlah_bruto = amount

        if not nomor:
            return None

        return PPhData(
            row_number=row_number,
            nomor=nomor,
            masa_pajak=masa_pajak,
            npwp_nik=npwp_nik,
            nama_penerima=nama_penerima,
            jumlah_bruto=jumlah_bruto,
            pph_dipotong=pph_dipotong
        )

    # ==================== Conversion to MatchCandidates ====================

    def convert_faktur_to_candidates(
        self,
        faktur_data: List[FakturPajakData],
        source_file: str
    ) -> List[MatchCandidate]:
        """
        Convert Faktur Pajak data to MatchCandidate objects for matching engine

        Args:
            faktur_data: List of FakturPajakData from read_faktur_pajak()
            source_file: Source Excel filename

        Returns:
            List of MatchCandidate objects
        """
        candidates = []

        for faktur in faktur_data:
            candidate = MatchCandidate(
                source_type='faktur_pajak',
                source_file=source_file,
                source_row=faktur.row_number,
                date=faktur.tanggal,
                amount=faktur.total,
                vendor_name=faktur.nama_seller or faktur.nama or '',
                reference=faktur.nomor_faktur,
                raw_data={
                    # Seller information
                    'nama_seller': faktur.nama_seller,
                    'alamat_seller': faktur.alamat_seller,
                    'npwp_seller': faktur.npwp_seller,
                    # Buyer information
                    'nama_buyer': faktur.nama_buyer,
                    'alamat_buyer': faktur.alamat_buyer,
                    'npwp_buyer': faktur.npwp_buyer,
                    # Transaction info
                    'tanggal': faktur.tanggal,
                    'nomor_faktur': faktur.nomor_faktur,
                    'dpp': float(faktur.dpp),
                    'ppn': float(faktur.ppn),
                    'total': float(faktur.total),
                    'invoice_number': faktur.invoice_number,
                    # Legacy fields for backward compatibility
                    'nama': faktur.nama or faktur.nama_seller,
                    'npwp': faktur.npwp or faktur.npwp_seller,
                    'alamat': faktur.alamat or faktur.alamat_seller
                }
            )
            candidates.append(candidate)

        logger.info(f"✅ Converted {len(candidates)} Faktur Pajak records to MatchCandidates")
        return candidates

    def convert_rekening_to_candidates(
        self,
        rekening_data: List[RekeningKoranData],
        source_file: str
    ) -> List[MatchCandidate]:
        """
        Convert Rekening Koran data to MatchCandidate objects for matching engine

        Args:
            rekening_data: List of RekeningKoranData from read_rekening_koran()
            source_file: Source Excel filename

        Returns:
            List of MatchCandidate objects
        """
        candidates = []

        for rekening in rekening_data:
            # Use the larger amount (debit or credit) as the primary amount
            amount = max(rekening.nilai_masuk, rekening.nilai_keluar)

            # Extract vendor name from keterangan (description)
            # This is a simplified approach - may need AI enhancement later
            vendor_name = rekening.keterangan[:50] if rekening.keterangan else ''

            candidate = MatchCandidate(
                source_type='rekening_koran',
                source_file=source_file,
                source_row=rekening.row_number,
                date=rekening.tanggal,
                amount=amount,
                vendor_name=vendor_name,
                reference=rekening.keterangan or '',
                raw_data={
                    'tanggal': rekening.tanggal,
                    'nilai_masuk': float(rekening.nilai_masuk),
                    'nilai_keluar': float(rekening.nilai_keluar),
                    'saldo': float(rekening.saldo),
                    'keterangan': rekening.keterangan
                }
            )
            candidates.append(candidate)

        logger.info(f"✅ Converted {len(candidates)} Rekening Koran records to MatchCandidates")
        return candidates

    def convert_pph_to_candidates(
        self,
        pph_data: List[PPhData],
        source_file: str,
        pph_type: str = 'pph21'
    ) -> List[MatchCandidate]:
        """
        Convert PPh data to MatchCandidate objects for matching engine

        Args:
            pph_data: List of PPhData from read_pph()
            source_file: Source Excel filename
            pph_type: Type of PPh ('pph21' or 'pph23')

        Returns:
            List of MatchCandidate objects
        """
        candidates = []

        for pph in pph_data:
            # Extract date from masa_pajak (e.g., "09-2025" -> approximate date)
            # This is a simplified approach - may need enhancement
            date = pph.masa_pajak if pph.masa_pajak else ''

            candidate = MatchCandidate(
                source_type=pph_type,
                source_file=source_file,
                source_row=pph.row_number,
                date=date,
                amount=pph.pph_dipotong,  # Use PPh amount as primary amount
                vendor_name=pph.nama_penerima,
                reference=pph.nomor,
                raw_data={
                    'nomor': pph.nomor,
                    'masa_pajak': pph.masa_pajak,
                    'npwp_nik': pph.npwp_nik,
                    'nama_penerima': pph.nama_penerima,
                    'jumlah_bruto': float(pph.jumlah_bruto),
                    'pph_dipotong': float(pph.pph_dipotong)
                }
            )
            candidates.append(candidate)

        logger.info(f"✅ Converted {len(candidates)} {pph_type.upper()} records to MatchCandidates")
        return candidates


# ==================== Singleton Instance ====================

# Create global instance
excel_reader_service = ExcelReaderService()
