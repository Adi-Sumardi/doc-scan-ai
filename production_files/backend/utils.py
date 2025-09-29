import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from pathlib import Path
import json
from typing import Dict, Any, List, Union

def create_excel_export(data: Union[Dict[str, Any], List[Dict[str, Any]]], output_path: Path, is_batch: bool = False):
    """Create Excel export from scan results"""
    
    wb = Workbook()
    
    if is_batch:
        # Handle batch export (multiple results)
        ws = wb.active
        ws.title = "Batch Results Summary"
        
        # Create summary sheet
        headers = ["Document Type", "Original Filename", "Confidence", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add data rows
        for row, result in enumerate(data, 2):
            ws.cell(row=row, column=1, value=result["document_type"])
            ws.cell(row=row, column=2, value=result["original_filename"])
            ws.cell(row=row, column=3, value=f"{result['confidence']*100:.1f}%")
            ws.cell(row=row, column=4, value="Completed")
        
        # Create individual sheets for each document
        for result in data:
            create_document_sheet(wb, result)
    
    else:
        # Handle single document export
        ws = wb.active
        ws.title = "Document Data"
        create_document_sheet(wb, data, is_main_sheet=True)
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)

def create_document_sheet(wb: Workbook, result: Dict[str, Any], is_main_sheet: bool = False):
    """Create a worksheet for a single document result"""
    
    if not is_main_sheet:
        ws = wb.create_sheet(title=result["original_filename"][:30])
    else:
        ws = wb.active
    
    row = 1
    
    # Document header
    ws.cell(row=row, column=1, value="Document Information").font = Font(size=14, bold=True)
    row += 1
    
    ws.cell(row=row, column=1, value="Filename:")
    ws.cell(row=row, column=2, value=result["original_filename"])
    row += 1
    
    ws.cell(row=row, column=1, value="Document Type:")
    ws.cell(row=row, column=2, value=result["document_type"])
    row += 1
    
    ws.cell(row=row, column=1, value="Confidence:")
    ws.cell(row=row, column=2, value=f"{result['confidence']*100:.1f}%")
    row += 2
    
    # Extracted data
    ws.cell(row=row, column=1, value="Extracted Data").font = Font(size=14, bold=True)
    row += 1
    
    extracted_data = result["extracted_data"]
    
    if result["document_type"] == "faktur_pajak":
        # Special handling for Faktur Pajak
        row = add_faktur_pajak_data(ws, extracted_data, row)
    elif result["document_type"] in ["pph21", "pph23"]:
        # Handle PPh documents
        row = add_pph_data(ws, extracted_data, row)
    elif result["document_type"] == "rekening_koran":
        # Handle bank statements
        row = add_rekening_koran_data(ws, extracted_data, row)
    elif result["document_type"] == "invoice":
        # Handle invoices
        row = add_invoice_data(ws, extracted_data, row)
    else:
        # Generic data handling
        row = add_generic_data(ws, extracted_data, row)

def add_faktur_pajak_data(ws, data, start_row):
    """Add Faktur Pajak specific data to worksheet"""
    row = start_row
    
    # Faktur Pajak Masukan
    ws.cell(row=row, column=1, value="Faktur Pajak Masukan").font = Font(bold=True, color="0066CC")
    row += 1
    
    masukan = data.get("masukan", {})
    for key, value in masukan.items():
        ws.cell(row=row, column=1, value=key.replace("_", " ").title())
        ws.cell(row=row, column=2, value=str(value))
        row += 1
    
    row += 1
    
    # Faktur Pajak Keluaran
    ws.cell(row=row, column=1, value="Faktur Pajak Keluaran").font = Font(bold=True, color="009900")
    row += 1
    
    keluaran = data.get("keluaran", {})
    for key, value in keluaran.items():
        ws.cell(row=row, column=1, value=key.replace("_", " ").title())
        ws.cell(row=row, column=2, value=str(value))
        row += 1
    
    return row

def add_pph_data(ws, data, start_row):
    """Add PPh document data to worksheet"""
    row = start_row
    
    for key, value in data.items():
        ws.cell(row=row, column=1, value=key.replace("_", " ").title())
        ws.cell(row=row, column=2, value=str(value))
        row += 1
    
    return row

def add_rekening_koran_data(ws, data, start_row):
    """Add bank statement data to worksheet"""
    row = start_row
    
    # Account info
    account_info = ["nomor_rekening", "nama_pemilik", "periode", "saldo_awal", "saldo_akhir", "total_transaksi"]
    for key in account_info:
        if key in data:
            ws.cell(row=row, column=1, value=key.replace("_", " ").title())
            ws.cell(row=row, column=2, value=str(data[key]))
            row += 1
    
    row += 2
    
    # Transactions table
    ws.cell(row=row, column=1, value="Transactions").font = Font(bold=True)
    row += 1
    
    if "transaksi" in data:
        # Headers
        headers = ["Tanggal", "Keterangan", "Debit", "Kredit", "Saldo"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        row += 1
        
        # Transaction data
        for transaction in data["transaksi"]:
            ws.cell(row=row, column=1, value=transaction.get("tanggal", ""))
            ws.cell(row=row, column=2, value=transaction.get("keterangan", ""))
            ws.cell(row=row, column=3, value=transaction.get("debit", 0))
            ws.cell(row=row, column=4, value=transaction.get("kredit", 0))
            ws.cell(row=row, column=5, value=transaction.get("saldo", 0))
            row += 1
    
    return row

def add_invoice_data(ws, data, start_row):
    """Add invoice data to worksheet"""
    row = start_row
    
    # Invoice header
    header_fields = ["nomor_invoice", "tanggal", "due_date"]
    for field in header_fields:
        if field in data:
            ws.cell(row=row, column=1, value=field.replace("_", " ").title())
            ws.cell(row=row, column=2, value=str(data[field]))
            row += 1
    
    row += 1
    
    # From/To information
    if "dari" in data:
        ws.cell(row=row, column=1, value="From").font = Font(bold=True)
        row += 1
        for key, value in data["dari"].items():
            ws.cell(row=row, column=1, value=key.title())
            ws.cell(row=row, column=2, value=str(value))
            row += 1
    
    row += 1
    
    if "kepada" in data:
        ws.cell(row=row, column=1, value="To").font = Font(bold=True)
        row += 1
        for key, value in data["kepada"].items():
            ws.cell(row=row, column=1, value=key.title())
            ws.cell(row=row, column=2, value=str(value))
            row += 1
    
    row += 2
    
    # Items table
    if "items" in data:
        ws.cell(row=row, column=1, value="Items").font = Font(bold=True)
        row += 1
        
        # Headers
        headers = ["Item", "Description", "Qty", "Unit Price", "Subtotal"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        row += 1
        
        # Items data
        for item in data["items"]:
            ws.cell(row=row, column=1, value=item.get("item", ""))
            ws.cell(row=row, column=2, value=item.get("deskripsi", ""))
            ws.cell(row=row, column=3, value=item.get("qty", 0))
            ws.cell(row=row, column=4, value=item.get("harga_satuan", 0))
            ws.cell(row=row, column=5, value=item.get("subtotal", 0))
            row += 1
        
        row += 1
        
        # Totals
        totals = ["subtotal", "ppn", "total"]
        for field in totals:
            if field in data:
                ws.cell(row=row, column=4, value=field.upper())
                ws.cell(row=row, column=5, value=data[field])
                if field == "total":
                    ws.cell(row=row, column=4).font = Font(bold=True)
                    ws.cell(row=row, column=5).font = Font(bold=True)
                row += 1
    
    return row

def add_generic_data(ws, data, start_row):
    """Add generic data to worksheet"""
    row = start_row
    
    def add_data_recursive(data, indent=0):
        nonlocal row
        for key, value in data.items():
            if isinstance(value, dict):
                ws.cell(row=row, column=1 + indent, value=str(key).replace("_", " ").title()).font = Font(bold=True)
                row += 1
                add_data_recursive(value, indent + 1)
            elif isinstance(value, list):
                ws.cell(row=row, column=1 + indent, value=str(key).replace("_", " ").title()).font = Font(bold=True)
                row += 1
                for i, item in enumerate(value):
                    if isinstance(item, dict):
                        ws.cell(row=row, column=2 + indent, value=f"Item {i+1}").font = Font(italic=True)
                        row += 1
                        add_data_recursive(item, indent + 2)
                    else:
                        ws.cell(row=row, column=2 + indent, value=str(item))
                        row += 1
            else:
                ws.cell(row=row, column=1 + indent, value=str(key).replace("_", " ").title())
                ws.cell(row=row, column=2 + indent, value=str(value))
                row += 1
    
    add_data_recursive(data)
    return row

def create_pdf_export(result: Dict[str, Any], output_path: Path):
    """Create PDF export from scan result"""
    
    doc = SimpleDocTemplate(str(output_path), pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        textColor=colors.HexColor('#366092')
    )
    
    story.append(Paragraph("Document Scan Results", title_style))
    story.append(Spacer(1, 12))
    
    # Document info
    info_data = [
        ["Filename:", result["original_filename"]],
        ["Document Type:", result["document_type"]],
        ["Confidence:", f"{result['confidence']*100:.1f}%"],
        ["Processed:", result["created_at"][:10]]
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Extracted data
    story.append(Paragraph("Extracted Data", styles['Heading2']))
    story.append(Spacer(1, 12))
    
    extracted_data = result["extracted_data"]
    
    if result["document_type"] == "faktur_pajak":
        add_faktur_pajak_pdf(story, extracted_data, styles)
    elif result["document_type"] in ["pph21", "pph23"]:
        add_pph_pdf(story, extracted_data, styles)
    elif result["document_type"] == "rekening_koran":
        add_rekening_koran_pdf(story, extracted_data, styles)
    elif result["document_type"] == "invoice":
        add_invoice_pdf(story, extracted_data, styles)
    else:
        add_generic_pdf(story, extracted_data, styles)
    
    doc.build(story)

def add_faktur_pajak_pdf(story, data, styles):
    """Add Faktur Pajak data to PDF"""
    
    # Masukan
    story.append(Paragraph("Faktur Pajak Masukan", styles['Heading3']))
    masukan_data = []
    for key, value in data.get("masukan", {}).items():
        masukan_data.append([key.replace("_", " ").title(), str(value)])
    
    if masukan_data:
        table = Table(masukan_data, colWidths=[2.5*inch, 3.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6f3ff')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]))
        story.append(table)
    
    story.append(Spacer(1, 15))
    
    # Keluaran
    story.append(Paragraph("Faktur Pajak Keluaran", styles['Heading3']))
    keluaran_data = []
    for key, value in data.get("keluaran", {}).items():
        keluaran_data.append([key.replace("_", " ").title(), str(value)])
    
    if keluaran_data:
        table = Table(keluaran_data, colWidths=[2.5*inch, 3.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6ffe6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]))
        story.append(table)

def add_pph_pdf(story, data, styles):
    """Add PPh data to PDF"""
    table_data = []
    for key, value in data.items():
        table_data.append([key.replace("_", " ").title(), str(value)])
    
    if table_data:
        table = Table(table_data, colWidths=[2.5*inch, 3.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fff9e6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]))
        story.append(table)

def add_rekening_koran_pdf(story, data, styles):
    """Add bank statement data to PDF"""
    # Account info
    account_data = []
    account_fields = ["nomor_rekening", "nama_pemilik", "periode", "saldo_awal", "saldo_akhir"]
    for field in account_fields:
        if field in data:
            account_data.append([field.replace("_", " ").title(), str(data[field])])
    
    if account_data:
        table = Table(account_data, colWidths=[2.5*inch, 3.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f8ff')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]))
        story.append(table)
    
    story.append(Spacer(1, 15))
    
    # Transactions (show first 10)
    if "transaksi" in data and data["transaksi"]:
        story.append(Paragraph("Recent Transactions", styles['Heading3']))
        
        trans_data = [["Date", "Description", "Debit", "Credit", "Balance"]]
        for trans in data["transaksi"][:10]:  # Limit to first 10
            trans_data.append([
                trans.get("tanggal", ""),
                trans.get("keterangan", "")[:30] + "..." if len(trans.get("keterangan", "")) > 30 else trans.get("keterangan", ""),
                f"Rp {trans.get('debit', 0):,}" if trans.get('debit', 0) > 0 else "",
                f"Rp {trans.get('kredit', 0):,}" if trans.get('kredit', 0) > 0 else "",
                f"Rp {trans.get('saldo', 0):,}"
            ])
        
        trans_table = Table(trans_data, colWidths=[1*inch, 2*inch, 1*inch, 1*inch, 1.2*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(trans_table)

def add_invoice_pdf(story, data, styles):
    """Add invoice data to PDF"""
    # Invoice header
    header_data = []
    header_fields = ["nomor_invoice", "tanggal", "due_date"]
    for field in header_fields:
        if field in data:
            header_data.append([field.replace("_", " ").title(), str(data[field])])
    
    if header_data:
        table = Table(header_data, colWidths=[2*inch, 4*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5f5f5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]))
        story.append(table)
    
    story.append(Spacer(1, 15))
    
    # Items table
    if "items" in data and data["items"]:
        story.append(Paragraph("Items", styles['Heading3']))
        
        items_data = [["Item", "Qty", "Unit Price", "Subtotal"]]
        for item in data["items"]:
            items_data.append([
                item.get("item", ""),
                str(item.get("qty", 0)),
                f"Rp {item.get('harga_satuan', 0):,}",
                f"Rp {item.get('subtotal', 0):,}"
            ])
        
        # Add totals
        items_data.append(["", "", "Subtotal:", f"Rp {data.get('subtotal', 0):,}"])
        items_data.append(["", "", "PPN:", f"Rp {data.get('ppn', 0):,}"])
        items_data.append(["", "", "TOTAL:", f"Rp {data.get('total', 0):,}"])
        
        items_table = Table(items_data, colWidths=[2*inch, 1*inch, 1.5*inch, 1.5*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(items_table)

def add_generic_pdf(story, data, styles):
    """Add generic data to PDF"""
    def flatten_data(data, prefix=""):
        flat_data = []
        for key, value in data.items():
            full_key = f"{prefix}{key}" if prefix else key
            if isinstance(value, dict):
                flat_data.extend(flatten_data(value, f"{full_key}."))
            elif isinstance(value, list):
                flat_data.append([full_key.replace("_", " ").title(), f"{len(value)} items"])
            else:
                flat_data.append([full_key.replace("_", " ").title(), str(value)])
        return flat_data
    
    table_data = flatten_data(data)
    
    if table_data:
        table = Table(table_data, colWidths=[3*inch, 3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]))
        story.append(table)