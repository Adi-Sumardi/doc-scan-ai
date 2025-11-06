"""
Generate demo Excel files for PPN Reconciliation testing
Creates 3 files:
1. Demo_Faktur_Pajak.xlsx (40 rows - will be split into Point A & B)
2. Demo_Bukti_Potong.xlsx (40 rows - Point C with high match rate)
3. Demo_Rekening_Koran.xlsx (40 rows - Point E with high match rate)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

# Store generated data for matching purposes
faktur_data = {
    'point_a': [],  # Company as seller
    'point_b': []   # Company as buyer
}

def create_styled_header(ws, headers, header_fill, header_font, border):
    """Create styled header row"""
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Set row height
    ws.row_dimensions[1].height = 30

def generate_faktur_pajak():
    """Generate Faktur Pajak demo file with buyer and seller info"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faktur Pajak"

    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers (11+ columns as mentioned in the frontend)
    headers = [
        "Nomor Faktur",
        "Tanggal Faktur",
        "Masa Pajak",
        "Tahun Pajak",
        "Status",
        "NPWP Penjual",
        "Nama Penjual",
        "NPWP Pembeli",
        "Nama Pembeli",
        "DPP (Rp)",
        "PPN (Rp)",
        "Total (Rp)"
    ]

    create_styled_header(ws, headers, header_fill, header_font, border)

    # Company NPWP (this will be used to split Point A and Point B)
    company_npwp = "01.234.567.8-901.000"

    # Generate 40 rows of data
    start_date = datetime(2024, 1, 1)

    for i in range(40):
        # Random date in January-March 2024
        days_offset = random.randint(0, 89)
        tanggal = start_date + timedelta(days=days_offset)

        # Determine if this is Point A (company is seller) or Point B (company is buyer)
        is_point_a = i % 2 == 0  # Alternating - 20 Point A, 20 Point B

        # Generate DPP and calculate PPN (11%)
        dpp = random.randint(5000000, 50000000)
        ppn = int(dpp * 0.11)
        total = dpp + ppn

        if is_point_a:
            # Point A: Company is seller
            npwp_penjual = company_npwp
            nama_penjual = "PT DEMO COMPANY"
            npwp_pembeli = f"02.{100+i:03d}.{200+i:03d}.{i+1}-{300+i:03d}.000"
            nama_pembeli = f"PT BUYER {i+1}"

            # Store Point A data for Bukti Potong matching
            faktur_data['point_a'].append({
                'npwp_pembeli': npwp_pembeli,
                'nama_pembeli': nama_pembeli,
                'tanggal': tanggal,
                'dpp': dpp,
                'total': total
            })
        else:
            # Point B: Company is buyer
            npwp_penjual = f"03.{100+i:03d}.{200+i:03d}.{i+1}-{300+i:03d}.000"
            nama_penjual = f"PT VENDOR {i+1}"
            npwp_pembeli = company_npwp
            nama_pembeli = "PT DEMO COMPANY"

            # Store Point B data for Rekening Koran matching
            faktur_data['point_b'].append({
                'npwp_penjual': npwp_penjual,
                'nama_penjual': nama_penjual,
                'tanggal': tanggal,
                'dpp': dpp,
                'total': total
            })

        row_data = [
            f"010.000-{24:02d}.{i+1:08d}",  # Nomor Faktur
            tanggal.strftime("%d/%m/%Y"),   # Tanggal Faktur
            tanggal.strftime("%m"),          # Masa Pajak
            "2024",                          # Tahun Pajak
            "Normal",                        # Status
            npwp_penjual,                    # NPWP Penjual
            nama_penjual,                    # Nama Penjual
            npwp_pembeli,                    # NPWP Pembeli
            nama_pembeli,                    # Nama Pembeli
            dpp,                             # DPP
            ppn,                             # PPN
            total                            # Total
        ]

        ws.append(row_data)

        # Format currency cells
        row_num = ws.max_row
        for col_num in [10, 11, 12]:  # DPP, PPN, Total columns
            cell = ws.cell(row=row_num, column=col_num)
            cell.number_format = '#,##0'
            cell.border = border

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save("Demo_Faktur_Pajak.xlsx")
    print("✓ Generated Demo_Faktur_Pajak.xlsx (40 rows: 20 Point A, 20 Point B)")

def generate_bukti_potong():
    """Generate Bukti Potong PPh23 demo file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bukti Potong PPh23"

    # Styling
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Nomor Bukti Potong",
        "Tanggal Bukti Potong",
        "NPWP Pemotong",
        "Nama Pemotong",
        "NPWP Yang Dipotong",
        "Nama Yang Dipotong",
        "Jenis Penghasilan",
        "Jumlah Penghasilan Bruto (Rp)",
        "Tarif (%)",
        "PPh Dipotong (Rp)"
    ]

    create_styled_header(ws, headers, header_fill, header_font, border)

    company_npwp = "01.234.567.8-901.000"

    # Jenis penghasilan yang umum untuk PPh 23
    jenis_penghasilan = [
        "Jasa Teknik",
        "Jasa Manajemen",
        "Jasa Konsultan",
        "Sewa Gedung",
        "Jasa Konstruksi"
    ]

    # Generate 40 rows: 30 that match Point A, 10 that don't match
    for i in range(40):
        # 75% match rate (30 out of 40)
        if i < 30:
            # Match with Point A data
            point_a_index = i % len(faktur_data['point_a'])
            point_a = faktur_data['point_a'][point_a_index]

            npwp_pemotong = point_a['npwp_pembeli']
            nama_pemotong = point_a['nama_pembeli']
            tanggal = point_a['tanggal']
            bruto = point_a['dpp']  # Use same DPP as Point A
        else:
            # Random NPWP that won't match
            tanggal = datetime(2024, 1, 1) + timedelta(days=random.randint(0, 89))
            npwp_pemotong = f"04.{random.randint(100, 999):03d}.{random.randint(100, 999):03d}.{random.randint(1, 9)}-{random.randint(100, 999):03d}.000"
            nama_pemotong = f"PT CLIENT {i+1}"
            bruto = random.randint(5000000, 50000000)

        tarif = 2  # PPh 23 umumnya 2%
        pph = int(bruto * tarif / 100)

        row_data = [
            f"BP-{i+1:05d}/PPh23/{tanggal.strftime('%m')}/2024",
            tanggal.strftime("%d/%m/%Y"),
            npwp_pemotong,
            nama_pemotong,
            company_npwp,
            "PT DEMO COMPANY",
            random.choice(jenis_penghasilan),
            bruto,
            tarif,
            pph
        ]

        ws.append(row_data)

        # Format cells
        row_num = ws.max_row
        for col_num in [8, 10]:  # Bruto and PPh columns
            cell = ws.cell(row=row_num, column=col_num)
            cell.number_format = '#,##0'
            cell.border = border

        # Format percentage
        cell = ws.cell(row=row_num, column=9)
        cell.number_format = '0%'
        cell.border = border

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save("Demo_Bukti_Potong.xlsx")
    print("✓ Generated Demo_Bukti_Potong.xlsx (40 rows: 30 match Point A, 10 unmatched)")

def generate_rekening_koran():
    """Generate Rekening Koran (Bank Statement) demo file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekening Koran"

    # Styling
    header_fill = PatternFill(start_color="9966FF", end_color="9966FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Tanggal",
        "Keterangan",
        "Cabang",
        "Debet (Rp)",
        "Kredit (Rp)",
        "Saldo (Rp)",
        "Referensi"
    ]

    create_styled_header(ws, headers, header_fill, header_font, border)

    saldo = 100000000  # Starting balance

    # Generate 40 rows: 25 match Point B (payment out), 15 random
    for i in range(40):
        # 62.5% match rate (25 out of 40)
        if i < 25:
            # Match with Point B data (payments out to vendors)
            point_b_index = i % len(faktur_data['point_b'])
            point_b = faktur_data['point_b'][point_b_index]

            tanggal = point_b['tanggal']
            # Payment amount = Total from faktur (DPP + PPN)
            jumlah = point_b['total']
            debet = jumlah
            kredit = 0
            saldo -= jumlah
            keterangan = f"TRANSFER KE {point_b['nama_penjual']}"
        else:
            # Random transactions that won't match
            tanggal = datetime(2024, 1, 1) + timedelta(days=random.randint(0, 89))
            jumlah = random.randint(5000000, 50000000)

            # Mix of debit and credit for unmatched
            if i % 2 == 0:
                debet = jumlah
                kredit = 0
                saldo -= jumlah
                keterangan = f"TRANSFER KE PT SUPPLIER {i+1}"
            else:
                debet = 0
                kredit = jumlah
                saldo += jumlah
                keterangan = f"TERIMA TRANSFER DARI PT CLIENT {i+1}"

        row_data = [
            tanggal.strftime("%d/%m/%Y"),
            keterangan,
            "Jakarta",
            debet,
            kredit,
            saldo,
            f"REF{i+1:06d}"
        ]

        ws.append(row_data)

        # Format currency cells
        row_num = ws.max_row
        for col_num in [4, 5, 6]:  # Debet, Kredit, Saldo
            cell = ws.cell(row=row_num, column=col_num)
            cell.number_format = '#,##0'
            cell.border = border

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save("Demo_Rekening_Koran.xlsx")
    print("✓ Generated Demo_Rekening_Koran.xlsx (40 rows: 25 match Point B, 15 unmatched)")

if __name__ == "__main__":
    print("\n🚀 Generating demo Excel files for PPN Reconciliation...\n")
    generate_faktur_pajak()
    generate_bukti_potong()
    generate_rekening_koran()
    print("\n✅ All demo files generated successfully!")
    print("\nFiles created:")
    print("  1. Demo_Faktur_Pajak.xlsx - 40 rows (20 Point A, 20 Point B)")
    print("  2. Demo_Bukti_Potong.xlsx - 40 rows (30 match Point A, 10 unmatched)")
    print("  3. Demo_Rekening_Koran.xlsx - 40 rows (25 match Point B, 15 unmatched)")
    print("\nExpected match rates:")
    print("  - Point A vs C: ~75% (15/20 Point A should match)")
    print("  - Point B vs E: ~62.5% (12-13/20 Point B should match)")
    print("\nCompany NPWP: 01.234.567.8-901.000")
    print("\nYou can now upload these files in the PPN Reconciliation page.")
